#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build Date_MASTER-C02.xlsx (C02 CONTROL - CORESPONDENTA CU REALITATEA).
Implementator cap-coada conform BLUEPRINT-C02-CONTROL + SPEC-DATE-MASTER-C02 + D1/D2/D3.

Construieste: realitatea curata + 5 anomalii LOCK injectate controlat + coloane si
formule de semnalizare + foaia _SEMNALIZARE + reconciliere injectat==detectat +
conservare suma valoare_neta (mostenita). NU repara / NU modifica valori sursa /
NU sterge / NU previi / NU DV cascada / NU C03.
"""
import random, datetime as dt
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

random.seed(42)
SRC = 'c02/Date_MASTER-C02.xlsx'
OUT = 'c02/Date_MASTER-C02.xlsx'

# ============================================================ D1 LOCK (surse de adevar)
LOCALITATI = [
    ('Alba Iulia','Alba'),('Pitești','Argeș'),('Bacău','Bacău'),('Oradea','Bihor'),
    ('Brașov','Brașov'),('Făgăraș','Brașov'),('Cluj-Napoca','Cluj'),('Turda','Cluj'),
    ('Constanța','Constanța'),('Mangalia','Constanța'),('Craiova','Dolj'),('Galați','Galați'),
    ('Iași','Iași'),('Târgu Mureș','Mureș'),('Ploiești','Prahova'),('Câmpina','Prahova'),
    ('Sibiu','Sibiu'),('Mediaș','Sibiu'),('Timișoara','Timiș'),('București','București'),
]
ORAS2JUDET = {o: j for o, j in LOCALITATI}
REGULI_TVA = [('Hardware IT',0.19),('Licențe software',0.19),('Servicii consultanță',0.19),
    ('Echipamente birou',0.19),('Alimente',0.09),('Produse farmaceutice',0.09),('Cărți și manuale',0.05)]
CAT2COTA = {c: k for c, k in REGULI_TVA}
SARBATORI = [
    (dt.date(2026,1,1),'Anul Nou'),(dt.date(2026,1,2),'Anul Nou'),(dt.date(2026,1,6),'Boboteaza'),
    (dt.date(2026,1,7),'Sfântul Ioan'),(dt.date(2026,1,24),'Unirea Principatelor Române'),
    (dt.date(2026,4,10),'Vinerea Mare'),(dt.date(2026,4,12),'Paștele'),(dt.date(2026,4,13),'A doua zi de Paște'),
    (dt.date(2026,5,1),'Ziua Muncii'),(dt.date(2026,5,31),'Rusalii'),
    (dt.date(2026,6,1),'A doua zi de Rusalii / Ziua Copilului'),(dt.date(2026,8,15),'Adormirea Maicii Domnului'),
    (dt.date(2026,11,30),'Sfântul Andrei'),(dt.date(2026,12,1),'Ziua Națională'),
    (dt.date(2026,12,25),'Crăciunul'),(dt.date(2026,12,26),'A doua zi de Crăciun')]
SARB_SET = {d for d, _ in SARBATORI}
COD_JUDET = [('01','Alba'),('03','Argeș'),('04','Bacău'),('05','Bihor'),('08','Brașov'),
    ('12','Cluj'),('13','Constanța'),('16','Dolj'),('17','Galați'),('22','Iași'),
    ('26','Mureș'),('29','Prahova'),('32','Sibiu'),('35','Timiș'),('40','București')]
JUDET2COD = {j: c for c, j in COD_JUDET}
COD2JUDET = {c: j for c, j in COD_JUDET}
PRODUSE_CAT = {
    'Hardware IT': ['Laptop business 15"','Statie de lucru desktop','Monitor 27" 4K','Server rack 1U','Switch gigabit 24p'],
    'Licențe software': ['Licenta Office 365 anuala','Licenta antivirus enterprise','Licenta ERP contabilitate','Abonament suita cloud'],
    'Servicii consultanță': ['Consultanta implementare IT','Audit infrastructura','Suport tehnic anual','Training utilizatori'],
    'Echipamente birou': ['Multifunctionala laser','Tocator documente','Videoproiector sala','Telefon IP birou'],
    'Alimente': ['Cafea boabe 1kg','Apa plata bax','Pachet protocol sedinte','Ceai cutie 100buc'],
    'Produse farmaceutice': ['Trusa prim ajutor','Dezinfectant maini 5L','Masti protectie cutie','Termometru digital'],
    'Cărți și manuale': ['Manual proceduri interne','Carte tehnica IT','Ghid legislatie GDPR','Set carti formare']}
CAT_POOL = []
for c, w in [('Hardware IT',30),('Licențe software',18),('Echipamente birou',18),
             ('Servicii consultanță',12),('Alimente',9),('Produse farmaceutice',8),('Cărți și manuale',5)]:
    CAT_POOL += [c] * w
NUME_PERS = ['Andrei Popescu','Maria Ionescu','Vlad Dumitrescu','Elena Stan','Mihai Radu',
    'Ana Gheorghe','Cristian Marin','Ioana Tudor','George Stoica','Diana Florea','Razvan Barbu',
    'Carmen Nistor','Adrian Pop','Laura Constantin','Bogdan Dinu','Simona Lazar','Florin Matei',
    'Roxana Voicu','Catalin Serban','Gabriela Ene']

# ============================================================ citeste sursa (mosteneste valoare_neta)
wb = openpyxl.load_workbook(SRC)
ws_old = wb['Vanzari']
src_rows = list(ws_old.iter_rows(min_row=2, values_only=True))
N = len(src_rows)
clienti = sorted({r[2] for r in src_rows})
orase = [o for o, _ in LOCALITATI]
client_home = {cl: (orase[i % len(orase)], ORAS2JUDET[orase[i % len(orase)]]) for i, cl in enumerate(clienti)}

zile_lucr = []
d = dt.date(2026,1,1)
while d <= dt.date(2026,12,31):
    if d.weekday() < 5 and d not in SARB_SET:
        zile_lucr.append(d)
    d += dt.timedelta(days=1)
duminici = [d for d in (dt.date(2026,1,1)+dt.timedelta(n) for n in range(365)) if d.weekday()==6]

# ============================================================ realitatea curata
recs = []
for r in src_rows:
    client = r[2]; cat = random.choice(CAT_POOL); cota = CAT2COTA[cat]
    oras, judet = client_home[client]
    fact = random.choice(zile_lucr); scad = fact + dt.timedelta(days=random.choice([15,30,30,30,45,60]))
    neta = r[9]; tva = round(neta*cota,2)
    recs.append(dict(nr_factura=r[0], data_factura=fact, client_nume=client, judet=judet, oras=oras,
        cod_produs=r[4], produs_nume=random.choice(PRODUSE_CAT[cat]), categorie=cat,
        cantitate=r[7], pret_unitar=r[8], valoare_neta=neta, cota_tva=cota,
        valoare_tva=tva, valoare_totala=round(neta+tva,2), moneda=r[13] or 'RON', data_scadentei=scad))
suma_sursa = round(sum(x['valoare_neta'] for x in recs), 2)

# ============================================================ D4 injectie anomalii (registru ground-truth)
idx = list(range(N)); random.shuffle(idx)
inj = {'oras': set(), 'tva': set(), 'scad': set(), 'zi': set()}
p = 0
oras_idx = idx[p:p+50]; p += 50
tva_idx  = idx[p:p+35]; p += 35
scad_idx = idx[p:p+25]; p += 25
zi_idx   = idx[p:p+20]; p += 20

scriere_pool = ['Cluj','CJ','Cluj Napoca','Clum Napoca','Ploiesti Centru','Sibiu Oras',
                'Craioba','Iassy','Bv','Tm','Galati Port','Pitesci']
for k, ri in enumerate(oras_idx):
    if k < 30:
        recs[ri]['oras'] = random.choice(scriere_pool)        # nu exista in nomenclator (vizibil uman)
    else:
        wrong = random.choice([j for _, j in LOCALITATI if j != recs[ri]['judet']])
        recs[ri]['judet'] = wrong                              # oras oficial intr-un judet imposibil
    inj['oras'].add(ri)

for ri in tva_idx:
    legal = CAT2COTA[recs[ri]['categorie']]
    wrong = random.choice([c for c in (0.19,0.09,0.05) if c != legal])
    recs[ri]['cota_tva'] = wrong
    recs[ri]['valoare_tva'] = round(recs[ri]['valoare_neta']*wrong,2)
    recs[ri]['valoare_totala'] = round(recs[ri]['valoare_neta']+recs[ri]['valoare_tva'],2)
    inj['tva'].add(ri)

for ri in scad_idx:
    recs[ri]['data_scadentei'] = recs[ri]['data_factura'] - dt.timedelta(days=random.choice([3,5,7,10,15]))
    inj['scad'].add(ri)

for k, ri in enumerate(zi_idx):
    recs[ri]['data_factura'] = random.choice(list(SARB_SET)) if k % 2 == 0 else random.choice(duminici)
    recs[ri]['data_scadentei'] = recs[ri]['data_factura'] + dt.timedelta(days=30)
    inj['zi'].add(ri)

# multi-anomalii: 12 randuri din setul orase primesc si o anomalie TVA (boala dubla)
for ri in list(inj['oras'])[:12]:
    legal = CAT2COTA[recs[ri]['categorie']]
    recs[ri]['cota_tva'] = 0.19 if legal != 0.19 else 0.09
    recs[ri]['valoare_tva'] = round(recs[ri]['valoare_neta']*recs[ri]['cota_tva'],2)
    recs[ri]['valoare_totala'] = round(recs[ri]['valoare_neta']+recs[ri]['valoare_tva'],2)
    inj['tva'].add(ri)

suma_dupa = round(sum(x['valoare_neta'] for x in recs), 2)  # neatins de injectie

# ============================================================ CONTACTE (CNP coerent + anomalii)
def cnp_control(d12):
    w = [2,7,9,1,4,6,3,5,8,2,7,9]
    s = sum(int(d12[i])*w[i] for i in range(12)) % 11
    return 1 if s == 10 else s

def build_cnp(sex, bd, judet, seq):
    century = 5 if bd.year >= 2000 else 1
    s = century + (0 if sex == 'M' else 1)   # 1=M/2=F 1900s ; 5=M/6=F 2000s
    cod = JUDET2COD[judet]
    d12 = f"{s}{bd.year%100:02d}{bd.month:02d}{bd.day:02d}{cod}{seq:03d}"
    return d12 + str(cnp_control(d12))

contacts = []
inj_cnp = {'sex': set(), 'judet': set(), 'data': set(), 'control': set()}
for i in range(50):
    client = clienti[i % len(clienti)]
    judet = client_home[client][1]
    sex = random.choice(['M','F'])
    bd = dt.date(random.randint(1962,2001), random.randint(1,12), random.randint(1,28))
    cnp = build_cnp(sex, bd, judet, random.randint(1,899))
    contacts.append(dict(id_contact=f"CTC-{i+1:04d}", client_nume=client,
        nume_persoana=NUME_PERS[i % len(NUME_PERS)], CNP=cnp, Sex=sex, Judet=judet, Data_Nasterii=bd))

cnp_idx = list(range(50)); random.shuffle(cnp_idx)
for k, ci in enumerate(cnp_idx[:25]):
    c = contacts[ci]
    sub = ['sex','judet','data','control'][k % 4]
    if sub == 'sex':
        c['Sex'] = 'F' if c['Sex'] == 'M' else 'M'                       # contrazice CNP
        inj_cnp['sex'].add(ci)
    elif sub == 'judet':
        c['Judet'] = random.choice([j for _, j in LOCALITATI if j != c['Judet']])
        inj_cnp['judet'].add(ci)
    elif sub == 'data':
        c['Data_Nasterii'] = c['Data_Nasterii'] + dt.timedelta(days=400) # nu mai corespunde AALLZZ
        inj_cnp['data'].add(ci)
    else:
        wrong = (int(c['CNP'][12]) + 1) % 10
        c['CNP'] = c['CNP'][:12] + str(wrong)                            # cifra control gresita
        inj_cnp['control'].add(ci)
# cateva multi pe CNP: 5 contacte primesc si o a doua contradictie (judet)
for ci in list(inj_cnp['sex'])[:5]:
    contacts[ci]['Judet'] = random.choice([j for _, j in LOCALITATI if j != contacts[ci]['Judet']])
    inj_cnp['judet'].add(ci)

# ============================================================ DETECTIE INDEPENDENTA (reconciliere)
def det_oras(r):
    if r['oras'] not in ORAS2JUDET: return 'ORAS NEALINIAT'
    if ORAS2JUDET[r['oras']] != r['judet']: return 'JUDET NECORESPUNZATOR'
    return ''
def det_tva(r):
    return 'TVA GRESIT' if CAT2COTA[r['categorie']] != r['cota_tva'] else ''
def det_scad(r):
    return 'SCADENTA INAINTEA FACTURII' if r['data_scadentei'] < r['data_factura'] else ''
def det_zi(r):
    return 'ZI NELUCRATOARE' if (r['data_factura'].weekday() >= 5 or r['data_factura'] in SARB_SET) else ''

det = {'oras': set(), 'tva': set(), 'scad': set(), 'zi': set()}
for i, r in enumerate(recs):
    if det_oras(r): det['oras'].add(i)
    if det_tva(r): det['tva'].add(i)
    if det_scad(r): det['scad'].add(i)
    if det_zi(r): det['zi'].add(i)

def det_cnp(c):
    cnp = c['CNP']; flags = []
    fd = int(cnp[0]); dsex = 'M' if fd in (1,3,5,7) else 'F'
    if dsex != c['Sex']: flags.append('sex')
    if COD2JUDET.get(cnp[7:9]) != c['Judet']: flags.append('judet')
    base = 2000 if fd >= 5 else 1900
    try:
        dd = dt.date(base+int(cnp[1:3]), int(cnp[3:5]), int(cnp[5:7]))
        if dd != c['Data_Nasterii']: flags.append('data')
    except ValueError:
        flags.append('data')
    if cnp_control(cnp[:12]) != int(cnp[12]): flags.append('control')
    return flags

det_cnp_sets = {'sex': set(), 'judet': set(), 'data': set(), 'control': set()}
for i, c in enumerate(contacts):
    for f in det_cnp(c):
        det_cnp_sets[f].add(i)

# ============================================================ SCRIE WORKBOOK
wb.remove(ws_old)
ws = wb.create_sheet('Vanzari', 0)
COLS = ['nr_factura','data_factura','client_nume','judet','cod_produs','produs_nume','categorie',
        'cantitate','pret_unitar','valoare_neta','cota_tva','valoare_tva','valoare_totala','moneda',
        'oras','data_scadentei','flag_oras','flag_tva','flag_scadenta','flag_zi','status_validare','motiv_anomalie']
ws.append(COLS)
for ri, r in enumerate(recs, start=2):
    ws.append([r['nr_factura'], r['data_factura'], r['client_nume'], r['judet'], r['cod_produs'],
        r['produs_nume'], r['categorie'], r['cantitate'], r['pret_unitar'], r['valoare_neta'],
        r['cota_tva'], r['valoare_tva'], r['valoare_totala'], r['moneda'], r['oras'], r['data_scadentei'],
        f'=IF(COUNTIF(tbl_Localitati[oras_oficial],[@oras])=0,"ORAS NEALINIAT",IF(XLOOKUP([@oras],tbl_Localitati[oras_oficial],tbl_Localitati[judet],"?")<>[@judet],"JUDET NECORESPUNZATOR",""))',
        '=IF(XLOOKUP([@categorie],tbl_RegulriTVA[categorie],tbl_RegulriTVA[cota_legala],-1)<>[@cota_tva],"TVA GRESIT","")',
        '=IF([@data_scadentei]<[@data_factura],"SCADENTA INAINTEA FACTURII","")',
        '=IF(NETWORKDAYS.INTL([@data_factura],[@data_factura],1,tbl_SarbatoriLegale[data])=0,"ZI NELUCRATOARE","")',
        '=IF([@motiv_anomalie]="","OK","DE INVESTIGAT")',
        '=TEXTJOIN("; ",TRUE,[@flag_oras],[@flag_tva],[@flag_scadenta],[@flag_zi])'])
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2): row[0].number_format = 'yyyy-mm-dd'
for row in ws.iter_rows(min_row=2, min_col=16, max_col=16): row[0].number_format = 'yyyy-mm-dd'
for row in ws.iter_rows(min_row=2, min_col=11, max_col=11): row[0].number_format = '0%'
tv = Table(displayName='tbl_Vanzari', ref=f"A1:{get_column_letter(len(COLS))}{N+1}")
tv.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
ws.add_table(tv)

# CONTACTE
wsc = wb.create_sheet('CONTACTE')
CC = ['id_contact','client_nume','nume_persoana','CNP','Sex','Judet','Data_Nasterii',
      'flag_sex','flag_judet','flag_data','flag_control','status_validare','motiv_anomalie']
wsc.append(CC)
for c in contacts:
    wsc.append([c['id_contact'], c['client_nume'], c['nume_persoana'], c['CNP'], c['Sex'], c['Judet'], c['Data_Nasterii'],
        '=IF(IF(OR(VALUE(LEFT([@CNP],1))=1,VALUE(LEFT([@CNP],1))=3,VALUE(LEFT([@CNP],1))=5,VALUE(LEFT([@CNP],1))=7),"M","F")<>[@Sex],"SEX CONTRADICTORIU","")',
        '=IF(XLOOKUP(MID([@CNP],8,2),tbl_CodJudetCNP[cod],tbl_CodJudetCNP[judet],"?")<>[@Judet],"JUDET CONTRADICTORIU","")',
        '=IF(DATE(IF(VALUE(LEFT([@CNP],1))>=5,2000,1900)+VALUE(MID([@CNP],2,2)),VALUE(MID([@CNP],4,2)),VALUE(MID([@CNP],6,2)))<>[@Data_Nasterii],"DATA CONTRADICTORIE","")',
        '=IF(IF(MOD(SUMPRODUCT(VALUE(MID([@CNP],{1;2;3;4;5;6;7;8;9;10;11;12},1)),{2;7;9;1;4;6;3;5;8;2;7;9}),11)=10,1,MOD(SUMPRODUCT(VALUE(MID([@CNP],{1;2;3;4;5;6;7;8;9;10;11;12},1)),{2;7;9;1;4;6;3;5;8;2;7;9}),11))<>VALUE(MID([@CNP],13,1)),"CIFRA CONTROL GRESITA","")',
        '=IF([@motiv_anomalie]="","OK","DE INVESTIGAT")',
        '=TEXTJOIN("; ",TRUE,[@flag_sex],[@flag_judet],[@flag_data],[@flag_control])'])
for row in wsc.iter_rows(min_row=2, min_col=4, max_col=4): row[0].number_format = '@'
for row in wsc.iter_rows(min_row=2, min_col=7, max_col=7): row[0].number_format = 'yyyy-mm-dd'
tc = Table(displayName='tbl_Contacte', ref=f"A1:{get_column_letter(len(CC))}{len(contacts)+1}")
tc.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
wsc.add_table(tc)

# nomenclatoare (tabele)
def add_nom(name, headers, rows, tname):
    w = wb.create_sheet(name); w.append(headers)
    for rr in rows: w.append(list(rr))
    t = Table(displayName=tname, ref=f"A1:{get_column_letter(len(headers))}{len(rows)+1}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight11", showRowStripes=True)
    w.add_table(t)
    return w
add_nom('tbl_Localitati', ['oras_oficial','judet'], LOCALITATI, 'tbl_Localitati')
add_nom('tbl_RegulriTVA', ['categorie','cota_legala'], REGULI_TVA, 'tbl_RegulriTVA')
wsar = add_nom('tbl_SarbatoriLegale', ['data','denumire'], SARBATORI, 'tbl_SarbatoriLegale')
for row in wsar.iter_rows(min_row=2, min_col=1, max_col=1): row[0].number_format = 'yyyy-mm-dd'
wcod = add_nom('tbl_CodJudetCNP', ['cod','judet'], COD_JUDET, 'tbl_CodJudetCNP')
for row in wcod.iter_rows(min_row=2, min_col=1, max_col=1): row[0].number_format = '@'
wb['tbl_RegulriTVA']
for row in wb['tbl_RegulriTVA'].iter_rows(min_row=2, min_col=2, max_col=2): row[0].number_format = '0%'

# _SEMNALIZARE (sinteza + COUNTIF live)
wss = wb.create_sheet('_SEMNALIZARE')
bold = Font(bold=True)
wss['A1'] = 'C02 CONTROL - SINTEZA SEMNALIZARII (anomalii de realitate)'; wss['A1'].font = Font(bold=True, size=13)
wss['A3'] = 'Fenomen'; wss['B3'] = 'Plantate'; wss['C3'] = 'Detectate (live)'
for c in ('A3','B3','C3'): wss[c].font = bold
rows_sem = [
    ('C02.01 Orase nealiniate', len(inj['oras']),
     '=COUNTIF(tbl_Vanzari[flag_oras],"ORAS NEALINIAT")+COUNTIF(tbl_Vanzari[flag_oras],"JUDET NECORESPUNZATOR")'),
    ('C02.02 TVA gresit', len(inj['tva']), '=COUNTIF(tbl_Vanzari[flag_tva],"TVA GRESIT")'),
    ('C02.03 Scadenta < Factura', len(inj['scad']), '=COUNTIF(tbl_Vanzari[flag_scadenta],"SCADENTA INAINTEA FACTURII")'),
    ('C02.04 Zi nelucratoare', len(inj['zi']), '=COUNTIF(tbl_Vanzari[flag_zi],"ZI NELUCRATOARE")'),
    ('C02.05 CNP contradictoriu (contacte)', len(set().union(*inj_cnp.values())),
     '=COUNTIF(tbl_Contacte[status_validare],"DE INVESTIGAT")'),
]
r = 4
for nume, pl, formula in rows_sem:
    wss.cell(r,1,nume); wss.cell(r,2,pl); wss.cell(r,3,formula); r += 1
wss.cell(r+1,1,'Randuri Vanzari de investigat (total)').font = bold
wss.cell(r+1,3,'=COUNTIF(tbl_Vanzari[status_validare],"DE INVESTIGAT")')
wss.cell(r+2,1,'Suma valoare_neta (neatinsa de semnalizare)')
wss.cell(r+2,3,'=SUM(tbl_Vanzari[valoare_neta])')
wss.cell(r+4,1,'Nota: semnalizarea MARCHEAZA, nu corecteaza. Valorile sursa raman neatinse (R-MET-1).')
wss.column_dimensions['A'].width = 42; wss.column_dimensions['C'].width = 60

# CF: marcheaza vizual randurile de investigat (semnalizare, nu corectie)
red = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
ws.conditional_formatting.add(f"A2:V{N+1}", FormulaRule(formula=['$U2="DE INVESTIGAT"'], fill=red))
wsc.conditional_formatting.add(f"A2:M{len(contacts)+1}", FormulaRule(formula=['$L2="DE INVESTIGAT"'], fill=red))

# reordoneaza: Vanzari, CONTACTE, nomenclatoare, _SEMNALIZARE, aux, _README, CONTROL_FINAL
order = ['Vanzari','CONTACTE','tbl_Localitati','tbl_RegulriTVA','tbl_SarbatoriLegale','tbl_CodJudetCNP',
         '_SEMNALIZARE','CLIENTI','PRODUSE','AGENTI','DEPOZITE','_README','CONTROL_FINAL']
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)

wb.save(OUT)

# ============================================================ RAPORT + RECONCILIERE
def line(): print('-'*64)
print('='*64); print('RAPORT BUILD Date_MASTER-C02 (C02 CONTROL)'); print('='*64)
print(f"Fisier: {OUT}")
print(f"Foi: {wb.sheetnames}")
line(); print('RECONCILIERE Vanzari (plantate == detectate):')
ok = True
for f, nume in [('oras','C02.01 Orase'),('tva','C02.02 TVA'),('scad','C02.03 Scadenta'),('zi','C02.04 NETWORKDAYS')]:
    same = inj[f] == det[f]
    ok = ok and same
    print(f"  {nume:22s} plantate={len(inj[f]):4d}  detectate={len(det[f]):4d}  {'MATCH' if same else 'MISMATCH'}")
line(); print('RECONCILIERE CONTACTE CNP (C02.05, plantate == detectate):')
for f in ['sex','judet','data','control']:
    same = inj_cnp[f] == det_cnp_sets[f]
    ok = ok and same
    print(f"  {f:10s} plantate={len(inj_cnp[f]):3d}  detectate={len(det_cnp_sets[f]):3d}  {'MATCH' if same else 'MISMATCH'}")
contacte_flag_inj = set().union(*inj_cnp.values())
contacte_flag_det = set().union(*det_cnp_sets.values())
print(f"  contacte cu anomalie: plantate={len(contacte_flag_inj)} detectate={len(contacte_flag_det)} "
      f"{'MATCH' if contacte_flag_inj==contacte_flag_det else 'MISMATCH'}")
line()
afectate = inj['oras']|inj['tva']|inj['scad']|inj['zi']
multi = sum(1 for i in range(N) if sum(1 for f in inj if i in inj[f])>=2)
print(f"Randuri Vanzari afectate (distinct): {len(afectate)} / {N}  ({100*len(afectate)/N:.1f}%)")
print(f"Randuri Vanzari cu anomalii multiple: {multi}")
print(f"SUMA valoare_neta sursa : {suma_sursa}")
print(f"SUMA valoare_neta dupa  : {suma_dupa}   {'CONSERVATA' if suma_sursa==suma_dupa else 'MODIFICATA!'}")
line()
print('REZULTAT GLOBAL:', 'TOATE MATCH - ZERO fals pozitiv/negativ' if ok else 'ESEC RECONCILIERE')
