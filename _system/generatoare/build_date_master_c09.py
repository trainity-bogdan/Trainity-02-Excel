#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c09.py - construieste Date_MASTER-C09.xlsx (T3 · RELATII).

C09 NU cloneaza C08. Preia ACELEASI date (continuitate R-V02.14, suma conservata)
si ACTIVEAZA relatiile pe care C08 doar le-a recunoscut descriptiv:
  - fact Vanzari (neatins, 15 col canonice, suma valoare_neta identica cu C08)
  - dimensiuni reale: PRODUSE (cheie cod_produs, deja in fact), CLIENTI (cheia
    cod_client ACTIVATA prin rezolvarea nume->cod), REGIUNI (judet->regiune macro,
    cheie noua), CALENDAR (data->an/luna/trimestru, cheie noua)
  - _MODEL: documentarea relatiilor 1:M + cardinalitati + status
  - _INTEGRITATE: zero orfani, cardinalitati, suma conservata
  - _CITIRE_DEMO: citire demonstrativa cross-tabel (FARA masuri numite / clasament)

INTERDICTII (garda C09, nu C10/C11/C12/T4/T5): zero masuri DAX numite, zero KPI
reutilizabil, zero ranking/top/bottom/Pareto, zero dashboard, zero cauza, zero
recomandare. Doar prima citire cross-tabel demonstrativa.

Uz: python3 _system/generatoare/build_date_master_c09.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

SRC = 'c08/Date_MASTER-C08.xlsx'
OUT = 'c09/Date_MASTER-C09.xlsx'

# regiune macro per judet (cheie noua REGIUNI; Vanzari are doar judet, nu regiune)
JUDET_REGIUNE = {
    'Brașov': 'Transilvania', 'Cluj': 'Transilvania', 'Mureș': 'Transilvania',
    'Sibiu': 'Transilvania', 'Timiș': 'Banat', 'București': 'Muntenia',
    'Constanța': 'Dobrogea', 'Iași': 'Moldova',
}
LUNI = ['ianuarie', 'februarie', 'martie', 'aprilie', 'mai', 'iunie', 'iulie',
        'august', 'septembrie', 'octombrie', 'noiembrie', 'decembrie']
ZILE = ['luni', 'marti', 'miercuri', 'joi', 'vineri', 'sambata', 'duminica']

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(bold=True, size=12, color='1F2A44')


def hdr(ws, row=1):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL; c.font = HDR_FONT


def main():
    os.makedirs('c09', exist_ok=True)
    src = openpyxl.load_workbook(SRC, data_only=True)
    V = src['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [list(r) for r in V.iter_rows(min_row=2, values_only=True)]
    nrows = len(vrows)
    last = nrows + 1  # ultimul rand cu date in foaia Vanzari

    out = openpyxl.Workbook()
    out.remove(out.active)

    # 1) FACT Vanzari (neatins, canonical)
    ws = out.create_sheet('Vanzari')
    ws.append(vhdr)
    for r in vrows:
        ws.append(r)
    hdr(ws)

    # 2) DIM PRODUSE, CLIENTI, AGENTI, DEPOZITE (preluate ca dimensiuni)
    for name in ['PRODUSE', 'CLIENTI', 'AGENTI', 'DEPOZITE']:
        s = src[name]
        w = out.create_sheet(name)
        for row in s.iter_rows(values_only=True):
            w.append(list(row))
        hdr(w)

    # bridge: nume_client -> cod_client (din dim CLIENTI)
    cl = {r[1]: r[0] for r in src['CLIENTI'].iter_rows(min_row=2, values_only=True)}

    # 3) DIM REGIUNI (cheie noua judet->regiune; activeaza relatia geografica)
    wr = out.create_sheet('REGIUNI')
    wr.append(['judet', 'regiune'])
    judete = sorted({r[ix['judet']] for r in vrows})
    for j in judete:
        wr.append([j, JUDET_REGIUNE.get(j, 'Necunoscuta')])
    hdr(wr)

    # 4) DIM CALENDAR (cheie noua data->an/luna/trimestru/zi)
    wc = out.create_sheet('CALENDAR')
    wc.append(['data', 'an', 'luna_nr', 'luna_nume', 'trimestru', 'zi_saptamana'])
    dates = sorted({r[ix['data_factura']] for r in vrows if r[ix['data_factura']]})
    for d in dates:
        q = (d.month - 1) // 3 + 1
        wc.append([d, d.year, d.month, LUNI[d.month - 1], f'T{q}', ZILE[d.weekday()]])
    hdr(wc)

    # 5) _REL_CLIENTI: cheia ACTIVATA (C08 avea clientul doar pe nume, fara cod)
    wb = out.create_sheet('_REL_CLIENTI')
    wb.append(['ACTIVAREA RELATIEI CLIENTI (C09): nume_client -> cod_client'])
    wb['A1'].font = TITLE_FONT
    wb.append(['In C08 clientul exista doar ca text in fact. C09 il leaga prin cheie.'])
    wb.append([])
    wb.append(['nume_client', 'cod_client_rezolvat', 'verificare'])
    hdr(wb, row=4)
    cnames = sorted({r[ix['client_nume']] for r in vrows})
    for nm in cnames:
        # rezolvare live prin XLOOKUP in dimensiunea CLIENTI
        wb.append([nm,
                   f'=XLOOKUP(A{wb.max_row+1},CLIENTI!B:B,CLIENTI!A:A,"LIPSA")',
                   f'=IF(B{wb.max_row+1}="LIPSA","ORFAN","OK")'])

    # 6) _MODEL: documentarea relatiilor (foaia ceruta)
    wm = out.create_sheet('_MODEL')
    rowsm = [
        ['MODEL RELATIONAL C09 · star schema (fact Vanzari + dimensiuni)'],
        ['Axa: C09 leaga sursele intr-un model interogabil. Nu masoara, nu compara, nu explica.'],
        [],
        ['fact', 'dimensiune', 'cheie', 'cardinalitate', 'status', 'observatie'],
        ['Vanzari', 'PRODUSE', 'cod_produs', 'M:1', 'ACTIVA',
         'cheie deja prezenta in fact'],
        ['Vanzari', 'CLIENTI', 'cod_client (din nume_client)', 'M:1', 'ACTIVATA in C09',
         'C08 avea doar nume; C09 rezolva cheia (vezi _REL_CLIENTI)'],
        ['Vanzari', 'REGIUNI', 'judet', 'M:1', 'ACTIVA',
         'judet -> regiune macro; fact nu are coloana regiune'],
        ['Vanzari', 'CALENDAR', 'data_factura', 'M:1', 'ACTIVA',
         'data -> an/luna/trimestru'],
        ['Vanzari', 'AGENTI', '(lipsa cheie in fact)', '-', 'NELEGATA',
         'dimensiune prezenta, fara cheie in fact: nu poate raspunde inca'],
        ['Vanzari', 'DEPOZITE', '(lipsa cheie in fact)', '-', 'NELEGATA',
         'idem AGENTI; candidata de extindere ulterioara'],
        [],
        ['Regula C09: o dimensiune raspunde DOAR daca are cheie activa in fact.'],
        ['Fara cheie, dimensiunea exista dar modelul nu o poate interoga.'],
    ]
    for r in rowsm:
        wm.append(r)
    wm['A1'].font = TITLE_FONT
    hdr(wm, row=4)

    # 7) _INTEGRITATE: validitatea modelului (zona de verificare)
    suma = round(sum(float(r[ix['valoare_neta']]) for r in vrows
                     if r[ix['valoare_neta']] not in (None, '')), 2)
    n_prod = len({r[ix['cod_produs']] for r in vrows})
    n_cli = len(cnames)
    n_reg = len(set(JUDET_REGIUNE.get(j) for j in judete))
    n_cal = len(dates)
    wi = out.create_sheet('_INTEGRITATE')
    rowsi = [
        ['INTEGRITATEA MODELULUI C09 (relatii curate, zero orfani)'],
        [],
        ['verificare', 'valoare', 'asteptat', 'metoda'],
        ['orfani cod_produs (fact fara dim)', None, '0',
         '=SUMPRODUCT(--(ISNA(MATCH(Vanzari!E2:E%d,PRODUSE!A2:A%d,0))))' % (last, src['PRODUSE'].max_row)],
        ['orfani client (nume fara cod)', None, '0',
         '=SUMPRODUCT(--(ISNA(MATCH(Vanzari!C2:C%d,CLIENTI!B2:B%d,0))))' % (last, src['CLIENTI'].max_row)],
        ['orfani judet (fact fara regiune)', None, '0',
         '=SUMPRODUCT(--(ISNA(MATCH(Vanzari!D2:D%d,REGIUNI!A2:A%d,0))))' % (last, len(judete)+1)],
        ['cardinalitate PRODUSE', n_prod, '13', '=SUMPRODUCT(1/COUNTIF(Vanzari!E2:E%d,Vanzari!E2:E%d))' % (last, last)],
        ['cardinalitate CLIENTI', n_cli, '15', 'COUNTA distinct nume'],
        ['cardinalitate REGIUNI', n_reg, '5', 'regiuni macro'],
        ['cardinalitate CALENDAR (zile)', n_cal, '', 'zile distincte cu tranzactii'],
        ['suma valoare_neta CONSERVATA', suma, '7986019.38',
         '=ROUND(SUM(Vanzari!J2:J%d),2)' % last],
    ]
    for r in rowsi:
        wi.append(r)
    wi['A1'].font = TITLE_FONT
    hdr(wi, row=3)

    # 8) _CITIRE_DEMO: prima citire cross-tabel (NU masura numita, NU clasament)
    wd = out.create_sheet('_CITIRE_DEMO')
    transilvania = [j for j in judete if JUDET_REGIUNE.get(j) == 'Transilvania']
    rng_t = ';'.join('"%s"' % j for j in transilvania)
    demo_client = cnames[0]
    rowsd = [
        ['CITIRE DEMONSTRATIVA CROSS-TABEL (C09) · dovada ca modelul raspunde'],
        ['Nota: este o singura citire de demonstratie, NU o masura reutilizabila si NU un clasament.'],
        [],
        ['Intrebare 1: cata valoare pe regiunea "Transilvania"?'],
        ['Un singur tabel (Vanzari) NU poate raspunde: nu are coloana "regiune", doar "judet".'],
        ['Raspuns prin relatia Vanzari[judet] -> REGIUNI:',
         '=SUMPRODUCT((ISNUMBER(MATCH(Vanzari!D2:D%d,{%s},0)))*Vanzari!J2:J%d)' % (last, rng_t, last)],
        [],
        ['Intrebare 2: cate tranzactii are clientul "%s"?' % demo_client],
        ['Raspuns prin relatia Vanzari[client_nume] -> CLIENTI (cheie activata):',
         '=COUNTIF(Vanzari!C2:C%d,"%s")' % (last, demo_client)],
        [],
        ['Concluzie: fara relatii, intrebarile cer un singur tabel si raman fara raspuns.'],
        ['Cu relatii, o intrebare traverseaza tabelele si primeste un raspuns. (AHA C09)'],
    ]
    for r in rowsd:
        wd.append(r)
    wd['A1'].font = TITLE_FONT

    # 9) _README C09 (NU textul ecosistem C08)
    wrd = out.create_sheet('_README')
    for line in [
        'DATE_MASTER-C09.XLSX · T3 ANALIZA · C09 RELATII',
        '',
        'Rol: setul cunoscut (T2) devine un MODEL RELATIONAL INTEROGABIL.',
        'C08 a recunoscut ecosistemul descriptiv (cine cu cine s-ar putea lega).',
        'C09 ACTIVEAZA relatiile reale: chei + cardinalitate 1:M + model interogabil.',
        '',
        'Continut:',
        '  Vanzari        - tabel fact (tranzactii), neatins fata de C08',
        '  PRODUSE/CLIENTI/REGIUNI/CALENDAR - dimensiuni cu chei catre fact',
        '  AGENTI/DEPOZITE - dimensiuni prezente, inca NELEGATE (fara cheie in fact)',
        '  _REL_CLIENTI    - activarea cheii client (nume -> cod)',
        '  _MODEL          - relatiile 1:M documentate',
        '  _INTEGRITATE    - zero orfani, cardinalitati, suma conservata',
        '  _CITIRE_DEMO    - prima citire cross-tabel (demonstrativa)',
        '',
        'Garda C09: doar leaga si permite prima citire cross-tabel.',
        'NU defineste masuri numite (C10), NU compara/claseaza (C11),',
        'NU explica de ce (C12), NU produce dashboard (T4), NU actioneaza (T5).',
        '',
        'Handoff: input de la C08 (set cunoscut). Output catre C10 (masuri peste model).',
        'Suma valoare_neta conservata cap-coada: 7986019.38 (R-V02.14).',
    ]:
        wrd.append([line])
    wrd['A1'].font = TITLE_FONT

    # 10) CONTROL_FINAL
    wcf = out.create_sheet('CONTROL_FINAL')
    for r in [
        ['CONTROL FINAL C09 · conservare + handoff'],
        [],
        ['verificare', 'valoare'],
        ['suma valoare_neta', suma],
        ['suma asteptata (C08 in)', 7986019.38],
        ['delta', round(suma - 7986019.38, 2)],
        ['randuri fact', nrows],
        ['relatii active', 4],
        ['relatii nelegate (candidate)', 2],
        ['predat catre', 'C10 MASURI (defineste masuri peste model)'],
    ]:
        wcf.append(r)
    wcf['A1'].font = TITLE_FONT
    hdr(wcf, row=3)

    out.save(OUT)
    # verificare finala
    print('SCRIS:', OUT)
    print('  sheets:', out.sheetnames)
    print('  suma valoare_neta:', suma, '| delta vs C08:', round(suma - 7986019.38, 2))
    print('  cardinalitati: PRODUSE=%d CLIENTI=%d REGIUNI=%d CALENDAR=%d' %
          (n_prod, n_cli, n_reg, n_cal))
    assert abs(suma - 7986019.38) < 0.01, 'SUMA NU se conserva!'
    print('  CONSERVARE OK')


if __name__ == '__main__':
    main()
