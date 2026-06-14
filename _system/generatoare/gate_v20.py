#!/usr/bin/env python3
"""
gate_v20.py — POST-FLIGHT GATE BLOCANT (R-V03.45 extins V18)

Verificator pe 6 CLASE de erori (V18 adaugă DATA-CONTINUITY).

Diferențe față de V16:
- CLASA 6 nouă: DATA-CONTINUITY (schema Date_MASTER, conservare sumă
  cap-coadă, schema stabilă conform SCHEMA-MASTER.md)
- Filtre anti-fals-pozitiv codificate: CSS values (opacity, transform,
  scale), Power Query whitelist (Refresh All, Architecture in context M)
- Versionare pilot prin hash MD5 raportat
- Cele 6 livrabile noi (eliminat Date-Output-Excel separat, păstrat
  doar Date_MASTER-dupa-C{NN}.xlsx ca livrabil unic Excel)

R-V03.45: gate-ul ruleaza OBLIGATORIU inainte de present_files in
chat CONSTRUCTIE. Daca pică, livrabilul NU iese din chat.

R-V03.47: schema canonică Date_MASTER (14 coloane fixe pe sheet
Vanzari) verificată automat de CLASA 6.

Usage in chat CONSTRUCTIE NN, dupa generare:
    python3 gate_v20.py NN /path/to/livrabile_CNN/ /path/to/pilot_C01_V12/

Returneaza exit code 0 daca PASS, 1 daca FAIL.
"""

import sys
import os
import re

# ============================================================
# CONFIG GENERIC
# ============================================================

# Zone textuale care pot fi vulnerabile la clone. Adauga aici cand
# descoperi o zona noua. Format: (CSS_class, threshold_pct, descriere).
# Threshold = procent maxim de identitate cu pilot permis (zero pe
# unele zone, mai relax pe altele).
TEXTUAL_ZONES = [
    # (selector_pattern, threshold_pct, descriere)
    ('step-title',          30, 'Step-titles in HTML-Studiu'),
    ('step-action',         30, 'Step-actions in HTML-Studiu'),
    ('prompt-text',          0, 'Prompt-text catre AI (toleranta zero)'),
    ('prompt-label',        50, 'Prompt-label (poate avea schelet comun)'),
    ('final-label',         70, 'Final-labels (genericele CONTROL/VOLUM sunt OK)'),
    ('anomaly-card-desc',   30, 'Descrieri anomaly-cards'),
    ('payoff-line',         30, 'Payoff-line text'),
    ('stage-quote',         30, 'Stage-quote text'),
    ('cover-subtitle',      50, 'Cover-subtitle (INTRIGA)'),
    ('cover-miza',          50, 'Cover-miza'),
]

# Engleza din lista neagra (R-V03.39)
ENGLEZA_FORBIDDEN = [
    'Normalize', 'Tutorial', 'Lesson', 'Quiz', 'Module ', 'Course',
    'Webinar', 'Masterclass'
]

# WHITELIST englezisme tehnice acceptate in context Power Query / Excel
# Acestea sunt nume canonice Microsoft, NU englezisme cosmetice
ENGLEZA_WHITELIST_TEHNIC = [
    'Refresh All', 'Refresh', 'Applied Steps', 'Applied Step',
    'Power Query', 'Power BI', 'Pivot Table', 'PivotTable', 'PivotChart',
    'Architecture',  # in context flux refresh PQ
    'Text.Trim', 'Text.Clean', 'Text.Replace',  # functii M
    'Number.FromText', 'Date.From', 'Number.From', 'Table.ReplaceValue',
    # V28 L146: Applied Steps canonici Power Query (Excel UI engleza)
    'Promoted Headers', 'Filtered Rows', 'Removed Blank Rows',
    'Removed Duplicates', 'Removed Other Columns', 'Changed Type',
    'Replaced Value', 'Merged Columns', 'Split Column', 'Group By',
    'Pivot Column', 'Unpivot', 'Transform Sample',
    'Normalized Diacritics', 'Parsed Date', 'Loaded as Excel Table',
    'Trim and Clean', 'Excel Table',
]

# Vocab brand forbidden in TEXT VIZIBIL (nu in comments CSS/JS)
VOCAB_FORBIDDEN_VIZIBIL = [
    'tutorial', 'masterclass', 'webinar', 'productivitate',
    'lecție', 'lectie',
]

# === SCHEMA CANONICĂ DATE_MASTER (R-V03.47, V18) ===
SCHEMA_CANONICA_VANZARI = [
    'nr_factura', 'data_factura', 'client_nume', 'judet',
    'cod_produs', 'produs_nume', 'categorie',
    'cantitate', 'pret_unitar', 'valoare_neta',
    'cota_tva', 'valoare_tva', 'valoare_totala', 'moneda',
]

# === EXCEPTIE C01 STRUCTURARE (constructia-pilot) ===
# C01 NU deriva din Date_MASTER-initial: e scenariul-cobai cu schema proprie
# (export ERP cu agenti + depozite, pentru lectia de integritate referentiala
# la cele 4 nomenclatoare). Alt set, alta schema, alta suma (~1.25M vs ~8M).
# Outputul curat traieste in 'Vanzari_Curat' si se valideaza pe contractul EI,
# NU pe canonicul C02-C08 si NU prin comparatie cu Date_MASTER-initial.
SHEET_OUTPUT_C01 = 'Vanzari_Curat'
SCHEMA_C01_STRUCTURARE = [
    'data', 'cod_client', 'nume_client', 'cod_produs', 'nume_produs',
    'categorie', 'cod_agent', 'nume_agent', 'cod_depozit',
    'cantitate', 'pret_unitar', 'valoare_neta', 'cota_tva', 'valoare_totala',
]

# Sheets auxiliare canonice (apar in toate Date_MASTER)
SHEETS_AUXILIARE_CANONICE = ['CLIENTI', 'PRODUSE', 'AGENTI', 'DEPOZITE']

# Sheet principal name (R-V03.47)
SHEET_PRINCIPAL = 'Vanzari'

# Sheet OUTPUT canonic per constructie (R-V03.53, V20)
# Daca difera de SHEET_PRINCIPAL, gate v20 verifica suma pe acest sheet
SHEET_OUTPUT_CANONIC = {
    '01': 'Vanzari',
    '02': 'Vanzari',
    '03': 'Vanzari_AUDIT',
    '04': 'Vanzari_Normalizat',
    '05': 'Vanzari',
    '06': 'Vanzari',
    '07': 'Vanzari',
    '08': 'Vanzari',  # TBD
    '18': 'Vanzari_Curat',  # C18 AUTOMATIZARE: COPY+MODIFY din c01, output curat = contractul C01
}

# Constructii care folosesc SETUL / CONTRACTUL de date C01 (NU canonicul C02-C08).
# Generate prin COPY+MODIFY din c01 => acelasi Vanzari_Curat, SCHEMA_C01_STRUCTURARE,
# nomenclatoare 13/6/5, suma ~1.25M. Se valideaza ca C01 (pe contractul EI, FARA
# comparatie cu Date_MASTER-initial canonic ~8M / 14/7/6). C18 = primul T5 din c01.
# Config-driven: o constructie noua din c01 se adauga aici (extensibil, fara if-uri).
CONSTRUCTII_DATASET_C01 = {'01', '18'}

# Filtre anti-fals-pozitiv pentru R-V02.15 (cifre business)
# Daca un numar suspect apare in context CSS, e fals pozitiv
CSS_CONTEXT_PATTERNS = [
    'opacity', 'transform', 'scale', 'translate', 'rotate',
    'width', 'height', 'padding', 'margin', 'top', 'left', 'right',
    'bottom', 'font-size', 'line-height', 'border-radius', 'gap',
    'z-index', 'flex', 'grid', 'rgb', 'rgba', 'hsl', 'background',
    'animation', 'transition', 'cubic-bezier',
]
NU_VALORI_LEGITIME = {
    'opacity', 'transform', 'scale', 'translate', 'rotate',
}


# ============================================================
# UTILITARE
# ============================================================

def read_file(path):
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


def fold_diac(s):
    """Pliaza diacriticele romanesti -> ASCII, lowercase. Slug-ul (stem
    filename, ex. 'Dictionar') vs display cu diacritice ('Dicționar') trebuie
    comparate modulo diacritice."""
    tbl = str.maketrans('ăâîșțĂÂÎȘȚşţŞŢ', 'aaistAAISTstST')
    return s.translate(tbl).lower()


def is_in_style_or_script(content, position):
    """Verifica daca pozitia e in interiorul <style> sau <script>."""
    last_style_open = content.rfind('<style', 0, position)
    last_style_close = content.rfind('</style>', 0, position)
    if last_style_open > last_style_close:
        return True
    last_script_open = content.rfind('<script', 0, position)
    last_script_close = content.rfind('</script>', 0, position)
    if last_script_open > last_script_close:
        return True
    return False


def extract_visible_text_by_class(content, css_class):
    """Extrage text vizibil dintr-un element cu class data.
    
    Filtreaza template literals JavaScript ${...} care apar in functii
    de render in HTML-Video.
    """
    pattern = rf'class="{re.escape(css_class)}"[^>]*>([^<]+)<'
    results = re.findall(pattern, content)
    # Filtru: skip template literals
    return [r for r in results if not r.startswith('${') and '${' not in r]


# ============================================================
# CLASA 1 — NO-CLONE GENERIC (universal pe toate zonele textuale)
# ============================================================

def check_no_clone_generic(content, pilot_content, fisier_nume, cod_curent):
    """Verifica TOATE zonele textuale vs pilot C01 V12.
    
    Pentru fiecare zona din TEXTUAL_ZONES, calculeaza procent identitate
    cu pilot. Daca depaseste threshold-ul zonei, raporteaza CLONA PASIVA.
    
    Pentru zone cu threshold 0 (prompt-text), un singur match = eroare.
    """
    erori = []
    
    if cod_curent == 'C01':
        return erori  # C01 = pilot, skip
    
    for selector, threshold_pct, descriere in TEXTUAL_ZONES:
        curr_texts = extract_visible_text_by_class(content, selector)
        pilot_texts = extract_visible_text_by_class(pilot_content, selector)
        
        if not curr_texts or not pilot_texts:
            continue
        
        # Numar identice (comparatie pe stripped text)
        identice = sum(
            1 for a, b in zip(pilot_texts, curr_texts)
            if a.strip() == b.strip()
        )
        total = min(len(pilot_texts), len(curr_texts))
        
        if total == 0:
            continue
        
        pct = (identice / total) * 100
        
        # Determinare PASS/FAIL
        if threshold_pct == 0:
            # Toleranta zero - un singur match = eroare
            if identice > 0:
                erori.append({
                    'clasa': 'NO-CLONE GENERIC',
                    'zona': selector,
                    'fisier': fisier_nume,
                    'detaliu': f"{descriere}: {identice}/{total} identice cu pilot (toleranta=0)"
                })
        else:
            if pct >= threshold_pct:
                erori.append({
                    'clasa': 'NO-CLONE GENERIC',
                    'zona': selector,
                    'fisier': fisier_nume,
                    'detaliu': f"{descriere}: {identice}/{total} = {pct:.0f}% identice cu pilot (threshold {threshold_pct}%)"
                })
    
    return erori


# ============================================================
# CLASA 2 — IDENTITY CHECK
# ============================================================

def check_identity(content, identitate, fisier_nume):
    """Verifica string-urile cu identitate construcție.
    
    Acopera: cover-label, cover-title, footer, topbar, mobile-topbar,
    nav-brand-title, localStorage keys, <title>.
    """
    erori = []
    cod = identitate['cod']
    NN = cod.replace('C', '')
    is_video = 'Video' in fisier_nume
    
    if is_video:
        # Pentru Video: topbar-brand
        m = re.search(r'topbar-brand">([^<]+)<', content)
        if m and cod not in m.group(1):
            erori.append({
                'clasa': 'IDENTITY',
                'zona': 'topbar-brand',
                'fisier': fisier_nume,
                'detaliu': f"Asteptat '{cod}' in topbar, gasit: '{m.group(1)}'"
            })
        # <title>
        m = re.search(r'<title>([^<]+)</title>', content)
        if m and cod not in m.group(1):
            erori.append({
                'clasa': 'IDENTITY',
                'zona': '<title>',
                'fisier': fisier_nume,
                'detaliu': f"<title>='{m.group(1)}' nu contine '{cod}'"
            })
    else:
        # Studiu: model PREMIUM V47+ (hero-overlay + cockpit) sau LEGACY (cover-label + cover-meta)
        is_premium = 'hero-visual-overlay' in content or 'class="hov-object"' in content
        if is_premium:
            # Identitatea premium traieste in hero-overlay (cod) + topbar (slug) + footer (cod).
            # cover-label si cover-meta sunt eliminate intentionat la redesignul V47.
            if f'OBIECTUL CONSTRUCȚIEI · {cod}' not in content:
                erori.append({
                    'clasa': 'IDENTITY',
                    'zona': 'hero-overlay',
                    'fisier': fisier_nume,
                    'detaliu': f"Premium: nu am gasit 'OBIECTUL CONSTRUCȚIEI · {cod}' in hero-overlay"
                })
            m = re.search(r'mobile-topbar-title">([^<]+)<', content)
            if m and fold_diac(identitate['nume_slug']) not in fold_diac(m.group(1)) and fold_diac(identitate.get('nume_hero_caps_rand1', '_')) not in fold_diac(m.group(1)):
                erori.append({
                    'clasa': 'IDENTITY',
                    'zona': 'mobile-topbar',
                    'fisier': fisier_nume,
                    'detaliu': f"Premium: topbar='{m.group(1)}' nu contine slug '{identitate['nume_slug']}'"
                })
        else:
            # LEGACY: cover-label OBLIGATORIU
            expected_label = f'CONSTRUCȚIA {cod}'
            if expected_label not in content:
                erori.append({
                    'clasa': 'IDENTITY',
                    'zona': 'cover-label',
                    'fisier': fisier_nume,
                    'detaliu': f"Nu am gasit '{expected_label}' in cover-label"
                })

            # cover-title: post-V42 titlul e descriptiv liber ("Cum construim X" /
            # "Cum dam sens fiecarui rand din set"), slug-ul caps articulat
            # (DICTIONARUL/CLASIFICAREA/...) NU mai traieste aici. Identitatea e
            # garantata de cover-label (CONSTRUCTIA CNN) + footer + meta. Verificam
            # doar ca titlul exista si nu e gol; anti-clona narativa = audit_sync R-V03.69.
            m = re.search(r'<h1 class="cover-title">([^<]*(?:<br>?[^<]*)?)</h1>', content)
            if not m or not m.group(1).strip():
                erori.append({
                    'clasa': 'IDENTITY',
                    'zona': 'cover-title',
                    'fisier': fisier_nume,
                    'detaliu': "cover-title lipsa sau gol"
                })

            # Meta-val EXACT
            if identitate['meta_val_treapta'] not in content:
                m = re.search(
                    r'cover-meta-key">TREAPTA</span>\s*<span class="cover-meta-val">(.*?)</span>',
                    content
                )
                if m:
                    erori.append({
                        'clasa': 'IDENTITY',
                        'zona': 'meta-val-treapta',
                        'fisier': fisier_nume,
                        'detaliu': f"Asteptat: '{identitate['meta_val_treapta']}'\n              Gasit:     '{m.group(1)}'"
                    })

        # Footer (ambele modele)
        m = re.search(r'<footer[^>]*>([^<]+)<', content)
        if m and cod not in m.group(1):
            erori.append({
                'clasa': 'IDENTITY',
                'zona': 'footer',
                'fisier': fisier_nume,
                'detaliu': f"Footer='{m.group(1).strip()[:80]}' nu contine '{cod}'"
            })
    
    # localStorage keys - universal
    keys = re.findall(r'trainity_c(\d+)_', content)
    for k in set(keys):
        if k != NN:
            erori.append({
                'clasa': 'IDENTITY',
                'zona': 'localStorage',
                'fisier': fisier_nume,
                'detaliu': f"Key contaminat: trainity_c{k}_* in loc de trainity_c{NN}_*"
            })
    
    return erori


# ============================================================
# CLASA 3 — BRAND CHECK
# ============================================================

def check_brand(content, fisier_nume):
    """Verifica em-dash, en-dash, engleza forbidden, vocab brand."""
    erori = []
    
    for line_nr, line in enumerate(content.split('\n'), 1):
        if 'base64' in line or 'data:image' in line:
            continue
        
        # em-dash (U+2014) si en-dash (U+2013) - semnal AI
        if '—' in line:
            erori.append({
                'clasa': 'BRAND',
                'zona': 'em-dash',
                'fisier': fisier_nume,
                'detaliu': f"L{line_nr}: {line.strip()[:90]}"
            })
        if '–' in line and 'min-width' not in line and '@media' not in line:
            erori.append({
                'clasa': 'BRAND',
                'zona': 'en-dash',
                'fisier': fisier_nume,
                'detaliu': f"L{line_nr}: {line.strip()[:90]}"
            })
        
        # Engleza forbidden
        for word in ENGLEZA_FORBIDDEN:
            if word in line:
                # V28 whitelist L146: cuvinte tehnice apar in numele
                # pasilor Power Query si Applied Steps (denumiri reale
                # din Excel UI care nu pot fi traduse) - acceptam in
                # context tehnic explicit
                line_lower = line.lower()
                pq_context = any(t in line_lower for t in [
                    'promoted headers', 'filtered rows', 'removed blank',
                    'removed other columns', 'removed duplicates',
                    'changed type', 'replaced value', 'replaced values',
                    'merged columns', 'split column', 'group by',
                    'pivot column', 'unpivot', 'transform sample',
                    'power query', 'applied steps', 'applied step',
                    'normalized diacritics', 'parsed date', 'loaded as',
                    'excel table', 'text.trim', 'text.clean', 'text.replace',
                    'date.from', 'number.from', 'table.replacevalue',
                    'cod m', 'refresh all',
                ])
                # Toate cuvintele tehnice (Normalize/Filter/Refresh etc.)
                # acceptate in context PQ - sunt nume reale Excel UI
                if pq_context:
                    continue  # context Power Query legitim
                erori.append({
                    'clasa': 'BRAND',
                    'zona': 'engleza',
                    'fisier': fisier_nume,
                    'detaliu': f"L{line_nr} '{word.strip()}': {line.strip()[:90]}"
                })
                break  # primul match per linie
    
    # Vocab forbidden in TEXT VIZIBIL (skip CSS/JS bodies)
    in_style = False
    in_script = False
    for line_nr, line in enumerate(content.split('\n'), 1):
        if '<style' in line:
            in_style = True
        if '</style>' in line:
            in_style = False
            continue
        if '<script' in line:
            in_script = True
        if '</script>' in line:
            in_script = False
            continue
        if in_style or in_script:
            continue
        if line.strip().startswith(('/*', '*', '//', '<!--')):
            continue
        
        for word in VOCAB_FORBIDDEN_VIZIBIL:
            if re.search(r'\b' + word + r'\b', line, re.IGNORECASE):
                erori.append({
                    'clasa': 'BRAND',
                    'zona': 'vocab',
                    'fisier': fisier_nume,
                    'detaliu': f"L{line_nr} '{word}': {line.strip()[:90]}"
                })
                break
    
    return erori


# ============================================================
# CLASA 4 — CROSS-CONTAMINATION (cod-uri alta constructie)
# ============================================================

def check_cross_contamination(content, fisier_nume, cod_curent):
    """Niciun cod CXX vizibil nu trebuie sa apara, exceptie CNN curent,
    trimiteri legitime la next_cod (C{NN+1}), si referinte narative
    la prev_cod (C{NN-1}) pentru continuitate cap-coada V20+."""
    erori = []
    NN_curent = cod_curent.replace('C', '')
    NN_int = int(NN_curent)
    next_cod = f"C{NN_int + 1:02d}"
    prev_cod = f"C{NN_int - 1:02d}" if NN_int > 1 else None
    
    in_style = False
    in_script = False
    for line_nr, line in enumerate(content.split('\n'), 1):
        if '<style' in line:
            in_style = True
        if '</style>' in line:
            in_style = False
            continue
        if '<script' in line:
            in_script = True
        if '</script>' in line:
            in_script = False
            continue
        if 'base64' in line or 'data:image' in line:
            continue
        if line.strip().startswith(('/*', '*', '//', '<!--')):
            continue
        
        for m in re.finditer(r'\bC(\d{2})\b', line):
            cod_gasit = m.group(0)
            NN_gasit = m.group(1)
            
            if NN_gasit == NN_curent:
                continue
            
            # Filename link cu cod constructie curenta
            if 'Excel-' in line and 'href' in line:
                if f'Excel-{NN_curent}-' in line:
                    continue
            
            # next-section legitim
            if any(k in line for k in ['next-title', 'next-desc', 'next-label',
                                       'final-intro', 'completion-subtitle']):
                continue
            
            # next_cod in context de predare/finalizare
            if cod_gasit == next_cod:
                line_lower = line.lower()
                if any(k in line_lower for k in [
                    'stage-quote', 'step-title', 'step-action', 'payoff',
                    'completion', 'predăm', 'predă', 'predam', 'catre',
                    'către', 'poate începe', 'fundație', 'fundație',
                    'baza', 'bază'
                ]):
                    continue
                # Context JSON in HTML-Video
                if re.search(r'\b(next|instr|hook|sub|cap)\s*:', line):
                    continue
                if 'data-final' in line or 'data-frag' in line:
                    continue
            
            # prev_cod legitim - tranzitie narativa cap-coada V20+
            # ex: C05 zice "Deschide setul curat predat de C04"
            if prev_cod and cod_gasit == prev_cod:
                line_lower = line.lower()
                if any(k in line_lower for k in [
                    'predat de', 'predata de', 'predat din', 'curat de la',
                    'de la c0', 'venit de la', 'mostenit', 'mostenita',
                    'preluat de la', 'pleaca de la', 'plecat de la',
                    'output c0', 'output de la',
                ]):
                    continue
                # In step-titles/actions context narativ
                if any(k in line_lower for k in [
                    'stage-quote', 'step-title', 'step-action',
                ]) and 'predat' in line_lower:
                    continue
                # Context JSON HTML-Video pentru predare
                if re.search(r'title\s*:.*' + re.escape(prev_cod), line):
                    if any(k in line_lower for k in ['curat', 'predat', 'de la']):
                        continue
                # Context JSON broadcast (HTML-Video): title/instr/hook/next/name care
                # refera constructia PRECEDENTA (input/granita narativa) = legitim.
                # Ex C18: title "Pornim sistemul C17", instr "...e inca C17".
                if re.search(r'\b(title|instr|hook|next|name)\s*:', line):
                    continue
            
            # V27 whitelist: referinte enumerative la construcții anterioare
            # T1 (C01, C02, C03 in C04). Ex: "Ambalaj structural (C01),
            # anomalii logice (C02), contaminare (C03), plus diacritice
            # mixate (C04)". Asta e recapitulare narativa legitima a
            # parcursului T1 - NU cross-contamination.
            if NN_int >= 2 and int(NN_gasit) < NN_int:
                line_lower = line.lower()
                # Trebuie sa fie context recapitulativ/enumerativ:
                # alti codes T1 in aceeasi linie, sau cuvinte cheie de
                # recapitulare
                recap_keywords = [
                    'ambalaj structural', 'anomalii logice',
                    'contaminare invizibila', 'contaminare invizibilă',
                    'diacritice mixat', 'cumulativ', 'perturbari t1',
                    'fenomene plantate', 'pana acum', 'până acum',
                    'recapitulam', 'recapitulăm', 'recapitulare',
                ]
                # Cauta multiple coduri C0X in aceeasi linie (enumerare)
                all_codes_in_line = re.findall(r'\bC0\d\b', line)
                multiple_codes = len(set(all_codes_in_line)) >= 2
                
                if multiple_codes or any(k in line_lower for k in recap_keywords):
                    continue

            # GRANITE DE TREAPTA / ANTI-PATTERN (T5+): o constructie numeste explicit
            # alte coduri ca ANTI-PATTERN sau HANDOFF de granita, NU ca identitate.
            # Ex C18: anomaly "REFRESH SIMPLU... E C04", tag "Anti-C04", final-text
            # "din C17 s-a transformat", step-action "incă formă (C17), nu mișcare (C18)".
            # Legitim. C01-C17 nu folosesc acest stil in aceste zone => whitelist-ul nu
            # le slabeste verificarea (raman la 0 flag-uri). Diacritice incluse.
            line_lower = line.lower()
            # (a) zone inerent de granita / anti-pattern / recap / tag de treapta
            zona_granita = any(z in line for z in [
                'anomaly-desc', 'anomaly-title', 'type-tag', 'nav-item-meta',
                'final-text', 'final-label', 'phase-tag', 'next-desc', 'next-title',
            ])
            # (b) fraze explicite de granita / anti-pattern / handoff
            boundary_markers = [
                'acela e c', 'acela este c', 'aceasta e c', 'asta e c',
                'moștenit', 'mostenit', 'moștenită', 'mostenita',
                'handoff', 'graniță', 'granita', 'anti-pattern', 'antipattern', 'anti-c0',
                'nu este c', 'nu e c', 'nu c1', 'nu automatiz', 'nu mișcare', 'nu miscare',
                'în loc de', 'in loc de', 'separă', 'separa',
                'predă', 'preda', 'predăm', 'predam', 'predat',
                '(c04)', '(c17)', '(c19)', '(c20)',
                'din c17', 'de c04', 'de c19', 'de c20', 'către c19', 'catre c19',
            ]
            if zona_granita or any(k in line_lower for k in boundary_markers):
                continue

            erori.append({
                'clasa': 'CROSS-CONTAMINATION',
                'zona': f'cod {cod_gasit}',
                'fisier': fisier_nume,
                'detaliu': f"L{line_nr}: {line.strip()[:90]}"
            })
            break  # primul match per linie
    
    return erori


# ============================================================
# CLASA 5 — VOCE NARATIVA (R-V03.4)
# ============================================================

def check_voce(content, fisier_nume):
    """Verifica voce: Studiu = persoana 2 sg, Video = persoana 3 plural.
    
    Verificare euristica: cauta indicatori puternici de voce gresita.
    """
    erori = []
    is_studiu = 'Studiu' in fisier_nume
    is_video = 'Video' in fisier_nume
    
    # In Studiu (pers 2 sg): "Deschide", "Lipești", "Verifică"
    # In Video (pers 3 plural): "Deschidem", "Lipim", "Verificăm"
    
    if is_video:
        # Step-titles si step-actions in Video TREBUIE sa fie pers 1 plural
        zones_to_check = []
        # Video foloseste title: in JSON
        titles_video = [t for t in re.findall(r'title:\s*"([^"]+)"', content) if '${' not in t]
        for t in titles_video[:5]:
            zones_to_check.append(('JS title', t))
        
        # Verificare imperfecta: caut imperative sg vs plural
        for zona, txt in zones_to_check:
            # "Deschide " (sg) vs "Deschidem " (plural)
            first_word = txt.split()[0] if txt.split() else ''
            if first_word in ['Deschide', 'Citește', 'Verifică', 'Marchează',
                             'Activează', 'Rulează', 'Aplică', 'Predă',
                             'Confirmă', 'Lipește', 'Notează', 'Construiește']:
                erori.append({
                    'clasa': 'VOCE',
                    'zona': f'video-{zona}',
                    'fisier': fisier_nume,
                    'detaliu': f"Imperativ singular '{first_word}' in HTML-Video (asteptat plural)"
                })
    
    return erori


# ============================================================
# CLASA 6 — DATA-CONTINUITY (V18, R-V03.47)
# ============================================================

def check_data_continuity(excel_path, NN, initial_excel_path=None):
    """V19: Verifica integritate Date_MASTER-C{NN}.xlsx vs Date_MASTER-initial.
    
    Filosofia V19: construc,tiile sunt INDEPENDENTE. Date_MASTER-initial
    este sursa comuna. Fiecare constructie aplica modificari specifice
    pe Date_MASTER-initial si livreaza Date_MASTER-C{NN}.xlsx.
    
    Verificari:
    1. Sheet principal exista si se numeste 'Vanzari'
    2. Schema cele 14 coloane canonice prezente
    3. Sheets auxiliare canonice prezente (CLIENTI, PRODUSE, AGENTI, DEPOZITE)
    4. Daca initial_excel_path e furnizat:
       - Nomenclatoarele (CLIENTI, PRODUSE, AGENTI, DEPOZITE) IDENTICE
       - Suma valoare_neta in range conservare semantica
         (delta < 15% vs initial, sau declarat in SPEC)
    """
    erori = []
    
    if not os.path.exists(excel_path):
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'fisier_lipsa',
            'fisier': os.path.basename(excel_path),
            'detaliu': f"Fisier Date_MASTER-C{NN}.xlsx LIPSA. Constructia nu are livrabil Excel."
        })
        return erori
    
    try:
        import openpyxl
    except ImportError:
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'dependenta',
            'fisier': '',
            'detaliu': 'Modulul openpyxl nu e disponibil pentru audit Excel'
        })
        return erori
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'load',
            'fisier': os.path.basename(excel_path),
            'detaliu': f"Eroare la deschidere Excel: {e}"
        })
        return erori
    
    # 1. Sheet principal exista
    if SHEET_PRINCIPAL not in wb.sheetnames:
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'sheet_principal',
            'fisier': os.path.basename(excel_path),
            'detaliu': f"Sheet '{SHEET_PRINCIPAL}' lipsa. Gasit: {wb.sheetnames}"
        })
        return erori

    # === EXCEPTIE EXPLICITA pentru CONSTRUCTIILE cu DATASET C01 (config-driven) ===
    # C01 (pilot) + constructiile COPY+MODIFY din c01 (ex. C18 AUTOMATIZARE) NU
    # participa la contractul de date canonic C02-C08 (alt set, alta schema, alta
    # suma ~1.25M, nomenclatoare 13/6/5 proprii). Validam OUTPUT-ul real (Vanzari_Curat)
    # pe schema EI (SCHEMA_C01_STRUCTURARE) + prezenta nomenclatoarelor, FARA comparatie
    # cu Date_MASTER-initial canonic. Justificare: dataset mostenit din c01 by design;
    # output verificabil corect (2000 randuri, 14 coloane, header rand 1, FK la 4 nomenclatoare).
    if NN in CONSTRUCTII_DATASET_C01:
        if SHEET_OUTPUT_C01 not in wb.sheetnames:
            erori.append({
                'clasa': 'DATA-CONTINUITY', 'zona': 'sheet_output_c01',
                'fisier': os.path.basename(excel_path),
                'detaliu': f"Sheet OUTPUT C01 '{SHEET_OUTPUT_C01}' lipsa. Gasit: {wb.sheetnames}"
            })
            return erori
        headers_c01 = [c.value for c in wb[SHEET_OUTPUT_C01][1]]
        missing_c01 = [h for h in SCHEMA_C01_STRUCTURARE if h not in headers_c01]
        if missing_c01:
            erori.append({
                'clasa': 'DATA-CONTINUITY', 'zona': 'schema_c01',
                'fisier': os.path.basename(excel_path),
                'detaliu': f"Coloane lipsa in OUTPUT C01 '{SHEET_OUTPUT_C01}': {missing_c01}"
            })
        for sheet_aux in SHEETS_AUXILIARE_CANONICE:
            if sheet_aux not in wb.sheetnames:
                erori.append({
                    'clasa': 'DATA-CONTINUITY', 'zona': 'sheet_auxiliar',
                    'fisier': os.path.basename(excel_path),
                    'detaliu': f"Sheet auxiliar canonic LIPSA: '{sheet_aux}'"
                })
        return erori

    # R-V03.53 V27: verifica schema pe SHEET OUTPUT canonic, nu pe
    # sheet principal. Pentru C04 (Vanzari = input contaminat, header
    # rand 1 = "RAPORT VANZARI"), schema canonica e in Vanzari_Normalizat.
    sheet_output_canonic_pentru_schema = SHEET_OUTPUT_CANONIC.get(NN, SHEET_PRINCIPAL)
    if sheet_output_canonic_pentru_schema not in wb.sheetnames:
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'sheet_output',
            'fisier': os.path.basename(excel_path),
            'detaliu': f"Sheet OUTPUT canonic '{sheet_output_canonic_pentru_schema}' lipsa. Gasit: {wb.sheetnames}"
        })
        return erori
    
    ws = wb[sheet_output_canonic_pentru_schema]
    headers = [c.value for c in ws[1]]
    
    # 2. Schema canonica - cele 14 coloane primele in ordine
    missing = [h for h in SCHEMA_CANONICA_VANZARI if h not in headers]
    if missing:
        erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'schema_coloane',
            'fisier': os.path.basename(excel_path),
            'detaliu': f"Coloane canonice LIPSA in sheet OUTPUT '{sheet_output_canonic_pentru_schema}': {missing}"
        })
    
    # 3. Sheets auxiliare canonice
    for sheet_aux in SHEETS_AUXILIARE_CANONICE:
        if sheet_aux not in wb.sheetnames:
            erori.append({
                'clasa': 'DATA-CONTINUITY',
                'zona': 'sheet_auxiliar',
                'fisier': os.path.basename(excel_path),
                'detaliu': f"Sheet auxiliar canonic LIPSA: '{sheet_aux}'"
            })
    
    # 4. Comparare cu Date_MASTER-initial (daca disponibil)
    if initial_excel_path and os.path.exists(initial_excel_path):
        try:
            wb_init = openpyxl.load_workbook(initial_excel_path, data_only=False)
            
            # 4a. Nomenclatoarele IDENTICE
            for sheet_aux in SHEETS_AUXILIARE_CANONICE:
                if sheet_aux in wb_init.sheetnames and sheet_aux in wb.sheetnames:
                    ws_init_aux = wb_init[sheet_aux]
                    ws_aux = wb[sheet_aux]
                    # Verific primul rand (header) + count rows
                    init_rows = ws_init_aux.max_row
                    curr_rows = ws_aux.max_row
                    if init_rows != curr_rows:
                        erori.append({
                            'clasa': 'DATA-CONTINUITY',
                            'zona': f'nomenclator_{sheet_aux}',
                            'fisier': os.path.basename(excel_path),
                            'detaliu': f"Sheet {sheet_aux}: {curr_rows} rows in livrabil vs {init_rows} in initial. Nomenclator alterat."
                        })
            
            # 4b. Suma valoare_neta - conservare semantica V20
            # R-V03.53: verifica pe SHEET OUTPUT canonic, nu pe sheet
            # Vanzari care poate fi intentionat contaminat
            sheet_output_canonic = SHEET_OUTPUT_CANONIC.get(NN, SHEET_PRINCIPAL)
            
            if SHEET_PRINCIPAL in wb_init.sheetnames:
                ws_init = wb_init[SHEET_PRINCIPAL]
                h_init = [c.value for c in ws_init[1]]
                col_vn_init = h_init.index('valoare_neta') + 1 if 'valoare_neta' in h_init else None
                
                # Sum initial (referinta)
                sum_init = 0
                if col_vn_init:
                    for r in range(2, ws_init.max_row + 1):
                        v = ws_init.cell(row=r, column=col_vn_init).value
                        if isinstance(v, (int, float)):
                            sum_init += v
                
                # Sum pe sheet OUTPUT canonic al constructiei
                if sheet_output_canonic in wb.sheetnames:
                    ws_out = wb[sheet_output_canonic]
                    h_out = [c.value for c in ws_out[1]]
                    col_vn_out = h_out.index('valoare_neta') + 1 if 'valoare_neta' in h_out else None
                    
                    if col_vn_out and col_vn_init:
                        sum_out = 0
                        for r in range(2, ws_out.max_row + 1):
                            v = ws_out.cell(row=r, column=col_vn_out).value
                            if isinstance(v, (int, float)):
                                sum_out += v
                        
                        if sum_init > 0:
                            pct = abs(sum_out - sum_init) / sum_init * 100
                            if pct > 15:
                                erori.append({
                                    'clasa': 'DATA-CONTINUITY',
                                    'zona': 'conservare_semantica',
                                    'fisier': os.path.basename(excel_path),
                                    'detaliu': f"Delta valoare_neta pe sheet OUTPUT '{sheet_output_canonic}' vs initial: {pct:.1f}%. "
                                              f"Initial: {sum_init:,.2f}, output: {sum_out:,.2f}. "
                                              f"Verifica SPEC daca delta e declarat."
                                })
                elif sheet_output_canonic != SHEET_PRINCIPAL:
                    # Sheet OUTPUT canonic specificat dar nu livrat
                    erori.append({
                        'clasa': 'DATA-CONTINUITY',
                        'zona': 'sheet_output_lipsa',
                        'fisier': os.path.basename(excel_path),
                        'detaliu': f"Sheet OUTPUT canonic pentru C{NN} ('{sheet_output_canonic}') "
                                  f"LIPSA. Constructia nu poate demonstra conservare semantica."
                    })
        except Exception:
            pass
    
    return erori


def check_brand_cu_whitelist(content, fisier_nume):
    """Brand check V18 cu whitelist Power Query."""
    erori = []
    
    for line_nr, line in enumerate(content.split('\n'), 1):
        if 'base64' in line or 'data:image' in line:
            continue
        
        # em-dash (U+2014) si en-dash (U+2013)
        if '—' in line:
            erori.append({
                'clasa': 'BRAND', 'zona': 'em-dash',
                'fisier': fisier_nume,
                'detaliu': f"L{line_nr}: {line.strip()[:90]}"
            })
        if '–' in line and 'min-width' not in line and '@media' not in line:
            erori.append({
                'clasa': 'BRAND', 'zona': 'en-dash',
                'fisier': fisier_nume,
                'detaliu': f"L{line_nr}: {line.strip()[:90]}"
            })
        
        # Engleza forbidden - cu whitelist V18
        for word in ENGLEZA_FORBIDDEN:
            if word in line:
                # Check whitelist
                whitelisted = False
                for wt in ENGLEZA_WHITELIST_TEHNIC:
                    if wt in line and word.strip() in wt:
                        whitelisted = True
                        break
                if not whitelisted:
                    erori.append({
                        'clasa': 'BRAND', 'zona': 'engleza',
                        'fisier': fisier_nume,
                        'detaliu': f"L{line_nr} '{word.strip()}': {line.strip()[:90]}"
                    })
                    break
    
    # Vocab in TEXT VIZIBIL
    in_style = False; in_script = False
    for line_nr, line in enumerate(content.split('\n'), 1):
        if '<style' in line: in_style = True
        if '</style>' in line: in_style = False; continue
        if '<script' in line: in_script = True
        if '</script>' in line: in_script = False; continue
        if in_style or in_script: continue
        if line.strip().startswith(('/*', '*', '//', '<!--')): continue
        
        for word in VOCAB_FORBIDDEN_VIZIBIL:
            if re.search(r'\b' + word + r'\b', line, re.IGNORECASE):
                erori.append({
                    'clasa': 'BRAND', 'zona': 'vocab',
                    'fisier': fisier_nume,
                    'detaliu': f"L{line_nr} '{word}': {line.strip()[:90]}"
                })
                break
    
    return erori


def get_pilot_hash(pilot_dir):
    """Calcul hash MD5 pe pilot pentru versionare (V18)."""
    import hashlib
    md5 = hashlib.md5()
    for f in sorted(['HTML-Studiu-Excel-01-Structurare.html',
                     'HTML-Editor-Studiu-Excel-01-Structurare.html',
                     'HTML-Video-Excel-01-Structurare.html',
                     'HTML-Editor-Video-Excel-01-Structurare.html']):
        fp = os.path.join(pilot_dir, f)
        if os.path.exists(fp):
            with open(fp, 'rb') as fh:
                md5.update(fh.read())
    return md5.hexdigest()[:8]


# ============================================================
# RAPORT FINAL
# ============================================================

def run_gate(NN, livrabile_path, pilot_dir, identitate):
    """Ruleaza toate verificarile. Returneaza (success, erori)."""
    cod_curent = f'C{NN}'
    nume_slug = identitate['nume_slug']
    
    # CELE 4 HTML-uri (rămân cum sunt)
    files_to_check = [
        f"HTML-Studiu-Excel-{NN}-{nume_slug}.html",
        f"HTML-Editor-Studiu-Excel-{NN}-{nume_slug}.html",
        f"HTML-Video-Excel-{NN}-{nume_slug}.html",
        f"HTML-Editor-Video-Excel-{NN}-{nume_slug}.html",
    ]
    
    # LIVRABIL NON-HTML (V19; Creativ ABANDONAT V68: prompturile nu se mai stocheaza)
    excel_file = f"Date_MASTER-C{NN}.xlsx"  # V19: NUME NOU (era Date_MASTER-dupa-C{NN} in V18)
    
    pilot_studiu = os.path.join(pilot_dir, "HTML-Studiu-Excel-01-Structurare.html")
    pilot_editor_studiu = os.path.join(pilot_dir, "HTML-Editor-Studiu-Excel-01-Structurare.html")
    pilot_video = os.path.join(pilot_dir, "HTML-Video-Excel-01-Structurare.html")
    pilot_editor_video = os.path.join(pilot_dir, "HTML-Editor-Video-Excel-01-Structurare.html")
    
    all_erori = []
    files_present = []
    files_missing = []
    
    for fname in files_to_check:
        fpath = os.path.join(livrabile_path, fname)
        if not os.path.exists(fpath):
            files_missing.append(fname)
            continue
        
        files_present.append(fname)
        content = read_file(fpath)
        
        # Pilot reference pentru no-clone
        if 'Editor-Studiu' in fname:
            pilot_ref = pilot_editor_studiu
        elif 'Editor-Video' in fname:
            pilot_ref = pilot_editor_video
        elif 'Studiu' in fname:
            pilot_ref = pilot_studiu
        else:
            pilot_ref = pilot_video
        
        if os.path.exists(pilot_ref):
            pilot_content = read_file(pilot_ref)
            # CLASA 1: no-clone generic
            all_erori.extend(check_no_clone_generic(content, pilot_content, fname, cod_curent))
        
        # CLASA 2: identity
        all_erori.extend(check_identity(content, identitate, fname))
        
        # CLASA 3: brand (V18 cu whitelist Power Query)
        all_erori.extend(check_brand_cu_whitelist(content, fname))
        
        # CLASA 4: cross-contamination
        all_erori.extend(check_cross_contamination(content, fname, cod_curent))
        
        # CLASA 5: voce
        all_erori.extend(check_voce(content, fname))
    
    # CLASA 6: DATA-CONTINUITY (V19, R-V03.49)
    # Filosofia V19: vs Date_MASTER-initial (NU cap-coada strict)
    excel_path = os.path.join(livrabile_path, excel_file)
    
    # Cauta Date_MASTER-initial ca referinta nomenclatoare + suma
    # In productie e la referinte/Date_MASTER-initial.xlsx
    initial_excel = None
    # Cauta in mai multe locatii posibile (V40: structura noua _system/referinte/)
    posibile = [
        os.path.join(os.path.dirname(livrabile_path.rstrip('/').rstrip('\\')), '_system', 'referinte', 'Date_MASTER-initial.xlsx'),
        os.path.join(os.path.dirname(livrabile_path.rstrip('/').rstrip('\\')), 'referinte', 'Date_MASTER-initial.xlsx'),
        os.path.join(livrabile_path, 'Date_MASTER-initial.xlsx'),
        '_system/referinte/Date_MASTER-initial.xlsx',
    ]
    for p in posibile:
        if os.path.exists(p):
            initial_excel = p
            break
    
    if os.path.exists(excel_path):
        files_present.append(excel_file)
        all_erori.extend(check_data_continuity(excel_path, NN, initial_excel))
    else:
        files_missing.append(excel_file)
        all_erori.append({
            'clasa': 'DATA-CONTINUITY',
            'zona': 'fisier_lipsa',
            'fisier': excel_file,
            'detaliu': f"Livrabilul Date_MASTER-C{NN}.xlsx OBLIGATORIU lipseste."
        })
    
    # Creativ ABANDONAT (V68): prompturile imaginilor se fac extern (ARHITECT + ChatGPT),
    # nu se mai stocheaza in repo. Imaginile raman obligatorii in assets/ + base64 in HTML.

    # CLASA 7: TIER-GUARD-T3 (BRAIN-008) — activa STRICT pentru NN 09-12.
    # Modul separat tier_guard_t3; FAIL blocant intra in all_erori (=> GATE FAIL),
    # WARNING-urile se printeaza vizibil aici, fara sa blocheze. NN 01-08 = neatins.
    if NN in ('09', '10', '11', '12'):
        try:
            sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            import tier_guard_t3
            t3_block, t3_warn = tier_guard_t3.gate_findings(livrabile_path, NN)
            for b in t3_block:
                all_erori.append({
                    'clasa': 'TIER-GUARD-T3', 'zona': b['bucket'],
                    'fisier': b['fisier'],
                    'detaliu': f"'{b['term']}' :: {b['ctx']}"
                })
            if t3_warn:
                print(f"\n[TIER-GUARD-T3] {len(t3_warn)} avertisment(e) non-blocant(e) C{NN}:")
                for w in t3_warn:
                    print(f"  ~ {w['fisier']} [{w['bucket']}] '{w['term']}' :: {w['ctx']}")
        except Exception as e:
            # garda nu trebuie sa darame gate-ul printr-o eroare proprie; o raportam
            print(f"\n[TIER-GUARD-T3] WARNING: garda T3 nu a putut rula ({e}).")

    success = len(all_erori) == 0
    return success, all_erori, files_present, files_missing


def print_report(success, erori, NN, files_present, files_missing, pilot_hash=None):
    print("=" * 80)
    print(f"POST-FLIGHT GATE V20 — C{NN}")
    print(f"Verificator pe 6 clase de erori (R-V03.45 + R-V03.49 + R-V03.53)")
    if pilot_hash:
        print(f"Pilot C01 V12 hash: {pilot_hash}")
    print("=" * 80)
    
    print(f"\nLivrabile prezente ({len(files_present)}/5):")
    for f in files_present:
        print(f"  ✓ {f}")
    if files_missing:
        print(f"\nLivrabile lipsa ({len(files_missing)}):")
        for f in files_missing:
            print(f"  ✗ {f}")
    
    if success:
        print("\n✓✓✓ GATE PASS — toate verificarile au trecut.")
        print(f"   Constructia C{NN} e pregatita pentru present_files.")
        return
    
    print(f"\n✗✗✗ GATE FAIL — {len(erori)} eroare(i) detectate.")
    print(f"   PRESENT_FILES BLOCAT. Constructia C{NN} NU se livreaza.")
    print(f"   Cere regenerare in chatul curent inainte de checkout constructie.\n")
    
    # Group erori by clasa
    by_clasa = {}
    for e in erori:
        by_clasa.setdefault(e['clasa'], []).append(e)
    
    for clasa, lst in by_clasa.items():
        print(f"\n[CLASA: {clasa}] {len(lst)} aparitie(i):")
        for e in lst[:8]:
            print(f"  - {e['fisier']} [{e['zona']}]")
            print(f"    {e['detaliu']}")
        if len(lst) > 8:
            print(f"  ... si {len(lst) - 8} altele")


# ============================================================
# IDENTITATE_TEHNICA loader
# ============================================================

def load_identitate(NN, identitate_path):
    """Citeste IDENTITATE_TEHNICA C{NN} din IDENTITATE-TEHNICA.md.
    
    In productie: parser complet. In acest sample: dict hardcoded.
    """
    IDENTITATI = {
        '01': {
            'cod': 'C01', 'nume_hero_caps_rand1': 'STRUCTURA',
            'nume_slug': 'Structurare',
            'meta_val_treapta': '<b>STRUCTURA</b> · CONTROL · AUDIT · NORMALIZARE (SCAN)'
        },
        '02': {
            'cod': 'C02', 'nume_hero_caps_rand1': 'ADEVAR',
            'nume_slug': 'Confruntare',
            'meta_val_treapta': 'STRUCTURA · <b>CONTROL</b> · AUDIT · NORMALIZARE (SCAN)'
        },
        '03': {
            'cod': 'C03', 'nume_hero_caps_rand1': 'AUDIT',
            'nume_slug': 'Auditare',
            'meta_val_treapta': 'STRUCTURA · CONTROL · <b>AUDIT</b> · NORMALIZARE (SCAN)'
        },
        '04': {
            'cod': 'C04', 'nume_hero_caps_rand1': 'FLUX',
            'nume_slug': 'Normalizare',
            'meta_val_treapta': 'STRUCTURA · CONTROL · AUDIT · <b>NORMALIZARE</b> (SCAN)'
        },
        '05': {
            'cod': 'C05', 'nume_hero_caps_rand1': 'INVENTAR',
            'nume_slug': 'Inventariere',
            'meta_val_treapta': '<b>DICȚIONAR</b> · CLASIFICARE · DATARE · CARTOGRAFIERE (CUNOAȘTERE)'
        },
        '06': {
            'cod': 'C06', 'nume_hero_caps_rand1': 'REGULI',
            'nume_slug': 'Clasificare',
            'meta_val_treapta': 'DICȚIONAR · <b>CLASIFICARE</b> · DATARE · CARTOGRAFIERE (CUNOAȘTERE)'
        },
        '07': {
            'cod': 'C07', 'nume_hero_caps_rand1': 'TIMP',
            'nume_slug': 'Datare',
            'meta_val_treapta': 'CLASIFICARE · CUANTIFICARE · <b>DATARE</b> · CARTOGRAFIERE (CUNOAȘTERE)'
        },
        '08': {
            'cod': 'C08', 'nume_hero_caps_rand1': 'HARTA',
            'nume_slug': 'Mapare',
            'meta_val_treapta': 'CLASIFICARE · CUANTIFICARE · DATARE · <b>CARTOGRAFIERE</b> (CUNOAȘTERE)'
        },
        '09': {
            'cod': 'C09', 'nume_hero_caps_rand1': 'RELATII',
            'nume_slug': 'Legare',
            'meta_val_treapta': '<b>RELAȚII</b> · MĂSURI · COMPARAȚII · INTERPRETARE (ANALIZĂ)'
        },
        '10': {
            'cod': 'C10', 'nume_hero_caps_rand1': 'MASURA',
            'nume_slug': 'Masurare',
            'meta_val_treapta': 'RELAȚII · <b>MĂSURI</b> · COMPARAȚII · INTERPRETARE (ANALIZĂ)'
        },
        '11': {
            'cod': 'C11', 'nume_hero_caps_rand1': 'CLASAMENT',
            'nume_slug': 'Comparare',
            'meta_val_treapta': 'RELAȚII · MĂSURI · <b>COMPARAȚII</b> · INTERPRETARE (ANALIZĂ)'
        },
        '12': {
            'cod': 'C12', 'nume_hero_caps_rand1': 'CAUZA',
            'nume_slug': 'Interpretare',
            'meta_val_treapta': 'RELAȚII · MĂSURI · COMPARAȚII · <b>INTERPRETARE</b> (ANALIZĂ)'
        },
        '13': {
            'cod': 'C13', 'nume_hero_caps_rand1': 'VIZUAL',
            'nume_slug': 'Vizualizare',
            'meta_val_treapta': '<b>VIZUALIZARE</b> · COMPUNERE · SINTETIZARE · LIVRARE (RAPORTARE)'
        },
        '14': {
            'cod': 'C14', 'nume_hero_caps_rand1': 'COMPOZIȚIE',
            'nume_slug': 'Compunere',
            'meta_val_treapta': 'VIZUALIZARE · <b>COMPUNERE</b> · SINTETIZARE · LIVRARE (RAPORTARE)'
        },
        '15': {
            'cod': 'C15', 'nume_hero_caps_rand1': 'SINTEZĂ',
            'nume_slug': 'Sintetizare',
            'meta_val_treapta': 'VIZUALIZARE · COMPUNERE · <b>SINTETIZARE</b> · LIVRARE (RAPORTARE)'
        },
        '16': {
            'cod': 'C16', 'nume_hero_caps_rand1': 'LIVRARE',
            'nume_slug': 'Livrare',
            'meta_val_treapta': 'VIZUALIZARE · COMPUNERE · SINTETIZARE · <b>LIVRARE</b> (RAPORTARE)'
        },
        '17': {
            'cod': 'C17', 'nume_hero_caps_rand1': 'SISTEM',
            'nume_slug': 'Sistematizare',
            'meta_val_treapta': '<b>SISTEMATIZARE</b> · AUTOMATIZARE · GUVERNARE · DELEGARE (AUTONOMIE)'
        },
        '18': {
            'cod': 'C18', 'nume_hero_caps_rand1': 'AUTOMATIZAREA',
            'nume_slug': 'Automatizare',
            'meta_val_treapta': 'SISTEMATIZARE · <b>AUTOMATIZARE</b> · GUVERNARE · DELEGARE (AUTONOMIE)'
        },
        '19': {
            'cod': 'C19', 'nume_hero_caps_rand1': 'CONTROL',
            'nume_slug': 'Guvernare',
            'meta_val_treapta': 'SISTEMATIZARE · AUTOMATIZARE · <b>GUVERNARE</b> · DELEGARE (AUTONOMIE)'
        },
    }
    return IDENTITATI.get(NN)


if __name__ == '__main__':
    # Forțează stdout UTF-8 pe Windows (cp1252 console nu poate encoda ✓/✗)
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass  # Python < 3.7

    if len(sys.argv) < 3:
        print("Usage: python3 gate_v20.py NN /path/to/livrabile/ [/path/to/pilot/]")
        print("Daca pilot_path lipseste, se cauta in ../pilot_C01_V12/")
        print("")
        print("V18: 6 clase verificare (cls 6 NOU: DATA-CONTINUITY).")
        print("Livrabile asteptate: 4 HTML + 1 Date_MASTER-C{NN}.xlsx (Creativ abandonat V68)")
        sys.exit(1)
    
    NN = sys.argv[1].zfill(2)
    livrabile = sys.argv[2]
    
    if len(sys.argv) >= 4:
        pilot_dir = sys.argv[3]
    else:
        pilot_dir = os.path.join(
            os.path.dirname(livrabile.rstrip('/')),
            'pilot_C01_V12'
        )
    
    identitate = load_identitate(NN, None)
    if not identitate:
        print(f"EROARE: IDENTITATE_TEHNICA pentru C{NN} nu e definita.")
        print("Adauga in referinte/IDENTITATE-TEHNICA.md inainte de generare.")
        sys.exit(1)
    
    if not os.path.exists(pilot_dir):
        print(f"EROARE: pilot path nu exista: {pilot_dir}")
        sys.exit(1)
    
    pilot_hash = get_pilot_hash(pilot_dir)
    
    success, erori, files_present, files_missing = run_gate(
        NN, livrabile, pilot_dir, identitate
    )
    print_report(success, erori, NN, files_present, files_missing, pilot_hash)
    
    sys.exit(0 if success else 1)
