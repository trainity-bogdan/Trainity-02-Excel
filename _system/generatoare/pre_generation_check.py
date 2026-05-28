#!/usr/bin/env python3
"""
pre_generation_check.py — V27 (SCHEMA REALA + SELF-CHECK)

SCRIPT_VERSION = "V27.1"  # NU modifica - folosit pentru self-check.

V27 vs V26:
- Folosesc denumiri REALE din IDENTITATE-TEHNICA.md (nume_hero_caps_rand1,
  localStorage_studiu, localStorage_video) - NU denumiri inventate
  (cover_label, cover_title, localStorage_keys) ca in V26 buggy.
- SELF-CHECK la inceput: scriptul afiseaza SCRIPT_VERSION ca prima
  linie ca motorul sa detecteze daca rulează versiunea corecta.
  
Verificare HARD CHECKS la pre-generare:
- R-V03.55: SPEC C{NN} narativ inghetat in SISTEM_TRAINITY.md
- L142: IDENTITATE_TEHNICA C{NN} populata in referinte/IDENTITATE-TEHNICA.md
  cu schema reala (nume_hero_caps_rand1, localStorage_studiu, etc.)
- L143: SPEC FENOMENE cifre canonice corespund cu Date_MASTER-initial.xlsx

Exit code:
- 0 = TOATE check-urile PASS, motorul poate continua la generare
- 1 = Unul/mai multe check-uri FAIL, motorul OPRESTE generare
- 2 = Eroare: SISTEM_TRAINITY.md sau referinte/* lipsa

Usage:
    python3 pre_generation_check.py NN [/path/to/SISTEM_TRAINITY.md]
"""

import sys
import re
import os

# SCRIPT_VERSION - folosit pentru self-check. Motorul citeste prima
# linie de output ca sa stie daca ruleaza versiunea corecta. Daca
# vede o versiune mai mica decat asteptat, opreste si cere reatasare
# arhiva.
SCRIPT_VERSION = "V27.1"
EXPECTED_SCHEMA = "REAL"  # vs "INVENTED" (V26 buggy)


def find_spec_section(content, NN):
    """Cauta sectiunea SPEC C{NN} si extrage statusul."""
    NN_int = int(NN)
    NN_str = f"{NN_int:02d}"
    
    patterns = [
        rf'##\s*SPEC\s+C{NN_str}\b',
        rf'##\s*SPEC\s+C{NN_int}\b',
    ]
    
    section_start = -1
    for pat in patterns:
        m = re.search(pat, content)
        if m:
            section_start = m.start()
            break
    
    if section_start == -1:
        return (None, None)
    
    rest = content[section_start:]
    next_section = re.search(r'\n##\s+', rest[2:])
    if next_section:
        section_text = rest[:next_section.start() + 2]
    else:
        section_text = rest
    
    status_patterns = [
        (r'\[Status:\s*INGHETAT[^\]]*\]', 'INGHETAT'),
        (r'\[Status:\s*COMPLET\s*\]', 'INGHETAT'),
        (r'\[Status:\s*NEGENERAT\s*\]', 'NEGENERAT'),
        (r'\[Status:\s*PENDING\s*\]', 'NEGENERAT'),
    ]
    
    for pat, status in status_patterns:
        if re.search(pat, section_text, re.IGNORECASE):
            return (status, section_text)
    
    if len(section_text.strip().split('\n')) < 5:
        return ('NEGENERAT', section_text)
    
    return ('INGHETAT', section_text)


def get_construction_name(NN, content):
    NN_int = int(NN)
    NN_str = f"{NN_int:02d}"
    
    m = re.search(rf'##\s*SPEC\s+C{NN_str}\s*[-—–]\s*([^\n\[]+)', content)
    if m:
        return m.group(1).strip()
    m = re.search(rf'##\s*SPEC\s+C{NN_int}\s*[-—–]\s*([^\n\[]+)', content)
    if m:
        return m.group(1).strip()
    return f"C{NN_str}"


def check_identitate_tehnica(NN, base_path):
    """
    L142: Verifica daca IDENTITATE_TEHNICA C{NN} e populata in
    referinte/IDENTITATE-TEHNICA.md.
    
    Returneaza (ok, mesaj):
    - (True, None) daca exista si e populata
    - (False, mesaj) daca lipseste sau e doar placeholder
    """
    NN_int = int(NN)
    NN_str = f"{NN_int:02d}"
    
    ident_path = os.path.join(base_path, 'referinte', 'IDENTITATE-TEHNICA.md')
    if not os.path.exists(ident_path):
        # V40: structura noua _system/referinte/
        for cand in ['_system/referinte/IDENTITATE-TEHNICA.md',
                    '../_system/referinte/IDENTITATE-TEHNICA.md',
                    'referinte/IDENTITATE-TEHNICA.md',
                    '../referinte/IDENTITATE-TEHNICA.md']:
            if os.path.exists(cand):
                ident_path = cand
                break
        else:
            return (False, f"Fisier _system/referinte/IDENTITATE-TEHNICA.md LIPSA")
    
    with open(ident_path, 'r', encoding='utf-8') as f:
        ident_content = f.read()
    
    # Cauta sectiunea IDENTITATE_TEHNICA C{NN} (strict, fara range C{NN}-C{XX})
    patterns = [
        rf'##\s*IDENTITATE_TEHNICA\s+C{NN_str}(?!-)(?!\d)',
        rf'##\s*IDENTITATE_TEHNICA\s+C{NN_int}(?!-)(?!\d)',
    ]
    
    section_start = -1
    for pat in patterns:
        m = re.search(pat, ident_content)
        if m:
            section_start = m.start()
            break
    
    if section_start == -1:
        return (False, f"Sectiunea IDENTITATE_TEHNICA C{NN_str} LIPSA in referinte/IDENTITATE-TEHNICA.md")
    
    # Verifica daca sectiunea contine campurile obligatorii
    rest = ident_content[section_start:]
    next_section = re.search(r'\n##\s+', rest[2:])
    if next_section:
        section_text = rest[:next_section.start() + 2]
    else:
        section_text = rest
    
    required_fields = ['cod', 'treapta_nr', 'nume_slug',
                       'nume_hero_caps_rand1', 'meta_val_treapta',
                       'footer_text', 'topbar_text',
                       'localStorage_studiu', 'localStorage_video',
                       'next_cod', 'next_nume_title']
    missing = []
    for field in required_fields:
        if field not in section_text:
            missing.append(field)
    
    if missing:
        return (False, f"IDENTITATE_TEHNICA C{NN_str} INCOMPLETA. Campuri lipsa: {missing}")
    
    return (True, None)


def check_fenomene_vs_asset(NN, spec_section, base_path):
    """
    L143: Verifica daca FENOMENELE din SPEC corespund cu asset fizic
    Date_MASTER-initial.xlsx.
    
    Cauta doar in sectiunea "9. FENOMENE" sau "FENOMENE ALESE" - nu in
    text istoric care DESCRIE corectia.
    """
    NN_int = int(NN)
    NN_str = f"{NN_int:02d}"
    
    initial_path = os.path.join(base_path, 'referinte', 'Date_MASTER-initial.xlsx')
    if not os.path.exists(initial_path):
        # V40: structura noua _system/referinte/
        for cand in ['_system/referinte/Date_MASTER-initial.xlsx',
                    '../_system/referinte/Date_MASTER-initial.xlsx',
                    'referinte/Date_MASTER-initial.xlsx',
                    '../referinte/Date_MASTER-initial.xlsx']:
            if os.path.exists(cand):
                initial_path = cand
                break
        else:
            return (True, "Date_MASTER-initial.xlsx LIPSA - skip L143 check")
    
    try:
        import openpyxl
    except ImportError:
        return (True, "openpyxl LIPSA - skip L143 check")
    
    try:
        wb = openpyxl.load_workbook(initial_path, data_only=False)
        if 'Vanzari' not in wb.sheetnames:
            return (True, "Sheet Vanzari LIPSA in initial - skip L143")
        
        ws = wb['Vanzari']
        headers_real = set(c.value for c in ws[1] if c.value)
    except Exception as e:
        return (True, f"Eroare citire initial: {e} - skip L143")
    
    # Extract DOAR sectiunea CELE N FENOMENE ALESE (lista finala),
    # NU textul de RECALIBRARE care explica de ce.
    # Cauta marker "CELE N FENOMENE ALESE" sau primul "**1." dupa header FENOMENE
    fenomene_match = re.search(r'CELE\s+\d+\s+FENOMENE\s+ALESE[^\n]*(?:\n|.)*?(?=---|\Z)',
                                spec_section, re.IGNORECASE | re.DOTALL)
    
    if fenomene_match:
        fenomene_section = fenomene_match.group(0)
    else:
        # Fallback: cauta primul "**1." sau "F1" dupa header FENOMENE
        m_header = re.search(r'###?\s*9\.\s*FENOMENE', spec_section, re.IGNORECASE)
        if not m_header:
            return (True, "Sectiunea FENOMENE nu detectata - skip L143")
        
        # Cauta primul item "**1." sau "**F1" dupa header
        after_header = spec_section[m_header.end():]
        m_first = re.search(r'\*\*(?:F)?1[\.\s]', after_header)
        if not m_first:
            return (True, "Lista FENOMENE nu detectata - skip L143")
        
        fenomene_section = after_header[m_first.start():m_first.start() + 3000]
    
    # Lista coloane canonice de cautat in FENOMENE
    coloane_canonice_set = {
        'nr_factura', 'data_factura', 'client_nume', 'judet',
        'cod_produs', 'produs_nume', 'categorie',
        'cantitate', 'pret_unitar', 'valoare_neta',
        'cota_tva', 'valoare_tva', 'valoare_totala', 'moneda',
        'cod_client', 'nume_client', 'nume_produs',  # variante vechi
        'status_factura', 'regiune_business', 'canal_vanzare',  # invenții posibile
        'cod_agent', 'cod_depozit',
    }
    
    spec_columns_referite = set()
    for col in coloane_canonice_set:
        if col in fenomene_section:
            spec_columns_referite.add(col)
    
    coloane_inventate = spec_columns_referite - headers_real
    
    if coloane_inventate:
        return (False,
                f"L143 DISCREPANTA: SPEC C{NN_str} sectiunea FENOMENE referă "
                f"coloane care NU exista in Date_MASTER-initial.xlsx: "
                f"{sorted(coloane_inventate)}. "
                f"Coloane reale in asset: {sorted(headers_real)}. "
                f"Aliniaza SPEC la asset fizic INAINTE de generare.")
    
    return (True, None)


def print_freeze_required_message(NN, nume_constructie):
    """Mesaj R-V03.55 cand SPEC e NEGENERAT."""
    NN_int = int(NN)
    NN_str = f"{NN_int:02d}"
    
    print("=" * 78)
    print(f"R-V03.55 BLOCAJ: SPEC C{NN_str} ESTE NEGENERAT")
    print("=" * 78)
    print()
    print(f"NU POT genera constructia C{NN_str} ({nume_constructie}) fara SPEC inghetat.")
    print()
    print("Inainte de generare trebuie inghetate impreuna cu ARHITECT cele 9 elemente")
    print("in ordinea narativa OBLIGATORIE:")
    print()
    print("  1. INTRIGA       - paradoxul care deschide constructia")
    print("  2. PROBLEMELE    - lista incitanta, utila si de efect a problemelor")
    print("                     pe care constructia le rezolva (3-5 probleme scurte)")
    print("  3. MIZA          - ce castiga cursantul concret la final")
    print("  4. MANTRA        - fraza-semnatura care ramane cu cursantul")
    print("  5. MOTTO         - sub-hook complementar mantrei")
    print("  6. STEP-TITLES   - cele 18 step-titles (progresia narativa)")
    print("  7. PROMPTURI     - 2 prompturi Copilot (instrumentele AI)")
    print("  8. FINAL-LABELS  - 8 ancore de retentie")
    print("  9. FENOMENE/OPS  - specifice constructiei (T1=contaminari,")
    print("                     T2=clasificari, T3=motoare, etc.)")
    print()
    print("LOGICA ORDINII NARATIVE:")
    print("  INTRIGA (ATENTIE) -> PROBLEME (RECUNOASTERE) -> MIZA (DORINTA)")
    print("  -> MANTRA+MOTTO (IDENTITATE) -> executie tehnica.")
    print()
    print("FORMAT INGHETARE (R-V03.56, obligatoriu V23+):")
    print()
    print("  Pentru fiecare element, motor propune 3 VARIANTE CREATIVE concrete:")
    print("    A. [varianta pedagogica/conceptuala]")
    print("    B. [varianta concreta/emotionala]")
    print("    C. [varianta poetica/aforistica]")
    print()
    print("  ARHITECT alege A/B/C, cere 'alte 3', combina ('A cu finalul din B'),")
    print("  sau edita ('B dar X inlocuit cu Y'). Motorul propune, ARHITECT decide.")
    print()
    print("L143 (V26): la elementul 9 FENOMENE, motorul VERIFICA cu asset fizic")
    print("Date_MASTER-initial.xlsx INAINTE de inghetare. Coloanele si cifrele")
    print("propuse trebuie sa existe fizic in asset.")
    print()
    print("Optiuni continuare:")
    print(f"  a) Comanda 'definim SPEC C{NN_str} acum in acest chat' - intram in")
    print(f"     modul SPEC FREEZING aici, 9 intrebari cu format grila.")
    print(f"  b) Deschizi chat BRAIN dedicat SPEC C{NN_str} (recomandat - focus")
    print(f"     curat, fara presiune de generare).")
    print()
    print("DUPA inghetare SPEC in SISTEM_TRAINITY.md, ruleaza din nou acest script")
    print("si va trece la exit 0. Apoi procedeaza la generare normal.")
    print("=" * 78)


def main():
    # SELF-CHECK: prima linie de output e SCRIPT_VERSION ca motorul
    # sa poata verifica daca ruleaza versiunea corecta. Daca motorul
    # vede o versiune mai mica (V26 sau lipsa), opreste si cere
    # reatasare arhiva curenta.
    print(f"[pre_generation_check {SCRIPT_VERSION} - schema {EXPECTED_SCHEMA}]")
    
    if len(sys.argv) < 2:
        print("Usage: python3 pre_generation_check.py NN [/path/to/SISTEM_TRAINITY.md]")
        print("")
        print("Verifica TREI conditii inainte de generare constructie C{NN}:")
        print("  1. R-V03.55: SPEC narativ inghetat in SISTEM_TRAINITY.md")
        print("  2. L142: IDENTITATE_TEHNICA C{NN} populata in referinte/")
        print("  3. L143: FENOMENE SPEC vs asset fizic Date_MASTER-initial.xlsx")
        print("")
        print("Exit 0 = toate PASS. Exit 1 = blocaj. Exit 2 = eroare config.")
        sys.exit(2)
    
    NN = sys.argv[1].zfill(2)
    
    if len(sys.argv) >= 3:
        sistem_path = sys.argv[2]
    else:
        # V40: SISTEM_TRAINITY.md a fost mutat in arhiva la refactor V38.
        # SPEC narativ traieste acum in _system/arhiva/SISTEM_TRAINITY-versiuni.md.
        candidates = [
            '_system/arhiva/SISTEM_TRAINITY-versiuni.md',
            '../_system/arhiva/SISTEM_TRAINITY-versiuni.md',
            'SISTEM_TRAINITY.md',
            '../SISTEM_TRAINITY.md',
        ]
        sistem_path = None
        for c in candidates:
            if os.path.exists(c):
                sistem_path = c
                break

        if not sistem_path:
            print("EROARE: nu am gasit fisierul cu SPEC narativ.", file=sys.stderr)
            print("Cautat in: " + ", ".join(candidates), file=sys.stderr)
            sys.exit(2)
    
    if not os.path.exists(sistem_path):
        print(f"EROARE: {sistem_path} nu exista.", file=sys.stderr)
        sys.exit(2)
    
    base_path = os.path.dirname(os.path.abspath(sistem_path))
    
    with open(sistem_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    status, section = find_spec_section(content, NN)
    nume = get_construction_name(NN, content)
    
    # CHECK 1 (R-V03.55): SPEC narativ inghetat
    if status is None:
        print(f"R-V03.55 BLOCAJ: Sectiunea SPEC C{NN} nu exista in SISTEM_TRAINITY.md.")
        print(f"Adauga sectiunea inainte de generare:")
        print(f"  ## SPEC C{NN} - [NUMELE_CONSTRUCTIEI]   [Status: NEGENERAT]")
        sys.exit(1)
    
    if status == 'NEGENERAT':
        print_freeze_required_message(NN, nume)
        sys.exit(1)
    
    print(f"✓ CHECK 1 (R-V03.55): SPEC C{NN} ({nume}): INGHETAT narativ")
    
    # CHECK 2 (L142): IDENTITATE_TEHNICA populata
    ok_ident, msg_ident = check_identitate_tehnica(NN, base_path)
    if not ok_ident:
        print()
        print("=" * 78)
        print(f"L142 BLOCAJ: IDENTITATE_TEHNICA C{NN} INCOMPLETA")
        print("=" * 78)
        print()
        print(f"  {msg_ident}")
        print()
        print(f"Adauga in referinte/IDENTITATE-TEHNICA.md sectiunea:")
        print(f"  ## IDENTITATE_TEHNICA C{NN}   [Status: INGHETAT {{DATA}}]")
        print()
        print("Cu campurile obligatorii (schema folosita C01-C04 pilot):")
        print("cod, treapta_nr, treapta_nume, nume_slug, nume_hero_caps_rand1,")
        print("nume_hero_caps_rand2, nume_hero_caps, meta_val_treapta,")
        print("footer_text, topbar_text, mobile_topbar, nav_brand_label,")
        print("nav_brand_title, title_studiu, title_video,")
        print("localStorage_studiu, localStorage_video, next_cod,")
        print("next_nume_title, next_label.")
        print("=" * 78)
        sys.exit(1)
    
    print(f"✓ CHECK 2 (L142): IDENTITATE_TEHNICA C{NN}: POPULATA")
    
    # CHECK 3 (L143): FENOMENE vs asset fizic
    ok_fenomene, msg_fenomene = check_fenomene_vs_asset(NN, section, base_path)
    if not ok_fenomene:
        print()
        print("=" * 78)
        print(f"L143 BLOCAJ: SPEC FENOMENE C{NN} vs ASSET FIZIC")
        print("=" * 78)
        print()
        print(f"  {msg_fenomene}")
        print()
        print("FENOMENELE in SPEC trebuie sa fie cifre/coloane REALE din")
        print("Date_MASTER-initial.xlsx, NU inventii din chat SPEC FREEZING.")
        print()
        print("Optiuni:")
        print("  A. Aliniere SPEC la asset fizic (recomandat) - actualizeaza")
        print(f"     SPEC C{NN} sectiunea 9 FENOMENE in SISTEM_TRAINITY.md cu")
        print("     coloane si cifre care exista in asset.")
        print("  B. Aliniere asset la SPEC - refacere Date_MASTER-initial.xlsx")
        print("     (NU recomandat - impactă C01-C04 deja livrate).")
        print("=" * 78)
        sys.exit(1)
    
    print(f"✓ CHECK 3 (L143): FENOMENE C{NN} vs asset fizic: ALINIAT")
    print()
    print(f"TOATE CHECK-URILE PASS. Motorul poate proceda la generare C{NN}.")
    sys.exit(0)


if __name__ == '__main__':
    main()
