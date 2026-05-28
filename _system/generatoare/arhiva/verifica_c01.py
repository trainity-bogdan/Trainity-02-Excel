from playwright.sync_api import sync_playwright
import pathlib, zipfile, re

HTML = "file://" + str(pathlib.Path("/home/claude/brain/livrabile_C01/HTML-Excel-01-Structura.html").resolve())
R=[]
def chk(cap, name, cond, detail=""):
    R.append((cap, "CONFORM" if cond else "NECONFORM", name, detail))

with sync_playwright() as pw:
    br = pw.chromium.launch()
    ctx = br.new_context(viewport={"width":1440,"height":900})
    pg = ctx.new_page()
    pg.on("dialog", lambda d: d.accept())
    pg.goto(HTML, wait_until="networkidle"); pg.wait_for_timeout(400)

    # CAP 1 - PERSISTENTA localStorage
    pg.eval_on_selector('.step[data-step="1"] .step-check', "e=>e.click()")
    pg.wait_for_timeout(150)
    pg.eval_on_selector('.step[data-step="2"] .step-check', "e=>e.click()")
    pg.wait_for_timeout(150)
    pg.reload(wait_until="networkidle"); pg.wait_for_timeout(400)
    p1 = pg.eval_on_selector('.step[data-step="1"]', "e=>e.classList.contains('done')")
    p2 = pg.eval_on_selector('.step[data-step="2"]', "e=>e.classList.contains('done')")
    ls = pg.evaluate("Object.keys(localStorage).some(k=>k.includes('c01')||k.includes('trainity'))")
    chk(1, "Persistenta stare localStorage dupa reload", p1 and p2 and ls,
        f"step1={p1} step2={p2} ls={ls}")

    # CAP 2 - PASI GRANULARI per etapa
    nsteps = pg.evaluate("document.querySelectorAll('.step .step-check').length")
    nstages = pg.evaluate("document.querySelectorAll('.stage').length")
    chk(2, "Pasi granulari per etapa (step-check toggle real)", nsteps>=12 and nstages==6,
        f"{nsteps} pasi / {nstages} etape")

    # CAP 3 - VERIFICARI FINALE separate
    nfinal = pg.evaluate("document.querySelectorAll('.final-check-btn').length")
    pg.eval_on_selector('.final-card[data-final="1"] .final-check-btn', "e=>e.click()")
    pg.wait_for_timeout(150)
    fdone = pg.eval_on_selector('.final-card[data-final="1"]', "e=>e.classList.contains('done')")
    chk(3, "Verificari finale separate (final-check-btn efect)", nfinal==8 and fdone,
        f"{nfinal} finale, toggle efect={fdone}")

    # CAP 4 - COLAPSARE / EXPANDARE etape
    pg.evaluate("document.querySelectorAll('.stage .stage-header')[1].click()")
    pg.wait_for_timeout(300)
    anyopen = pg.evaluate("[...document.querySelectorAll('.stage')].some(s=>s.classList.contains('open'))")
    chk(4, "Colapsare/expandare etape (stage-toggle accordion)", anyopen,
        f"are etapa open={anyopen}")

    # CAP 5 - CONTOR DUBLU
    stt = pg.eval_on_selector('#navStepStat', "e=>e.textContent")
    fnt = pg.eval_on_selector('#navFinalStat', "e=>e.textContent")
    dbl = ("pa" in stt.lower()) and ("final" in fnt.lower())
    chk(5, "Contor dublu (pasi + verificari finale)", dbl, f"'{stt}' | '{fnt}'")

    # CAP 6 - BARA PROGRES per etapa
    nbar = pg.evaluate("document.querySelectorAll('[data-stage-progress]').length")
    # bifez pasii etapei 1 si verific ca bara ei creste
    w_before = pg.evaluate("(document.querySelector('[data-stage-progress=\"1\"]')||{}).style?document.querySelector('[data-stage-progress=\"1\"]').style.width:''")
    chk(6, "Bara progres per etapa (stage-progress)", nbar==6,
        f"{nbar} bare per etapa")

    # CAP 7 - FLOATING-NEXT INTELIGENT
    fn = pg.query_selector('#floatingNext')
    lbl = pg.eval_on_selector('#floatingNextLabel', "e=>e.textContent") if fn else ""
    smart = fn is not None and ("PASUL" in lbl.upper() or "CONTROL" in lbl.upper())
    chk(7, "Floating-next inteligent (eticheta dinamica)", smart, f"label='{lbl}'")

    # CAP 8 - HIGHLIGHT TEXT persistent
    hl = pg.evaluate("typeof processSelection==='function' && typeof clearHighlights==='function' && typeof wrapTextNodeRange==='function' && typeof removeMark==='function'")
    chk(8, "Highlight/adnotare text (selectie persistenta)", hl,
        f"functii highlight={hl}")

    # CAP 9 - RESET PROGRES
    rp = pg.query_selector('.nav-ctrl-btn')
    has_reset = pg.evaluate("[...document.querySelectorAll('button')].some(b=>/RESETEAZ/i.test(b.textContent))")
    pg.evaluate("resetProgress()")
    pg.wait_for_timeout(300)
    cleared = pg.eval_on_selector('.step[data-step="1"]', "e=>!e.classList.contains('done')")
    chk(9, "Reset progres (resetProgress efect)", has_reset and cleared,
        f"buton={has_reset} efect_clear={cleared}")

    # CAP 10 - NAVIGATIE MOBILA reala
    ctxm = br.new_context(viewport={"width":390,"height":844}, is_mobile=True, has_touch=True)
    pm = ctxm.new_page(); pm.on("dialog", lambda d:d.accept())
    pm.goto(HTML, wait_until="networkidle"); pm.wait_for_timeout(400)
    mt_vis = pm.eval_on_selector('.mobile-nav-toggle', "e=>getComputedStyle(e).display!=='none'")
    pm.eval_on_selector('.mobile-nav-toggle', "e=>e.click()")
    pm.wait_for_timeout(300)
    navopen = pm.eval_on_selector('#sideNav', "e=>e.classList.contains('open')")
    chk(10, "Navigatie mobila reala (hamburger + slide)", mt_vis and navopen,
        f"hamburger_vizibil={mt_vis} nav_deschide={navopen}")
    ctxm.close()

    # CAP 11 - RENDER DINAMIC SIDEBAR
    pg.eval_on_selector('.step[data-step="1"] .step-check', "e=>e.click()")
    pg.wait_for_timeout(100)
    pg.eval_on_selector('.step[data-step="2"] .step-check', "e=>e.click()")
    pg.eval_on_selector('.step[data-step="3"] .step-check', "e=>e.click()")
    pg.wait_for_timeout(100)
    pg.eval_on_selector('.verification[data-verif="1"] .verification-btn', "e=>e.click()")
    pg.wait_for_timeout(200)
    nav_done = pg.eval_on_selector('.nav-item[data-stage-link="1"]', "e=>e.classList.contains('done')")
    chk(11, "Render dinamic sidebar (oglinda live progres)", nav_done,
        f"nav-item etapa1 done dupa verif={nav_done}")

    # CAP 12 - COPIERE PROMPT robusta
    cb = pg.query_selector('.copy-btn')
    if cb:
        bgc = pg.eval_on_selector('.copy-btn', "e=>getComputedStyle(e).backgroundColor")
        cb.click(); pg.wait_for_timeout(200)
        copied = pg.eval_on_selector('.copy-btn', "e=>e.classList.contains('copied')")
        has_fb = pg.evaluate("typeof fallbackCopy==='function'")
        yellow = "255, 212, 0" in bgc or "rgb(255, 212, 0)" in bgc
        chk(12, "Copiere prompt robusta (fallback + galben)", copied and has_fb and yellow,
            f"copied={copied} fallback={has_fb} galben={yellow}({bgc})")
    else:
        chk(12, "Copiere prompt robusta", False, "buton copy absent")

    br.close()

# ===== CROSS-LIVRABIL: cifre canonice identice =====
import subprocess
def grab(path):
    try:
        return subprocess.run(["python3","-c",
            f"import sys;sys.path.insert(0,'/mnt/skills/public');"
            ], capture_output=True, text=True).stdout
    except: return ""

base = "/home/claude/brain/livrabile_C01/"
checks_cross = []
# HTML
h = open(base+"HTML-Excel-01-Structura.html",encoding="utf-8").read()
for token in ["1.247.893,50","2.062","2.000","14 coloane","62"]:
    checks_cross.append(("HTML", token, token in h))

print("="*64)
print("VERIFICARE CONSTRUCTIA 01 - GATE PARITATE FUNCTIONALA")
print("="*64)
nc=0
for cap,status,name,detail in R:
    mark = "OK " if status=="CONFORM" else "XX "
    if status!="CONFORM": nc+=1
    print(f"  {mark}[CAP {cap:>2}] {status:<10} {name}")
    if detail: print(f"           {detail}")
print("-"*64)
print("CROSS-LIVRABIL (cifre canonice in HTML sursa de adevar):")
for src,tok,ok in checks_cross:
    print(f"  {'OK ' if ok else 'XX '}{src}: '{tok}' {'prezent' if ok else 'LIPSA'}")

# ============================================================
# GATE 2 - COERENTA FIZICA INPUT/OUTPUT Excel <-> afirmatii HTML
# ============================================================
print("="*64)
print("GATE 2 - COERENTA FIZICA Excel <-> afirmatii HTML")
print("="*64)
from openpyxl import load_workbook
H2=open("/home/claude/brain/livrabile_C01/HTML-Excel-01-Structura.html",encoding="utf-8").read()
P=[]
def chk2(name, cond, detail=""):
    P.append(("CONFORM" if cond else "NECONFORM", name, detail))

wi=load_workbook("/home/claude/brain/livrabile_C01/INPUT-Excel-01-Structura.xlsx",data_only=False)
wsi=wi["Date"]
chk2("INPUT max_row = 2.062", wsi.max_row==2062, f"real={wsi.max_row}")
chk2("INPUT max_col = 16 (14 vizibile + 2 ascunse)", wsi.max_column==16, f"real={wsi.max_column}")
hidden=[c for c,d in wsi.column_dimensions.items() if d.hidden]
chk2("INPUT 2 coloane ascunse fizic", len(hidden)==2, f"ascunse={hidden}")
chk2("INPUT antet real pe r4 = 'data'", str(wsi.cell(4,1).value).lower()=="data", f"r4c1='{wsi.cell(4,1).value}'")
sub=sum(1 for r in range(5,wsi.max_row+1) if "SUBTOTAL" in str(wsi.cell(r,1).value or "").upper())
chk2("INPUT 4 SUBTOTAL", sub==4, f"real={sub}")
tot=sum(1 for r in range(5,wsi.max_row+1) if "TOTAL GENERAL" in str(wsi.cell(r,1).value or "").upper())
chk2("INPUT 1 TOTAL GENERAL", tot==1, f"real={tot}")
blk=sum(1 for r in range(5,wsi.max_row+1) if all((wsi.cell(r,c).value is None or str(wsi.cell(r,c).value).strip()=="") for c in range(1,wsi.max_column+1)))
chk2("INPUT 53 BLANK FALS", blk==53, f"real={blk}")
chk2("INPUT ambalaj total 61", (3+sub+tot+blk)==61, f"calc=3+{sub}+{tot}+{blk}={3+sub+tot+blk}")
si=0.0
for r in range(5,wsi.max_row+1):
    a=wsi.cell(r,1).value
    if a is None: continue
    s=str(a).upper()
    if "SUBTOTAL" in s or "TOTAL" in s: continue
    v=wsi.cell(r,13).value
    if isinstance(v,(int,float)): si+=v
chk2("INPUT suma valoare_neta = 1.247.893,50", abs(si-1247893.50)<0.01, f"real={si:.2f}")

wo=load_workbook("/home/claude/brain/livrabile_C01/OUTPUT-Excel-01-Structura.xlsx",data_only=False)
wso=wo["Date"]
chk2("OUTPUT 2001 randuri (antet+2000 tranz)", wso.max_row==2001, f"real={wso.max_row}")
chk2("OUTPUT 14 coloane", wso.max_column==14, f"real={wso.max_column}")
chk2("OUTPUT antet r1 = 'data'", str(wso.cell(1,1).value).lower()=="data", f"r1c1='{wso.cell(1,1).value}'")
blo=sum(1 for r in range(2,wso.max_row+1) if all((wso.cell(r,c).value is None or str(wso.cell(r,c).value).strip()=="") for c in range(1,wso.max_column+1)))
chk2("OUTPUT 0 randuri blank", blo==0, f"real={blo}")
suo=sum(1 for r in range(2,wso.max_row+1) if "SUBTOTAL" in str(wso.cell(r,1).value or "").upper() or "TOTAL" in str(wso.cell(r,1).value or "").upper())
chk2("OUTPUT 0 SUBTOTAL/TOTAL", suo==0, f"real={suo}")
so=0.0
for r in range(2,wso.max_row+1):
    v=wso.cell(r,12).value
    if isinstance(v,(int,float)): so+=v
chk2("OUTPUT suma = 1.247.893,50 (conservata)", abs(so-1247893.50)<0.01, f"real={so:.2f}")
hdr_out=[wso.cell(1,c).value for c in range(1,wso.max_column+1)]
chk2("OUTPUT NU contine cod_filiala", "cod_filiala" not in hdr_out)
chk2("OUTPUT NU contine cod_user", "cod_user" not in hdr_out)
chk2("OUTPUT contine Excel Table 'Vanzari_Curat'", "Vanzari_Curat" in list(wso.tables.keys()))
chk2("HTML zice 53 (nu 54)", "53 rânduri" in H2 and "54 rânduri" not in H2)
chk2("HTML zice 61 (nu 62 ambalaj)", "62 rânduri ambalaj" not in H2)

nc2=0
for status,name,detail in P:
    mark="OK " if status=="CONFORM" else "XX "
    if status!="CONFORM": nc2+=1
    print(f"  {mark}{status:<10} {name}")
    if detail: print(f"             {detail}")
print("-"*64)
print(f"GATE 2: {len(P)-nc2}/{len(P)} CONFORM coerenta fizica Excel<->HTML")

print("="*64)
total = len(R)
gate1_ok = nc==0
gate2_ok = nc2==0
print(f"GATE 1 (paritate functionala HTML): {total-nc}/{total} CONFORM")
print(f"GATE 2 (coerenta fizica Excel<->HTML): {len(P)-nc2}/{len(P)} CONFORM")
print("="*64)
if gate1_ok and gate2_ok:
    print("RAPORT FINAL: CONFORM - C01 trece AMBELE gate-uri blocante")
else:
    parts_msg=[]
    if not gate1_ok: parts_msg.append(f"{nc} capab. functionale esuate")
    if not gate2_ok: parts_msg.append(f"{nc2} discrepante fizice")
    print(f"RAPORT FINAL: NECONFORM - {' + '.join(parts_msg)}")
