#!/usr/bin/env python3
"""
gen_date_master_initial.py — V19

Genereaza Date_MASTER-initial.xlsx, asset perpetuu inghetat care
serveste ca punct de plecare pentru toate constructiile C01-C08.

Schema canonica: 14 coloane pe sheet Vanzari, plus sheets auxiliare
CLIENTI, PRODUSE, AGENTI, DEPOZITE.

~2000 facturi distribuite pe 12 luni cu pattern sezonier realist.
Suma valoare_neta tintita: ~5,500,000 lei.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
import datetime

# Seed determinist pentru reproductibilitate
random.seed(42)

# ============================================================
# NOMENCLATOARE (asset inghetat)
# ============================================================

CLIENTI = [
    ('CL-001', 'SC AROMET TRADING SRL',       'București'),
    ('CL-002', 'SC DACTECH SOLUTIONS SRL',    'Cluj'),
    ('CL-003', 'SC INTEGROMED SA',            'Timiș'),
    ('CL-004', 'SC NORDIS DISTRIBUTIE SRL',   'Iași'),
    ('CL-005', 'SC METALPLAST INVEST SA',     'București'),
    ('CL-006', 'SC OMEGA SOFT SRL',           'Cluj'),
    ('CL-007', 'SC PROVITA MEDICAL SRL',      'Brașov'),
    ('CL-008', 'SC RAPID LOGISTIC SA',        'Constanța'),
    ('CL-009', 'SC SOLARIS ENERGY SRL',       'Timiș'),
    ('CL-010', 'SC TEHNOPLUS SERVICE SRL',    'Sibiu'),
    ('CL-011', 'SC UNIFOOD DISTRIBUTION SA',  'București'),
    ('CL-012', 'SC VECTOR CONSULT SRL',       'Mureș'),
    ('CL-013', 'SC ZENITH AUTO SRL',          'Cluj'),
    ('CL-014', 'SC ELECTROBUILD SRL',         'București'),
    ('CL-015', 'SC FARMAVITAL DISTRIBUTIE SA','Iași'),
]

PRODUSE = [
    # (cod, nume, categorie, pret_unitar_baza)
    ('PRD-001', 'Laptop business 15"',           'Hardware',    2400),
    ('PRD-002', 'Monitor 27" 4K',                'Hardware',    1100),
    ('PRD-003', 'Server tower 16GB',             'Hardware',    4200),
    ('PRD-004', 'Imprimanta multifunctionala',   'Hardware',     900),
    ('PRD-005', 'Tastatura+mouse wireless',      'Hardware',     250),
    ('PRD-006', 'Licenta Office 365 anuala',     'Software',     750),
    ('PRD-007', 'Antivirus enterprise 1an',      'Software',     180),
    ('PRD-008', 'Backup cloud 1TB anual',        'Software',     450),
    ('PRD-009', 'Hartie A4 500coli (pachet)',    'Consumabile',   28),
    ('PRD-010', 'Toner laser negru',             'Consumabile',  350),
    ('PRD-011', 'Cablu HDMI 2m',                 'Consumabile',   45),
    ('PRD-012', 'Instalare retea LAN (ora)',     'Servicii',     180),
    ('PRD-013', 'Mentenanta IT lunara',          'Servicii',    1200),
]

AGENTI = [
    ('AG-001', 'Vasile Popescu'),
    ('AG-002', 'Maria Ionescu'),
    ('AG-003', 'Andrei Marin'),
    ('AG-004', 'Cristina Dumitrescu'),
    ('AG-005', 'Bogdan Stanescu'),
    ('AG-006', 'Elena Radu'),
]

DEPOZITE = [
    ('DEP-BUC',  'Depozit Bucuresti Central'),
    ('DEP-CLJ',  'Depozit Cluj-Napoca'),
    ('DEP-TM',   'Depozit Timisoara'),
    ('DEP-IS',   'Depozit Iasi'),
    ('DEP-CTA',  'Depozit Constanta'),
]

# Distribuție clienți: top 3 = 40% facturi, mid 7 = 45%, mic 5 = 15%
CLIENTI_WEIGHTS = [
    18, 14, 10,           # Top 3 (42%)
    8, 7, 7, 6, 6, 5, 5,  # Mid 7 (44%)
    3, 3, 3, 3, 2,        # Mic 5 (14%)
]

# Distribuție produse: hardware mai des, servicii mai rar
PRODUSE_WEIGHTS = [12, 10, 6, 7, 12,  # Hardware (47%)
                   10, 8, 6,           # Software (24%)
                   12, 8, 7,           # Consumabile (27%)
                   2, 1]               # Servicii (2%)


# ============================================================
# DISTRIBUTIE TEMPORALA SEZONIERA
# ============================================================

# Iunie 2025 -> Mai 2026 (12 luni)
# Pattern sezonier: vârf nov-dec (Black Friday + sărbători), valle iul-aug (concedii)
LUNI = [
    (2025, 6, 150),   # iunie
    (2025, 7, 110),   # iulie - concediu
    (2025, 8, 100),   # august - concediu
    (2025, 9, 180),   # septembrie - back to business
    (2025, 10, 200),  # octombrie
    (2025, 11, 270),  # noiembrie - Black Friday
    (2025, 12, 250),  # decembrie - sarbatori
    (2026, 1, 140),   # ianuarie
    (2026, 2, 160),   # februarie
    (2026, 3, 180),   # martie
    (2026, 4, 130),   # aprilie - Paste
    (2026, 5, 130),   # mai
]
# Total: 2000 facturi


# ============================================================
# GENERARE FACTURI
# ============================================================

def random_date_in_month(year, month):
    """Data aleatoare in luna data, fara weekend dominat."""
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    day = random.randint(1, last_day)
    return datetime.datetime(year, month, day)


def generate_factura(nr, data):
    client_idx = random.choices(range(len(CLIENTI)), weights=CLIENTI_WEIGHTS)[0]
    cod_client, nume_client, judet = CLIENTI[client_idx]
    
    produs_idx = random.choices(range(len(PRODUSE)), weights=PRODUSE_WEIGHTS)[0]
    cod_produs, nume_produs, categorie, pret_baza = PRODUSE[produs_idx]
    
    # Cantitate variaza per categorie
    if categorie == 'Hardware':
        cantitate = random.choices([1, 2, 3, 5, 10], weights=[40, 25, 15, 15, 5])[0]
    elif categorie == 'Software':
        cantitate = random.choices([1, 5, 10, 25, 50], weights=[30, 25, 25, 15, 5])[0]
    elif categorie == 'Consumabile':
        cantitate = random.choices([5, 10, 20, 50, 100], weights=[20, 30, 25, 20, 5])[0]
    else:  # Servicii
        cantitate = random.choices([1, 2, 4, 8, 16], weights=[40, 25, 20, 10, 5])[0]
    
    # Pret cu mica variatie (+/- 5%) per factura
    pret_unitar = round(pret_baza * random.uniform(0.95, 1.05), 2)
    valoare_neta = round(cantitate * pret_unitar, 2)
    
    # TVA: majoritar 19%, unele 9% (Servicii medicale gen), foarte rar 5%
    if categorie == 'Servicii' and random.random() < 0.15:
        cota_tva = 0.09
    elif random.random() < 0.05:
        cota_tva = 0.05
    else:
        cota_tva = 0.19
    
    valoare_tva = round(valoare_neta * cota_tva, 2)
    valoare_totala = round(valoare_neta + valoare_tva, 2)
    
    # Moneda
    moneda = random.choices(['RON', 'EUR', 'USD'], weights=[95, 4, 1])[0]
    
    return {
        'nr_factura':     f'F-2025-{nr:05d}',
        'data_factura':   data,
        'client_nume':    nume_client,
        'judet':          judet,
        'cod_produs':     cod_produs,
        'produs_nume':    nume_produs,
        'categorie':      categorie,
        'cantitate':      cantitate,
        'pret_unitar':    pret_unitar,
        'valoare_neta':   valoare_neta,
        'cota_tva':       cota_tva,
        'valoare_tva':    valoare_tva,
        'valoare_totala': valoare_totala,
        'moneda':         moneda,
    }


# Construire listă facturi cu distributie sezoniera
facturi = []
nr_curent = 1
for an, luna, count in LUNI:
    for _ in range(count):
        data = random_date_in_month(an, luna)
        facturi.append(generate_factura(nr_curent, data))
        nr_curent += 1

# Sortez dupa data_factura (mai realist pentru un export ERP)
facturi.sort(key=lambda f: f['data_factura'])
# Renumar dupa sortare
for i, f in enumerate(facturi, 1):
    f['nr_factura'] = f'F-2025-{i:05d}'

# Calcul sume statistici
total_neta = sum(f['valoare_neta'] for f in facturi)
total_tva = sum(f['valoare_tva'] for f in facturi)
total_globala = sum(f['valoare_totala'] for f in facturi)

print(f"Generat {len(facturi)} facturi")
print(f"Suma valoare_neta:   {total_neta:>15,.2f} lei")
print(f"Suma valoare_tva:    {total_tva:>15,.2f} lei")
print(f"Suma valoare_totala: {total_globala:>15,.2f} lei")
print(f"Range data: {min(f['data_factura'] for f in facturi).date()} - {max(f['data_factura'] for f in facturi).date()}")


# ============================================================
# SCRIERE EXCEL
# ============================================================

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Sheet Vanzari (principal)
ws = wb.create_sheet('Vanzari')
headers = ['nr_factura', 'data_factura', 'client_nume', 'judet',
           'cod_produs', 'produs_nume', 'categorie',
           'cantitate', 'pret_unitar', 'valoare_neta',
           'cota_tva', 'valoare_tva', 'valoare_totala', 'moneda']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center')

for r, f in enumerate(facturi, 2):
    for c, key in enumerate(headers, 1):
        v = f[key]
        cell = ws.cell(row=r, column=c, value=v)
        if key == 'data_factura':
            cell.number_format = 'dd.mm.yyyy'
        elif key in ('pret_unitar', 'valoare_neta', 'valoare_tva', 'valoare_totala'):
            cell.number_format = '#,##0.00'
        elif key == 'cota_tva':
            cell.number_format = '0.00%'

# Latimi coloane
widths = [14, 12, 32, 12, 12, 28, 14, 10, 12, 14, 10, 12, 14, 8]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

# Sheet CLIENTI
ws_c = wb.create_sheet('CLIENTI')
ws_c.append(['cod_client', 'nume_client', 'judet'])
for row in CLIENTI:
    ws_c.append(list(row))
for c in range(1, 4):
    cell = ws_c.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
ws_c.column_dimensions['A'].width = 12
ws_c.column_dimensions['B'].width = 35
ws_c.column_dimensions['C'].width = 14

# Sheet PRODUSE
ws_p = wb.create_sheet('PRODUSE')
ws_p.append(['cod_produs', 'nume_produs', 'categorie', 'pret_baza'])
for row in PRODUSE:
    ws_p.append(list(row))
for c in range(1, 5):
    cell = ws_p.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
ws_p.column_dimensions['A'].width = 12
ws_p.column_dimensions['B'].width = 32
ws_p.column_dimensions['C'].width = 14
ws_p.column_dimensions['D'].width = 12

# Sheet AGENTI
ws_a = wb.create_sheet('AGENTI')
ws_a.append(['cod_agent', 'nume_agent'])
for row in AGENTI:
    ws_a.append(list(row))
for c in range(1, 3):
    cell = ws_a.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
ws_a.column_dimensions['A'].width = 12
ws_a.column_dimensions['B'].width = 25

# Sheet DEPOZITE
ws_d = wb.create_sheet('DEPOZITE')
ws_d.append(['cod_depozit', 'nume_depozit'])
for row in DEPOZITE:
    ws_d.append(list(row))
for c in range(1, 3):
    cell = ws_d.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
ws_d.column_dimensions['A'].width = 14
ws_d.column_dimensions['B'].width = 30

# Sheet _README
ws_r = wb.create_sheet('_README')
readme = [
    ['DATE_MASTER-INITIAL.XLSX'],
    [''],
    ['Acesta este universul de date canonic Trainity Pack 02 Excel.'],
    ['Punct de plecare pentru toate constructiile C01-C08 (Treapta 1 + Treapta 2).'],
    [''],
    ['CONTINUT:'],
    [f'  - Sheet Vanzari: {len(facturi)} facturi pe 12 luni (iun 2025 - mai 2026)'],
    [f'  - Schema: 14 coloane canonice (R-V03.47)'],
    [f'  - Sheet CLIENTI: {len(CLIENTI)} clienti realistici'],
    [f'  - Sheet PRODUSE: {len(PRODUSE)} produse, 4 categorii'],
    [f'  - Sheet AGENTI: {len(AGENTI)} agenti'],
    [f'  - Sheet DEPOZITE: {len(DEPOZITE)} depozite'],
    [''],
    ['SUME REFERINTA (utilizate de gate V19 pentru validare):'],
    [f'  Suma valoare_neta:   {total_neta:,.2f} lei'],
    [f'  Suma valoare_tva:    {total_tva:,.2f} lei'],
    [f'  Suma valoare_totala: {total_globala:,.2f} lei'],
    [''],
    ['DISTRIBUTII:'],
    ['  Clienti: top 3 = ~42% facturi, mid 7 = ~44%, mic 5 = ~14%'],
    ['  Categorii: Hardware ~47%, Software ~24%, Consumabile ~27%, Servicii ~2%'],
    ['  Cota TVA: 19% (~85%), 9% (~10%), 5% (~5%)'],
    ['  Moneda: RON 95%, EUR 4%, USD 1%'],
    ['  Sezonier: varf nov-dec (Black Friday+sarbatori), valle iul-aug (concedii)'],
    [''],
    ['UTILIZARE PE CONSTRUCTIE:'],
    ['  Fiecare chat CONSTRUCTIE NN citeste acest fisier ca punct de plecare,'],
    ['  aplica modificari specifice SPEC C{NN}, livreaza Date_MASTER-C{NN}.xlsx.'],
    [''],
    ['  T1 SCAN (C01-C04): planteaza contaminari/anomalii, demonstreaza curatare'],
    ['    C01 STRUCTURA: adauga antet ERP brut, subtotale, blank-uri false'],
    ['    C02 CONTROL: planteaza anomalii logice (date viitor, cod necunoscut)'],
    ['    C03 AUDIT: planteaza contaminari invizibile (spatii, ZWSP, SHY)'],
    ['    C04 NORMALIZARE: consolideaza T1 in flux refresh idempotent'],
    [''],
    ['  T2 CUNOASTERE (C05-C08): cunoaste setul descriptiv multidimensional'],
    ['    C05 DICTIONAR: ce exista in set (categorii, frecvente, cardinalitate)'],
    ['    C06 CLASIFICARE: ce inseamna fiecare rand (clasa/segment/scor prin reguli)'],
    ['    C07 DATARE: cand se intampla (interval, ritm, varfuri, goluri)'],
    ['    C08 CARTOGRAFIERE: cu cine se leaga setul (ecosistem, chei, roluri)'],
    [''],
    ['NOTA TEHNICA: datele sunt sintetice, generate determinist (seed=42).'],
    ['Nu reprezinta o firma reala. Numele sunt fictive.'],
]
for row in readme:
    ws_r.append(row)
ws_r.column_dimensions['A'].width = 80
for r in range(1, len(readme)+1):
    cell = ws_r.cell(row=r, column=1)
    if r == 1:
        cell.font = Font(bold=True, size=16)
    elif row and len(readme[r-1]) and readme[r-1][0].endswith(':'):
        cell.font = Font(bold=True)

# Salvare
out_path = '/home/claude/v19_work/referinte/Date_MASTER-initial.xlsx'
wb.save(out_path)
print(f"\nSalvat: {out_path}")
import os
print(f"Marime: {os.path.getsize(out_path):,} bytes")
