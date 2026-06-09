#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c15.py - Date_MASTER-C15.xlsx (T4 · SINTETIZARE).

Nume fisier = Date_MASTER-C15.xlsx (cerut de gate B2).

C15 FORMULEAZA mesajul esential al paginii compuse (produsa la C14). Verb: a SINTETIZA.
Axa: SINTEZA, nu rezumat. Output = o propozitie (headline/so-what), NU layout (=C14),
NU decizie (=C16), NU analiza noua (=T3/C12). SINTEZA != REZUMAT: rezumatul scurteaza tot
proportional; sinteza alege singurul mesaj care conteaza.

Strategie (continuare compozitionala a Date_MASTER-C14, mandat generare C15): deschid
Date_MASTER-C14.xlsx (pastreaza TOATE foile, formulele si stilurile), rescriu START
pentru C15 si ADAUG foaia 'Sinteza' = STRATUL MESAJ. Vanzari ramane neatins -> suma
conservata (R-V02.14). Stratul MESAJ CITESTE din pagina (valori deja calculate in amonte,
in T3/C13), nu produce pagina, nu calculeaza insight, nu descopera cauza, nu recomanda.

Uz: python3 c15/build_date_master_c15.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import os

SRC = 'c14/Date_MASTER-C14.xlsx'
OUT = 'c15/Date_MASTER-C15.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def main():
    os.makedirs('c15', exist_ok=True)

    # --- pass valori (data_only): CITIM raspunsul deja produs in amonte (T3/C13) ---
    # NU calculam aici un insight nou; doar preluam valoarea existenta ca sa o
    # REFORMULAM intr-o propozitie (stratul MESAJ citeste, nu produce).
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)

    cat_val = defaultdict(float)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r)
    cats = sorted(cat_val.items(), key=lambda x: -x[1])
    top_cat, top_val = cats[0]
    second_cat, second_val = cats[1]
    dif_pct = round(100 * (top_val - second_val) / top_val, 1)

    # --- pass formule (pastreaza tot C14 intact) ---
    out = openpyxl.load_workbook(SRC)

    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C15 · SINTETIZARE · formularea mesajului esential al paginii de raport'],
        [],
        ['Ideea mare:'],
        ['Treapta de compunere a dat o pagina coerenta: focar, ierarhie, traseu de citire.'],
        ['Dar o pagina arata; nu spune. Decidentul o deschide si intreaba "si ce-i cu asta?".'],
        ['C15 extrage si FORMULEAZA mesajul esential: o singura fraza pe care pagina o dovedeste.'],
        ['Nu o cifra noua, nu o cauza noua (acelea vin gata din analiza), ci mesajul.'],
        [],
        ['AHA: O pagina arata. O sinteza spune.'],
        ['MANTRA: Nu rezumam. Sintetizam.'],
        ['MOTTO: Dintr-o privire, mesajul.'],
        ['FORMULA: Rezumatul scurteaza tot. Sinteza spune ce conteaza.'],
        [],
        ['SINTEZA != REZUMAT:'],
        ['  rezumatul scurteaza tot proportional (automatizabil);'],
        ['  sinteza alege si enunta SINGURUL mesaj care conteaza pentru aceasta decizie si audienta.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins fata de C14, suma conservata)'],
        ['  Compunere      pagina compusa mostenita (focar, ierarhie, traseu) = pagina-dovada'],
        ['  Sinteza        STRATUL MESAJ: headline, so-what, legatura la pagina-dovada, reformulare'],
        [],
        ['Exercitiul:'],
        ['  1. Pleci de la pagina compusa (mostenita), care arata corect dar nu spune nimic.'],
        ['  2. Rezumatul automat (AI) scurteaza tot; vezi ca nu e inca mesajul.'],
        ['  3. Formulezi headline-ul: singura afirmatie pe care pagina o dovedeste.'],
        ['  4. Testezi: "fara asta, decidentul hotaraste la fel?" Daca da, nu e mesajul.'],
        ['  5. Cand datele se schimba in amonte, REFORMULEZI mesajul (verbal), nu recalculezi.'],
        [],
        ['C15 doar FORMULEAZA mesajul. Nu calculeaza (T3), nu rearanjeaza pagina (C14),'],
        ['nu livreaza decizia (C16). C15 spune ce conteaza; incadrarea pentru decizie vine la C16.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 14, 18, 23):
        ws['A%d' % rr].font = SECT

    # --- foaia Sinteza (STRATUL MESAJ: citeste din pagina, nu produce pagina) ---
    wi = out.create_sheet('Sinteza')
    A = wi.append
    A(['SINTEZA C15 · stratul MESAJ · formularea mesajului esential al paginii compuse'])
    A([])

    A(['1) PAGINA PRIMITA · de la treapta de compunere (pagina-dovada, nerearanjata)'])
    A(['element', 'continut'])
    A(['pagina compusa', 'focar, ierarhie, traseu de citire (mostenite de la C14)'])
    A(['raspunsul venit din analiza (T3), de reformulat ca mesaj',
       '%s conduce, cu %.1f%% peste %s' % (top_cat, dif_pct, second_cat)])
    A(['stare', 'pagina ARATA corect; inca nu SPUNE nimic'])
    A([])

    r2 = wi.max_row + 1
    A(['2) REZUMAT vs SINTEZA · aceeasi pagina, doua iesiri'])
    A(['tip', 'ce face', 'rezultat'])
    A(['rezumat (automat, AI)', 'scurteaza tot proportional',
       'lista de tot ce e pe pagina, fara ierarhie de sens'])
    A(['sinteza (om)', 'alege singurul mesaj care conteaza',
       'o singura fraza pe care pagina o dovedeste'])
    A([])

    r3 = wi.max_row + 1
    A(['3) MESAJUL ESENTIAL · headline + so-what (text formulat de cursant, nu calculat)'])
    A(['camp', 'continut (exemplu, fara cifre statice in raportul vizual)'])
    A(['headline (o fraza)',
       'In trimestrul analizat, %s duce rezultatul; restul categoriilor raman in urma.' % top_cat])
    A(['so-what (de ce conteaza)',
       'Daca o singura categorie poarta rezultatul, acolo se uita decidentul intai.'])
    A(['legatura la pagina-dovada', 'focarul paginii: clasamentul categoriilor, cu %s in frunte' % top_cat])
    A(['NOTA', 'cifra exacta traieste in pagina/Excel; mesajul o reformuleaza, nu o re-calculeaza'])
    A([])

    r4 = wi.max_row + 1
    A(['4) TESTUL "si ce-i cu asta?" · mesaj care muta decizia vs zgomot'])
    A(['fraza candidata', 'trece testul?', 'de ce'])
    A(['"Pagina contine 7 categorii si 4 indicatori."', 'NU', 'descrie, nu spune ce conteaza (e rezumat)'])
    A(['"O singura categorie duce rezultatul."', 'DA', 'fara ea, decidentul s-ar uita altundeva'])
    A(['filtru', 'fara aceasta fraza, decidentul hotaraste la fel? Daca da, nu e mesajul.', ''])
    A([])

    r5 = wi.max_row + 1
    A(['5) REFORMULARE (E5) · mesajul se adapteaza la schimbare (verbal, NU recalcul)'])
    A(['situatie', 'ce faci'])
    A(['datele s-au actualizat in amonte (model/T3)', 'mesajul vechi nu mai e exact'])
    A(['actiunea C15', 'reformulezi headline-ul (refaci sinteza); NU recalculezi nimic'])
    A(['cifrele', 'raman ale modelului; se schimba in amonte, nu in stratul MESAJ'])
    A(['criteriu pastrat', 'noul mesaj ramane o singura afirmatie sustinuta de pagina: ce conteaza'])
    A([])

    r6 = wi.max_row + 1
    A(['6) HANDOFF · C15 formuleaza mesajul, C16 il incadreaza pentru decizie'])
    A(['input de la treapta de compunere', 'pagina coerenta: focar, ierarhie, traseu'])
    A(['output C15', 'mesajul esential (o fraza) + so-what + legatura la pagina-dovada'])
    A(['predat catre treapta de livrare', 'incadrarea deciziei: ce e de hotarat, pachet pentru decident'])
    A(['C15 NU face', 'cifra/cauza noua (T3), rearanjare pagina (C14), recomandare/livrare (C16)'])
    A(['regula de aur', 'mesajul CITESTE din pagina, nu PRODUCE pagina'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wi['A1'].font = TITLE
    for rr in (3, r2, r3, r4, r5, r6):
        wi['A%d' % rr].font = SECT
    hdr(wi, 4)
    hdr(wi, r2 + 1)
    hdr(wi, r3 + 1)
    hdr(wi, r4 + 1)
    hdr(wi, r5 + 1)

    # Sinteza imediat dupa Compunere in ordinea foilor
    order = [s for s in out.sheetnames if s not in ('START', 'Sinteza')]
    if 'Compunere' in order:
        i = order.index('Compunere') + 1
    else:
        i = len(order)
    new_order = ['START'] + order[:i] + ['Sinteza'] + order[i:]
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  top_cat=%s (%.2f) | a doua=%s (%.2f) | dif=%.1f%%'
          % (top_cat, top_val, second_cat, second_val, dif_pct))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    print('  OK: suma conservata, Sinteza (strat MESAJ) adaugata, continuare din C14')


if __name__ == '__main__':
    main()
