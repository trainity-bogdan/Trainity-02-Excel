#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c17.py - Date_MASTER-C17.xlsx (T5 · SISTEMATIZARE), deschide T5.

Nume fisier = Date_MASTER-C17.xlsx (cerut de gate B2).

C17 = SISTEMATIZAREA. Verb: a SISTEMATIZA. Axa: RELUABILITATEA / forma reluabila.
Autorul iese din OCAZIE. Artefactul central = foaia _SISTEM: harta functionala
nativ-Excel a workbook-ului real (componente = foile reale, prin HYPERLINK + ROWS),
care se RUPE scoasa din workbook (#REF!) = anti-SOP.

Strategie (chain T5): deschid Date_MASTER-C16.xlsx (pastreaza TOATE foile, formulele
si stilurile intacte), rescriu START pentru C17 si ADAUG foaia '_SISTEM'.
Vanzari ramane neatins -> suma conservata cap-coada (R-V02.14).

Garda OGLINDA: in _SISTEM, HYPERLINK navigheaza, ROWS/COUNTA/FORMULATEXT oglindesc
starea reala, referintele live oglindesc rezultatul. NICIUNA nu executa (C18),
nu judeca praguri/validari (C19), nu muta ownership (C20). Oglinda + harta, atat.

_SISTEM cartografiaza procesul care a produs raportul (C09-C16): surse -> lucru ->
iesire. Componentele sunt foile reale; pasii reia cineva instruit, fara autor.

Uz: python3 c17/build_date_master_c17.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from collections import defaultdict
import os

SRC = 'c16/Date_MASTER-C16.xlsx'
OUT = 'c17/Date_MASTER-C17.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')
LINK = Font(color='1F5FBF', underline='single')

# clasificarea foilor reale in proces (ce rol au la reluare)
ROL = {
    'Vanzari': 'sursa (fact)', 'PRODUSE': 'sursa (dimensiune)', 'CLIENTI': 'sursa (dimensiune)',
    'Calendar': 'sursa (dimensiune)', 'AGENTI': 'sursa (dimensiune)', 'DEPOZITE': 'sursa (dimensiune)',
    'Regiuni': 'sursa (dimensiune)', 'Masuri': 'lucru (masuri)', 'Masuri_Context': 'lucru (masuri)',
    'Comparatii': 'lucru (comparatie)', 'Interpretare': 'lucru (interpretare)',
    'Vizualizare': 'iesire (forma)', 'Compunere': 'iesire (pagina)', 'Sinteza': 'iesire (mesaj)',
    'Livrare': 'iesire (raport de decizie)',
}


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def main():
    os.makedirs('c17', exist_ok=True)

    # --- pass valori (data_only) pentru a CITI starea reala ---
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]
    vlast = len(vrows) + 1

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)
    cat_val = defaultdict(float)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r)
    top_cat = sorted(cat_val.items(), key=lambda x: -x[1])[0][0]

    # --- pass formule (pastreaza tot C16 intact) ---
    out = openpyxl.load_workbook(SRC)
    ncol = len(vhdr)
    last_col_letter = openpyxl.utils.get_column_letter(ncol)

    # componente reale = foile mostenite (fara START), in ordinea din workbook
    comp = [s for s in out.sheetnames if s not in ('START', '_SISTEM')]

    # gaseste o celula-formula reala in 'Livrare' pentru OBJ_ + FORMULATEXT (oglinda transformarii)
    obj_addr = None
    if 'Livrare' in out.sheetnames:
        L = out['Livrare']
        for row in L.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    obj_addr = "Livrare!%s" % c.coordinate
                    break
            if obj_addr:
                break

    # ---------- START (coperta C17) ----------
    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C17 · SISTEMATIZARE · munca devine un sistem pe care altcineva il porneste'],
        [],
        ['Ideea mare:'],
        ['Raportul e gata (T4). Dar l-ai facut tu, o data. In concediu, munca se opreste:'],
        ['tu esti ocazia. C17 nu mai face raportul - face MUNCA reluabila de altcineva.'],
        ['O munca facuta de doua ori nu mai e o livrare. E un sistem ascuns in workbook.'],
        ['Raport gata (C16)  ->  harta reluabila (_SISTEM)  ->  altcineva porneste, fara mine.'],
        [],
        ['AHA: O munca facuta de doua ori nu mai e o livrare. E un sistem ascuns in workbook.'],
        ['MANTRA: Nu o fac din nou. O fac sistem.'],
        ['MOTTO: Pleci, si munca o reia altcineva.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins, suma conservata cap-coada)'],
        ['  ... (foile C09-C16)   procesul care a produs raportul de decizie'],
        ['  _SISTEM        harta functionala: componente reale + surse + ordine de reluare'],
        [],
        ['_SISTEM (harta reluabila a muncii):'],
        ['  A COMPONENTE   foile reale, prin HYPERLINK + cate randuri au (oglinda vie)'],
        ['  B SURSE        o singura sursa de adevar per intrare (named range SRC_)'],
        ['  C PASI         ordinea de reluare, fiecare pas legat de obiectul real'],
        ['  D CINE RELUA   profilul de competenta care poate porni (nu numele tau)'],
        ['  E PARAMETRI    convenstiile scoase din cap, ca named ranges PARAM_'],
        ['  F PORNIRE      START AICI + punct de reluare + granita C17->C18 + test anti-SOP'],
        [],
        ['Garda OGLINDA: _SISTEM oglindeste, leaga si navigheaza. Nu executa (C18),'],
        ['nu judeca praguri (C19), nu muta proprietatea (C20). Harta, nu motor, nu gardian.'],
        ['Test anti-SOP: stergi celelalte foi -> _SISTEM se umple de #REF!. Moare scoasa'],
        ['din workbook, fiindca e legata de componentele reale. Nu e o procedura de citit.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 13, 18):
        ws['A%d' % rr].font = SECT

    # ---------- _SISTEM (harta functionala) ----------
    si = out.create_sheet('_SISTEM')
    A = si.append

    A(['_SISTEM C17 · harta functionala reluabila a muncii (nativ-Excel, se rupe scoasa din workbook)'])
    A([])
    # START AICI -> primul pas (bloc C, ancora reala dupa ce stim randul)
    A(['START AICI:', '=HYPERLINK("#_SISTEM!A1","mergi la inceputul hartii")', '(intri de aici la fiecare reluare)'])
    A([])

    # BLOC A · COMPONENTE
    rA = si.max_row + 1
    A(['A) COMPONENTE · foile reale ale workbook-ului (navigabile, vii)'])
    A(['componenta (deschide)', 'rol in proces', 'randuri completate (oglinda live)'])
    rAhdr = si.max_row
    for s in comp:
        A([
            '=HYPERLINK("#\'%s\'!A1","%s")' % (s, s),
            ROL.get(s, 'lucru'),
            "=COUNTA('%s'!A:A)" % s,
        ])
    A([])

    # BLOC B · SURSE (o singura sursa de adevar per intrare)
    rB = si.max_row + 1
    A(['B) SURSE · o singura sursa de adevar per intrare (named range SRC_)'])
    A(['intrare', 'sursa unica (SRC_)', 'randuri (oglinda =ROWS(SRC_))'])
    rBhdr = si.max_row
    A(['tabelul fact', 'SRC_VANZARI -> Vanzari!A2:%s%d' % (last_col_letter, vlast), '=ROWS(SRC_VANZARI)'])
    A([])

    # BLOC C · PASI (ordine de reluare, legata de obiecte)
    rC = si.max_row + 1
    A(['C) PASI · ordinea de reluare a muncii, fiecare pas legat de obiectul real'])
    A(['# pas', 'actiune (deschide obiectul)', 'obiect atins', 'rezultat vizibil (oglinda)', 'candidat C18?'])
    rChdr = si.max_row
    pasi = [
        ('Vanzari', 'deschizi tabelul fact (sursa)', "=COUNTA('Vanzari'!A:A)-1 & \" randuri\"", 'nu'),
        ('Masuri', 'recalculezi masurile peste fact', "=COUNTA('Masuri'!A:A) & \" linii masuri\"", 'da (recalcul recurent)'),
        ('Comparatii', 'citesti clasamentul', "=COUNTA('Comparatii'!A:A) & \" linii\"", 'da'),
        ('Interpretare', 'iei explicatia (de ce)', "=COUNTA('Interpretare'!A:A) & \" linii\"", 'nu'),
        ('Compunere', 'organizezi pagina raportului', "=COUNTA('Compunere'!A:A) & \" linii\"", 'nu'),
        ('Sinteza', 'iei mesajul esential', "=COUNTA('Sinteza'!A:A) & \" linii\"", 'nu'),
        ('Livrare', 'predai raportul de decizie', "=COUNTA('Livrare'!A:A) & \" linii\"", 'da (predare recurenta)'),
    ]
    for n, (sheet, act, rez, c18) in enumerate(pasi, 1):
        link = '=HYPERLINK("#\'%s\'!A1","%s")' % (sheet, act) if sheet in out.sheetnames else act
        A([n, link, sheet, rez, c18])
    A([])

    # BLOC D · CINE POATE RELUA (profil de competenta, NU ownership)
    rD = si.max_row + 1
    A(['D) CINE POATE RELUA · profil de competenta (NU ownership, NU nume)'])
    A(['profil de competenta', 'competenta minima'])
    rDhdr = si.max_row
    A(['analist de date instruit', 'stie tabele structurate, SUMIFS/agregari, navigare intre foi'])
    A([])

    # BLOC E · PARAMETRI (convenstii scoase din cap, named ranges PARAM_)
    rE = si.max_row + 1
    A(['E) PARAMETRI · convenstiile scoase din cap, ca valori vii (named range PARAM_)'])
    A(['parametru', 'valoare', 'folosit la'])
    rEhdr = si.max_row
    A(['PARAM_SURSA_CANONICA', 'Vanzari', 'pasul 1 (sursa unica a faptelor)'])
    rp1 = si.max_row
    A(['PARAM_CATEGORIE_FOCUS', top_cat, 'pasii de raport (categoria pe care se concentreaza decizia)'])
    rp2 = si.max_row
    A(['PARAM_FOAIE_FINALA', 'Livrare', 'pasul final (unde sta raportul de decizie)'])
    rp3 = si.max_row
    A([])

    # BLOC F · PORNIRE / RELUARE + GRANITA + ANTI-SOP
    rF = si.max_row + 1
    A(['F) PORNIRE / RELUARE + GRANITA C17->C18 + TEST ANTI-SOP'])
    A(['START AICI', '=HYPERLINK("#_SISTEM!A%d","pasul 1: deschizi tabelul fact")' % (rChdr + 1)])
    A(['punct de reluare', 'reintri la primul pas necompletat din blocul C (ordinea e fixa)'])
    A(['granita C17->C18', 'pasii marcati "candidat C18" se vor automatiza la C18; AICI raman manuali'])
    A(['transformare (oglinda)', '=FORMULATEXT(%s)' % obj_addr if obj_addr else '(nicio formula gasita in Livrare)'])
    A(['TEST ANTI-SOP', '=COUNTA(\'Vanzari\'!A:A) & " randuri vii; sterge foile -> #REF!; copiat in Word = mort"'])
    A(['de ce e nativ-Excel', 'HYPERLINK, ROWS, COUNTA, FORMULATEXT depind de foile reale: scoase, se rup'])
    A([])
    A(['Garda OGLINDA: tot ce e mai sus OGLINDESTE si NAVIGHEAZA. Nu executa, nu judeca, nu deleaga.'])

    # stiluri _SISTEM
    si['A1'].font = TITLE
    for rr in (rA, rB, rC, rD, rE, rF):
        si['A%d' % rr].font = SECT
    for rr in (rAhdr, rBhdr, rChdr, rDhdr, rEhdr):
        hdr(si, rr)
    si.column_dimensions['A'].width = 30
    si.column_dimensions['B'].width = 46
    si.column_dimensions['C'].width = 30
    si.column_dimensions['D'].width = 34

    # named ranges (SRC_ / PARAM_) - oglinzi/convenstii, legate de celule reale
    out.defined_names['SRC_VANZARI'] = DefinedName(
        'SRC_VANZARI', attr_text="Vanzari!$A$2:$%s$%d" % (last_col_letter, vlast))
    out.defined_names['PARAM_SURSA_CANONICA'] = DefinedName(
        'PARAM_SURSA_CANONICA', attr_text="_SISTEM!$B$%d" % rp1)
    out.defined_names['PARAM_CATEGORIE_FOCUS'] = DefinedName(
        'PARAM_CATEGORIE_FOCUS', attr_text="_SISTEM!$B$%d" % rp2)
    out.defined_names['PARAM_FOAIE_FINALA'] = DefinedName(
        'PARAM_FOAIE_FINALA', attr_text="_SISTEM!$B$%d" % rp3)

    # ordine: START, _SISTEM, apoi restul mostenit
    order = [s for s in out.sheetnames if s not in ('START', '_SISTEM')]
    new_order = ['START', '_SISTEM'] + order
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  componente mapate in _SISTEM:', len(comp))
    print('  named ranges:', list(out.defined_names.keys()))
    print('  obj_addr (FORMULATEXT oglinda):', obj_addr)
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    assert '_SISTEM' in out.sheetnames, '_SISTEM lipseste'
    assert 'Vanzari' in out.sheetnames, 'Vanzari lipseste (anti-SOP depinde de el)'
    print('  OK: suma conservata, _SISTEM adaugata, T5 deschis (harta reluabila peste procesul C09-C16)')


if __name__ == '__main__':
    main()
