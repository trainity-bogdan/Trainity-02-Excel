#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c19.py - Date_MASTER-C19.xlsx (T5 - GUVERNAREA), controlul.

C19 = GUVERNAREA. Verb: a GUVERNA. Axa: DAI SISTEMULUI REGULI CARE IL TIN CORECT
FARA SUPRAVEGHERE. Autorul iese din ATENTIE. Artefactul central = foaia _GUVERNARE:
praguri vii (LIMIT_) + validare la sursa (Data Validation) + semnale + STARE
(STATUS OK/ATENTIE/OPRIT) + lista de exceptii + fail-safe (output legat de
STATUS=OPRIT). Nativ-Excel, pe datele reale (Vanzari_Curat).

DEMONSTRATIE AUTENTICA: datele mostenite de la C18 contin o eroare grava tacuta -
un rand cu valoare_neta imposibila (negativa, != cantitate*pret_unitar) care trage
totalul real (~6,97M) la cel inregistrat (~1,25M) fara sa se planga nimeni. Motorul
(C18) a lasat-o sa treaca. C19 o prinde prin reguli (consistenta rand + inchiderea
totalului), trece STATUS in OPRIT si retine rezultatul prin fail-safe. Exact testul
ochilor inchisi: regula prinde, nu ochiul.

Strategie (chain T5): deschid Date_MASTER-C18.xlsx (pastreaza TOATE foile, formulele
si stilurile intacte), adaug START (coperta C19) si _GUVERNARE.
Vanzari_Curat ramane neatins -> suma conservata cap-coada fata de C18 (R-V02.14).

Garda: _GUVERNARE marcheaza, semnaleaza si opreste (fail-safe). NU executa lantul (C18),
NU desemneaza responsabil / owner / escaladare (C20), NU e dashboard de citit (T4).

Uz: python3 c19/build_date_master_c19.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

SRC = 'c18/Date_MASTER-C18.xlsx'
OUT = 'c19/Date_MASTER-C19.xlsx'
CLEAN = 'Vanzari_Curat'
COL_CAT, COL_CANT, COL_PRET, COL_VAL = 'F', 'J', 'K', 'L'
ROW1, ROWN = 2, 2001
CATS = ['Lactate', 'Bacanie', 'Bauturi', 'Nealimentare']

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')
GREEN = PatternFill('solid', fgColor='C6EFCE')
AMBER = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def main():
    os.makedirs('c19', exist_ok=True)

    wv = openpyxl.load_workbook(SRC, data_only=True)
    VC = wv[CLEAN]
    hdr_cells = [c.value for c in next(VC.iter_rows(min_row=1, max_row=1))]
    ixv = {h: i for i, h in enumerate(hdr_cells) if h}
    drows = [r for r in VC.iter_rows(min_row=2, values_only=True)]

    def num(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None
    suma = round(sum(num(r[ixv['valoare_neta']]) for r in drows if num(r[ixv['valoare_neta']]) is not None), 2)
    nrows = len(drows)

    out = openpyxl.load_workbook(SRC)

    # ranges pe datele reale
    R_VAL = "'%s'!$%s$%d:$%s$%d" % (CLEAN, COL_VAL, ROW1, COL_VAL, ROWN)
    R_CANT = "'%s'!$%s$%d:$%s$%d" % (CLEAN, COL_CANT, ROW1, COL_CANT, ROWN)
    R_PRET = "'%s'!$%s$%d:$%s$%d" % (CLEAN, COL_PRET, ROW1, COL_PRET, ROWN)
    R_CAT = "'%s'!$%s$%d:$%s$%d" % (CLEAN, COL_CAT, ROW1, COL_CAT, ROWN)

    # ---------- START (coperta C19) ----------
    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C19 - GUVERNARE - sistemul se tine corect fara ochiul tau'],
        [],
        ['Ideea mare:'],
        ['Motorul ruleaza (C18): lantul se misca fara mana ta. Si totusi, pe o intrare'],
        ['gresita produce un rezultat gresit in tacere, iar tu tot verifici la fiecare ciclu.'],
        ['C19 muta verificarea din mintea ta in sistem: reguli care prind, semnaleaza si opresc.'],
        ['Motor care ruleaza (C18) -> reguli care il tin corect (_GUVERNARE) -> fara ochiul tau.'],
        [],
        ['HERO: Cum se tine corect fara ochiul meu?'],
        ['AHA: Un sistem in care ai incredere nu e cel pe care il urmaresti. E cel care se prinde singur cand greseste.'],
        ['MANTRA: Nu o supraveghez. O guvernez prin reguli.'],
        ['MOTTO: Pleci, si munca se tine singura.'],
        [],
        ['Ce vezi la deschidere (testul ochilor inchisi, pe date reale):'],
        ['  Datele mostenite de la C18 contin un rand imposibil (valoare_neta negativa,'],
        ['  diferita de cantitate x pret). Motorul l-a lasat sa treaca tacut. _GUVERNARE il'],
        ['  prinde prin reguli, trece STATUS in OPRIT si RETINE rezultatul (fail-safe).'],
        [],
        ['_GUVERNARE (controlul sistemului):'],
        ['  A PRAGURI      plicul previzibil de variatie, ca valori vii (named range LIMIT_)'],
        ['  B DIAGNOSTIC   oglinzi vii peste datele reale (totaluri, randuri in afara regulilor)'],
        ['  C REGULI       validare la sursa: intrarea gresita e respinsa (Data Validation)'],
        ['  D SEMNALE      fiecare regula intoarce OK / ATENTIE / OPRIT'],
        ['  E STARE        STATUS agregat: OK / ATENTIE / OPRIT, calculat din reguli'],
        ['  F EXCEPTII     cazul prins e marcat, nu trece tacut'],
        ['  G FAIL-SAFE    cand STATUS=OPRIT, rezultatul corupt nu mai curge in aval'],
        [],
        ['Granita: C18 ruleaza - C19 se tine corect - C20 preda responsabilitatea.'],
        ['_GUVERNARE marcheaza, semnaleaza si opreste. NU executa lantul (C18), NU desemneaza'],
        ['responsabil / owner / escaladare (C20), NU e dashboard de citit (T4).'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 14, 19):
        ws['A%d' % rr].font = SECT

    # ---------- _GUVERNARE ----------
    if '_GUVERNARE' in out.sheetnames:
        out.remove(out['_GUVERNARE'])
    g = out.create_sheet('_GUVERNARE')
    A = g.append

    A(['_GUVERNARE C19 - reguli care tin sistemul corect fara supraveghere (nativ-Excel, pe Vanzari_Curat)'])
    A([])

    # A) PRAGURI
    rA = g.max_row + 1
    A(['A) PRAGURI - plicul previzibil de variatie, ca valori vii (named range LIMIT_)'])
    A(['prag', 'valoare', 'rol']); rAhdr = g.max_row
    A(['LIMIT_CANT_MIN', 1, 'cantitate minima acceptata']); r_cmin = g.max_row
    A(['LIMIT_CANT_MAX', 500, 'cantitate maxima acceptata']); r_cmax = g.max_row
    A(['LIMIT_VAL_MIN', 0.01, 'valoare_neta trebuie sa fie pozitiva']); r_vmin = g.max_row
    A(['LIMIT_TOL', 0.5, 'toleranta la consistenta rand si la inchiderea totalului']); r_tol = g.max_row
    A([])

    # B) DIAGNOSTIC (oglinzi vii)
    rB = g.max_row + 1
    A(['B) DIAGNOSTIC - oglinzi vii peste datele reale'])
    A(['masura', 'valoare (oglinda vie)', 'ce arata']); rBhdr = g.max_row
    A(['TOTAL_INREGISTRAT', '=SUM(%s)' % R_VAL, 'suma valoare_neta asa cum a iesit din motor']); r_treg = g.max_row
    A(['TOTAL_CONSISTENT', '=SUMPRODUCT(%s,%s)' % (R_CANT, R_PRET), 'suma daca fiecare rand respecta cantitate x pret']); r_tcon = g.max_row
    A(['randuri_inconsistente', '=SUMPRODUCT(--(ABS(%s-%s*%s)>$B$%d))' % (R_VAL, R_CANT, R_PRET, r_tol), 'randuri unde valoare_neta != cantitate x pret']); r_rinc = g.max_row
    A(['randuri_nepozitive', '=COUNTIF(%s,"<="&0)' % R_VAL, 'randuri cu valoare_neta nula sau negativa']); r_rnp = g.max_row
    A(['cantitati_in_afara', '=COUNTIF(%s,"<"&$B$%d)+COUNTIF(%s,">"&$B$%d)' % (R_CANT, r_cmin, R_CANT, r_cmax), 'cantitati sub minim sau peste maxim']); r_rca = g.max_row
    A(['categorii_necunoscute', '=COUNTA(%s)-(%s)' % (R_CAT, '+'.join('COUNTIF(%s,"%s")' % (R_CAT, c) for c in CATS)), 'categorii in afara listei permise']); r_rcat = g.max_row
    A([])

    # C) REGULI / VALIDARE LA SURSA (Data Validation pe zona de intrare noua)
    rC = g.max_row + 1
    A(['C) REGULI - validare la sursa: intrarea gresita e respinsa inainte sa intre (Data Validation)'])
    A(['camp intrare noua', 'valoare (incearca o intrare)', 'regula aplicata']); rChdr = g.max_row
    A(['cantitate', None, 'numar intreg intre LIMIT_CANT_MIN si LIMIT_CANT_MAX']); r_in_cant = g.max_row
    A(['pret_unitar', None, 'numar zecimal strict pozitiv']); r_in_pret = g.max_row
    A(['categorie', None, 'doar din lista permisa: ' + ', '.join(CATS)]); r_in_cat = g.max_row
    A([])

    # D) SEMNALE
    rD = g.max_row + 1
    A(['D) SEMNALE - fiecare regula citeste diagnosticul si intoarce OK / ATENTIE / OPRIT'])
    A(['regula', 'semnal (oglinda vie)', 'ce prinde']); rDhdr = g.max_row
    A(['consistenta randurilor', '=IF($B$%d>0,"OPRIT","OK")' % r_rinc, 'un rand unde valoare_neta != cantitate x pret']); r_chk1 = g.max_row
    A(['valoare_neta pozitiva', '=IF($B$%d>0,"OPRIT","OK")' % r_rnp, 'o valoare_neta nula sau negativa']); r_chk2 = g.max_row
    A(['inchiderea totalului', '=IF(ABS($B$%d-$B$%d)>$B$%d,"OPRIT","OK")' % (r_treg, r_tcon, r_tol), 'totalul inregistrat nu inchide cu cel consistent']); r_chk3 = g.max_row
    A(['cantitate in plic', '=IF($B$%d>0,"ATENTIE","OK")' % r_rca, 'o cantitate sub minim sau peste maxim']); r_chk4 = g.max_row
    A(['categorie cunoscuta', '=IF($B$%d>0,"ATENTIE","OK")' % r_rcat, 'o categorie in afara listei permise']); r_chk5 = g.max_row
    A([])

    # E) STARE (STATUS)
    rE = g.max_row + 1
    A(['E) STARE - STATUS agregat, calculat din semnale'])
    A(['STATUS_SISTEM',
       '=IF(COUNTIF($B$%d:$B$%d,"OPRIT")>0,"OPRIT",IF(COUNTIF($B$%d:$B$%d,"ATENTIE")>0,"ATENTIE","OK"))' % (
           r_chk1, r_chk5, r_chk1, r_chk5),
       'OK = toate regulile trec; ATENTIE = abatere; OPRIT = eroare grava']); r_status = g.max_row
    A([])

    # F) EXCEPTII
    rF = g.max_row + 1
    A(['F) EXCEPTII - cazul prins e marcat, nu trece tacut (fara responsabil - granita C20)'])
    A(['exceptie', 'nr. cazuri (oglinda vie)', 'actiune (marcheaza / opreste, NU desemneaza)']); rFhdr = g.max_row
    A(['rand inconsistent (valoare != cantitate x pret)', '=$B$%d' % r_rinc, 'marcat; trece STATUS in OPRIT, fail-safe retine rezultatul'])
    A(['valoare_neta nepozitiva', '=$B$%d' % r_rnp, 'blocata; trece STATUS in OPRIT'])
    A(['categorie necunoscuta', '=$B$%d' % r_rcat, 'marcata ca exceptie, nu trece tacut'])
    A(['cantitate in afara plicului', '=$B$%d' % r_rca, 'respinsa la sursa (Data Validation)'])
    A([])

    # G) FAIL-SAFE
    rG = g.max_row + 1
    A(['G) FAIL-SAFE - cand STATUS=OPRIT, rezultatul corupt nu mai curge in aval'])
    A(['OUTPUT_GUVERNAT',
       '=IF($B$%d="OPRIT","BLOCAT: rezultat retinut (fail-safe); "&$B$%d&" rand(uri) imposibil(e) prins(e); total inregistrat "&TEXT($B$%d,"#,##0.00")&" nu inchide cu "&TEXT($B$%d,"#,##0.00"),"valoare neta totala validata: "&TEXT($B$%d,"#,##0.00"))'
       % (r_status, r_rinc, r_treg, r_tcon, r_treg),
       'iesirea e legata de STARE: OPRIT inseamna ca nu se livreaza un rezultat gresit']); r_out = g.max_row
    A([])

    # H) GRANITE / GARDA / TEST
    rH = g.max_row + 1
    A(['H) GRANITE + GARDA + TESTUL OCHILOR INCHISI'])
    A(['granita C18->C19', 'C18 ruleaza (motorul); C19 il tine corect (regulile). C19 nu construieste motorul.'])
    A(['granita C19->C20', 'C19 marcheaza si opreste; C20 preda responsabilitatea. C19 NU desemneaza responsabil.'])
    A(['anti-T4', 'semnalul ACTIONEAZA (schimba STAREA, opreste), nu e dashboard de citit pentru un om.'])
    A(['anti-babysitting', 'regulile prind, nu ochiul.'])
    A(['testul ochilor inchisi',
       '=IF($B$%d="OK","sistemul se tine corect singur (toate regulile trec)","regulile au prins singure abaterea: STATUS="&$B$%d&" (fara ochiul tau)")' % (r_status, r_status)])
    A([])
    A(['Garda: _GUVERNARE marcheaza, semnaleaza si opreste. Nu executa, nu desemneaza, nu deleaga.'])

    # stiluri
    g['A1'].font = TITLE
    for rr in (rA, rB, rC, rD, rE, rF, rG, rH):
        g['A%d' % rr].font = SECT
    for rr in (rAhdr, rBhdr, rChdr, rDhdr, rFhdr):
        hdr(g, rr)
    g['A%d' % r_status].font = Font(bold=True, size=11, color='1F2A44')
    g.column_dimensions['A'].width = 34
    g.column_dimensions['B'].width = 64
    g.column_dimensions['C'].width = 54

    # Data Validation la sursa
    dv_cant = DataValidation(type='whole', operator='between', formula1='$B$%d' % r_cmin, formula2='$B$%d' % r_cmax,
                             allow_blank=True, showErrorMessage=True)
    dv_cant.errorTitle = 'Regula C19 (GUVERNARE)'; dv_cant.error = 'Cantitate in afara plicului permis: respinsa la sursa.'
    dv_pret = DataValidation(type='decimal', operator='greaterThan', formula1='0', allow_blank=True, showErrorMessage=True)
    dv_pret.errorTitle = 'Regula C19 (GUVERNARE)'; dv_pret.error = 'pret_unitar trebuie sa fie strict pozitiv: respins la sursa.'
    dv_cat = DataValidation(type='list', formula1='"%s"' % ','.join(CATS), allow_blank=True, showErrorMessage=True)
    dv_cat.errorTitle = 'Regula C19 (GUVERNARE)'; dv_cat.error = 'Categorie in afara listei permise: respinsa la sursa.'
    for dv, rr in ((dv_cant, r_in_cant), (dv_pret, r_in_pret), (dv_cat, r_in_cat)):
        g.add_data_validation(dv); dv.add('B%d' % rr)

    # Conditional formatting pe STATUS (semnal care schimba culoarea dupa regula)
    sc = 'B%d' % r_status
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"OK"'], fill=GREEN))
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"ATENTIE"'], fill=AMBER))
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"OPRIT"'], fill=RED))

    # named ranges
    def dn(name, ref):
        out.defined_names[name] = DefinedName(name, attr_text=ref)
    dn('LIMIT_CANT_MIN', "_GUVERNARE!$B$%d" % r_cmin)
    dn('LIMIT_CANT_MAX', "_GUVERNARE!$B$%d" % r_cmax)
    dn('LIMIT_VAL_MIN', "_GUVERNARE!$B$%d" % r_vmin)
    dn('LIMIT_TOL', "_GUVERNARE!$B$%d" % r_tol)
    dn('TOTAL_INREGISTRAT', "_GUVERNARE!$B$%d" % r_treg)
    dn('TOTAL_CONSISTENT', "_GUVERNARE!$B$%d" % r_tcon)
    dn('STATUS_SISTEM', "_GUVERNARE!$B$%d" % r_status)
    dn('OUTPUT_GUVERNAT', "_GUVERNARE!$B$%d" % r_out)

    rest = [s for s in out.sheetnames if s not in ('START', '_GUVERNARE')]
    out._sheets.sort(key=lambda s: (['START', '_GUVERNARE'] + rest).index(s.title))
    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  named ranges:', list(out.defined_names.keys()))
    print('  STATUS=_GUVERNARE!B%d | OUTPUT=_GUVERNARE!B%d | TOTAL_INREG=B%d | TOTAL_CONS=B%d' % (r_status, r_out, r_treg, r_tcon))
    print('  suma Vanzari_Curat (conservata fata de C18):', suma, '| randuri:', nrows)
    assert '_GUVERNARE' in out.sheetnames and CLEAN in out.sheetnames
    assert abs(suma - 1247893.5) < 0.01, 'SUMA NU se conserva fata de C18'
    print('  OK: suma conservata, _GUVERNARE adaugata (praguri/diagnostic/validari/semnale/STATUS/exceptii/fail-safe)')


if __name__ == '__main__':
    main()
