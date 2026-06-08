#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c13.py - Date_MASTER-C13.xlsx (T4 · VIZUALIZARE), deschide T4.

Nume fisier = Date_MASTER-C13.xlsx (cerut de gate B2 / decizie BRAIN mandat 011).

C13 da FORMA ONESTA unui rezultat deja produs de analiza (T3). Verb: a VIZUALIZA.
Axa: ONESTITATEA FORMEI. C13 = obiect vizual onest, NU dashboard (=C14).

Strategie sigura (ca la C12 din C11): deschid Date_MASTER-C12.xlsx (pastreaza TOATE
foile, formulele si stilurile intacte), rescriu START pentru C13 si ADAUG foaia
'Vizualizare'. Vanzari ramane neatins -> suma conservata cap-coada (R-V02.14).

Garda C13: workbook = SUPORT pentru construirea si verificarea obiectului vizual onest
(demonstratie forma onesta vs forma care minte + verificare vizual contra cifrei brute).
ZERO dashboard / pagina / chart publicabil (=C14). ZERO 'de ce'/cauza ca identitate (=C12).

Uz: python3 c13/build_date_master_c13.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import os

SRC = 'c12/Date_MASTER-C12.xlsx'
OUT = 'c13/Date_MASTER-C13.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def xlstr(s):
    return '"' + str(s).replace('"', '""') + '"'


def main():
    os.makedirs('c13', exist_ok=True)

    # --- pass valori (data_only) pentru a CITI rezultatul de vizualizat ---
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]
    vlast = len(vrows) + 1

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)

    cat_val = defaultdict(float)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r)
    cats = sorted(cat_val.items(), key=lambda x: -x[1])
    top_cat, top_val = cats[0]
    second_cat, second_val = cats[1]
    # diferenta procentuala reala intre primele doua categorii (rezultatul de vizualizat onest)
    dif_pct = round(100 * (top_val - second_val) / top_val, 1)

    # --- pass formule (pastreaza tot C12 intact) ---
    out = openpyxl.load_workbook(SRC)

    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C13 · VIZUALIZARE · forma onesta a unui rezultat, verificata contra cifrei'],
        [],
        ['Ideea mare:'],
        ['Analiza a dat raspunsul corect. Dar in model e mut. C13 ii da FORMA care il'],
        ['face adevarat si vizibil dintr-o privire. Forma gresita minte cu cifra corecta.'],
        ['Rezultat corect  ->  forma aleasa  ->  vizual onest  ->  verificat contra cifrei brute.'],
        [],
        ['AHA: Forma gresita minte cu cifra corecta.'],
        ['MANTRA: Nu desenez frumos. Desenez adevarat.'],
        ['MOTTO: Forma nu se nimereste. Se alege.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins fata de C12, suma conservata)'],
        ['  Comparatii     clasamentul mostenit (rezultatul de vizualizat)'],
        ['  Interpretare   explicatia mostenita de la C12'],
        ['  Vizualizare    SUPORT: forma onesta vs forma care minte + verificare + 6 reguli'],
        [],
        ['Exercitiul:'],
        ['  1. Iei un rezultat corect (mostenit, ex: diferenta dintre primele categorii).'],
        ['  2. Vezi cum aceeasi cifra, in forma trunchiata, spune altceva.'],
        ['  3. Alegi tipul potrivit naturii rezultatului, cu axa de la zero.'],
        ['  4. Verifici: vizualul citit singur da aceeasi concluzie ca cifra bruta?'],
        ['  5. Aplici cele 6 reguli de onestitate a formei.'],
        [],
        ['C13 doar da FORMA. Nu compune pagina/dashboard (C14), nu formuleaza mesajul (C15),'],
        ['nu livreaza (C16), nu explica de ce (C12). C13 face obiectul adevarat; C14 il asaza.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 12, 18):
        ws['A%d' % rr].font = SECT

    # --- foaia Vizualizare (suport pentru obiectul vizual onest) ---
    wi = out.create_sheet('Vizualizare')
    A = wi.append
    A(['VIZUALIZARE C13 · forma onesta a unui rezultat, verificata contra cifrei brute'])
    A([])

    A(['1) REZULTATUL DE VIZUALIZAT · citit din model (nu inventat)'])
    A(['rezultat', 'sursa', 'natura'])
    A(['Diferenta valorii intre primele doua categorii', 'Comparatii (C11/C12)',
       'comparatie -> forma potrivita = bare cu axa de la zero'])
    A([])

    r2 = wi.max_row + 1
    A(['2) CIFRA BRUTA · valorile reale (referinta de adevar)'])
    A(['categorie', 'valoare (live)', 'pondere din total (%)'])
    r_hdr2 = wi.max_row
    for c, v in cats:
        A([c,
           '=ROUND(SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s),2)' % (vlast, vlast, xlstr(c)),
           '=ROUND(100*SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s)/SUM(Vanzari!J2:J%d),1)'
           % (vlast, vlast, xlstr(c), vlast)])
    A([])

    r3 = wi.max_row + 1
    A(['3) FORMA ONESTA vs FORMA CARE MINTE · aceeasi cifra, doua concluzii'])
    A(['forma', 'axa', 'ce concluzie sugereaza', 'verdict'])
    A(['bare cu axa de la zero', 'porneste de la 0',
       'diferenta reala de %.1f%% intre primele doua categorii' % dif_pct, 'ONESTA'])
    A(['bare cu axa trunchiata', 'porneste de la a doua valoare',
       'diferenta pare uriasa, desi cifra e aceeasi', 'MINTE'])
    A(['placinta cu prea multe felii', 'fara axa',
       'ordinea si marimea categoriilor devin greu de citit', 'MINTE'])
    A([])

    r4 = wi.max_row + 1
    A(['4) VERIFICARE · vizualul citit singur = cifra bruta?'])
    A(['proba', 'rezultat'])
    A(['concluzia din forma onesta', '%s conduce, cu %.1f%% peste %s' % (top_cat, dif_pct, second_cat)])
    A(['concluzia din cifra bruta (sectiunea 2)', 'identica: %s conduce cu aceeasi diferenta' % top_cat])
    A(['al doilea ochi (doar graficul, fara cifra)', 'trage aceeasi concluzie -> forma e onesta'])
    A([])

    r5 = wi.max_row + 1
    A(['5) CELE 6 REGULI DE ONESTITATE A FORMEI'])
    A(['#', 'deformare', 'regula onesta'])
    for n, (d, reg) in enumerate([
        ('Axa care exagereaza', 'axa pleaca de la zero, sau abaterea e declarata explicit'),
        ('Tipul de grafic nepotrivit', 'tipul urmeaza natura rezultatului'),
        ('Scala care inventeaza relatii', 'o singura scala liniara, declarata'),
        ('Codarea care sugereaza fals', 'o singura dimensiune codata coerent'),
        ('Agregarea care ascunde variatia', 'arati distributia / variatia, nu doar media'),
        ('Eticheta / unitatea / contextul lipsa', 'fiecare vizual poarta unitate, eticheta si context'),
    ], 1):
        A([n, d, reg])
    A([])

    r6 = wi.max_row + 1
    A(['6) HANDOFF · C13 face obiectul adevarat, C14 il asaza in pagina'])
    A(['input de la T3 (C12)', 'rezultat corect, explicat, dar mut in model'])
    A(['output C13', 'obiect vizual onest, verificat contra cifrei brute'])
    A(['predat catre C14', 'compunerea paginii / dashboard-ului (organizare spatiala)'])
    A(['C13 NU face', 'dashboard, pagina, layout, mesaj, livrare'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wi['A1'].font = TITLE
    for rr in (3, r2, r3, r4, r5, r6):
        wi['A%d' % rr].font = SECT
    hdr(wi, 4)
    hdr(wi, r_hdr2)
    hdr(wi, r3 + 1)
    hdr(wi, r4 + 1)
    hdr(wi, r5 + 1)

    # Vizualizare imediat dupa Interpretare in ordinea foilor
    order = [s for s in out.sheetnames if s not in ('START', 'Vizualizare')]
    if 'Interpretare' in order:
        i = order.index('Interpretare') + 1
    else:
        i = len(order)
    new_order = ['START'] + order[:i] + ['Vizualizare'] + order[i:]
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  top_cat=%s (%.2f) | a doua=%s (%.2f) | dif=%.1f%%'
          % (top_cat, top_val, second_cat, second_val, dif_pct))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    print('  OK: suma conservata, Vizualizare adaugata, T4 deschisa')


if __name__ == '__main__':
    main()
