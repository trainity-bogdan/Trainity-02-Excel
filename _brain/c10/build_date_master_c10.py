#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c10.py - Date_MASTER-C10.xlsx (T3 · MASURI), versiune SIMPLA UX.

C10 defineste masuri peste modelul relational primit de la C09.
Verb: a DEFINI. Intrebare-mama: Cat? Single source of truth.

Garda C10 (PUR a defini): zero ranking, zero contributie/pondere-in-total,
zero comparatie contextuala intre entitati (acelea = C11 'a compara'),
zero cauza/interpretare (C12), zero dashboard (T4). Masuri_Context demonstreaza
DOAR context-awareness (aceeasi definitie corecta pe orice taietura), NU clasament.

7 foi vizibile:
  1 START            - ghid scurt: ce e o masura si de ce
  2 Vanzari          - fact principal (neatins fata de C09, suma conservata)
  3 PRODUSE          - dimensiune
  4 CLIENTI          - dimensiune
  5 Calendar         - dimensiune
  6 Masuri           - DEFINITIILE: single source of truth + baza + reper/prag + semnal
  7 Masuri_Context   - context-awareness: aceeasi masura, taieturi diferite (NU clasament)

Ascunse (continuitate schema, nu fac parte din experienta): AGENTI, DEPOZITE, Regiuni.

Uz: python3 _brain/c10/build_date_master_c10.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

SRC = 'c09/Date_MASTER-C09.xlsx'
OUT = 'c10/Date_MASTER-C10.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def main():
    os.makedirs('c10', exist_ok=True)
    src = openpyxl.load_workbook(SRC, data_only=True)
    V = src['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [list(r) for r in V.iter_rows(min_row=2, values_only=True)]
    nrows = len(vrows)
    vlast = nrows + 1  # ultima linie de date in Vanzari (cu header pe rand 1)

    # --- cifre verificate independent (pentru coloana "rezultat" + assert) ---
    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)
    n_tranz = sum(1 for r in vrows if r[ix['nr_factura']] not in (None, ''))
    n_clienti = len({r[ix['client_nume']] for r in vrows})
    n_zile = len({r[ix['data_factura']] for r in vrows if r[ix['data_factura']]})
    avg_tranz = round(suma / n_tranz, 2)
    avg_client = round(suma / n_clienti, 2)
    avg_zi = round(suma / n_zile, 2)

    # reper/prag definit (tinta interna pentru valoarea medie per tranzactie).
    # Valoare clara, definita ca reper de business; semnalul se citeste fata de ea.
    prag_mediu = 3500.0
    semnal_txt = 'peste reper' if avg_tranz >= prag_mediu else 'sub reper'

    # context-awareness: aceeasi masura (medie per tranzactie) pe taieturi diferite
    cat_demo = 'Hardware'
    cat_rows = [r for r in vrows if r[ix['categorie']] == cat_demo]
    avg_cat = round(sum(vn(r) for r in cat_rows) / len(cat_rows), 2) if cat_rows else 0.0

    client_demo = sorted({r[ix['client_nume']] for r in vrows})[0]
    cli_rows = [r for r in vrows if r[ix['client_nume']] == client_demo]
    avg_cli = round(sum(vn(r) for r in cli_rows) / len(cli_rows), 2) if cli_rows else 0.0

    # luna demo: prima luna prezenta in set
    luni_set = sorted({(r[ix['data_factura']].year, r[ix['data_factura']].month)
                       for r in vrows if r[ix['data_factura']]})
    an_demo, mo_demo = luni_set[0]
    mo_rows = [r for r in vrows if r[ix['data_factura']]
               and r[ix['data_factura']].year == an_demo
               and r[ix['data_factura']].month == mo_demo]
    avg_mo = round(sum(vn(r) for r in mo_rows) / len(mo_rows), 2) if mo_rows else 0.0

    out = openpyxl.Workbook()
    out.remove(out.active)

    # --- 1 START ---
    ws = out.create_sheet('START')
    for line in [
        ['C10 · MASURI · cum transformi cifrele in indicatori controlabili'],
        [],
        ['Ideea mare:'],
        ['Fara masuri ai cifre. Cu masuri corecte ai decizie.'],
        ['Cifra bruta  ->  intrebare clara  ->  masura definita  ->  decizie controlabila.'],
        [],
        ['AHA: Nu cifra conteaza. Conteaza ce intrebare raspunde cifra.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (model legat corect, primit de la C09)'],
        ['  PRODUSE        dimensiune'],
        ['  CLIENTI        dimensiune'],
        ['  Calendar       dimensiune'],
        ['  Masuri         DEFINITIILE: o singura sursa de adevar, baza, reper, semnal'],
        ['  Masuri_Context aceeasi masura pe taieturi diferite (dovada de robustete)'],
        [],
        ['Exercitiul:'],
        ['  1. Pornesti de la modelul legat corect (mostenit din C09).'],
        ['  2. Formulezi intrebarea de business: "Cat?".'],
        ['  3. Definesti masura o singura data (single source of truth).'],
        ['  4. Ii declari baza de raportare si o ancorezi intr-un reper.'],
        ['  5. Verifici ca aceeasi definitie da rezultat corect pe orice taietura.'],
        [],
        ['C10 doar DEFINESTE masuri. CARE masura e mai mare se compara la C11.'],
        ['De ce arata asa, se interpreteaza la C12. Tabloul de bord vine la T4.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    ws['A3'].font = SECT
    ws['A9'].font = SECT
    ws['A17'].font = SECT

    # --- 2 Vanzari (fact neatins) ---
    wv = out.create_sheet('Vanzari')
    wv.append(vhdr)
    for r in vrows:
        wv.append(r)
    hdr(wv, 1)

    # --- 3 PRODUSE / 4 CLIENTI / 5 Calendar (dimensiuni carate) ---
    for name in ('PRODUSE', 'CLIENTI', 'Calendar'):
        wd = out.create_sheet(name)
        for row in src[name].iter_rows(values_only=True):
            wd.append(list(row))
        hdr(wd, 1)

    # ascunse pentru continuitate schema
    for name in ('AGENTI', 'DEPOZITE', 'Regiuni'):
        if name in src.sheetnames:
            wh = out.create_sheet(name)
            for row in src[name].iter_rows(values_only=True):
                wh.append(list(row))
            hdr(wh, 1)
            wh.sheet_state = 'hidden'

    # --- 6 Masuri (definitiile) ---
    wm = out.create_sheet('Masuri')
    A = wm.append
    A(['MASURI C10 · definitiile: o singura sursa de adevar, baza, reper, semnal'])
    A([])
    A(['1) MASURILE · single source of truth (definite o singura data, reutilizabile)'])
    A(['cod', 'masura', 'definitie', 'formula (live)', 'rezultat', 'baza de raportare'])
    # randurile masurilor; retin numerele de rand pentru referinte (M1=rand 6 etc.)
    r_m1 = wm.max_row + 1
    A(['M1', 'Total valoare', 'suma valorii nete pe tot setul',
       '=ROUND(SUM(Vanzari!J2:J%d),2)' % vlast, suma, 'tot setul (toate tranzactiile)'])
    r_m2 = wm.max_row + 1
    A(['M2', 'Numar tranzactii', 'cate tranzactii are setul',
       '=COUNTA(Vanzari!A2:A%d)' % vlast, n_tranz, 'tot setul'])
    r_m3 = wm.max_row + 1
    A(['M3', 'Valoare medie per tranzactie', 'M1 impartit la M2',
       '=ROUND(E%d/E%d,2)' % (r_m1, r_m2), avg_tranz, 'numar tranzactii (declarat)'])
    r_m4 = wm.max_row + 1
    A(['M4', 'Valoare medie per client', 'M1 impartit la numarul de clienti distincti',
       '=ROUND(E%d/SUMPRODUCT(1/COUNTIF(Vanzari!C2:C%d,Vanzari!C2:C%d)),2)'
       % (r_m1, vlast, vlast), avg_client, 'clienti distincti (declarat)'])
    r_m5 = wm.max_row + 1
    A(['M5', 'Valoare medie pe zi activa', 'M1 impartit la numarul de zile cu vanzari',
       '=ROUND(E%d/SUMPRODUCT(1/COUNTIF(Vanzari!B2:B%d,Vanzari!B2:B%d)),2)'
       % (r_m1, vlast, vlast), avg_zi, 'zile active (cu cel putin o vanzare)'])
    A([])
    r_sect2 = wm.max_row + 1
    A(['2) REPER / PRAG · masura capata sens raportata la o tinta definita'])
    A(['masura', 'reper (tinta interna)', 'formula semnal (live)', 'semnal'])
    r_sig = wm.max_row + 1
    A(['M3 Valoare medie per tranzactie', prag_mediu,
       '=IF(E%d>=B%d,"peste reper","sub reper")' % (r_m3, r_sig), semnal_txt])
    A([])
    r_sect3 = wm.max_row + 1
    A(['3) CIFRA UTILA vs CIFRA DECORATIVA · criteriul de selectie'])
    A(['cifra', 'statut', 'de ce'])
    A(['suma valoare_neta', 'UTILA -> masura M1', 'raspunde la o intrebare de business'])
    A(['media valorii per tranzactie', 'UTILA -> masura M3', 'controlabila, are baza si reper'])
    A(['suma numerelor de factura', 'DECORATIVA', 'nu raspunde la nicio intrebare reala'])
    A(['media cotei TVA', 'DECORATIVA', 'constanta fiscala, nu indicator de decizie'])
    A([])
    r_sect4 = wm.max_row + 1
    A(['4) HANDOFF'])
    A(['input de la C09', 'model relational interogabil (relatii activate)'])
    A(['output C10', 'masuri definite (single source of truth, baza, reper, semnal)'])
    A(['predat catre C11', 'compara masurile (clasament/diferenta); C10 nu compara'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wm['A1'].font = TITLE
    for rr in (3, r_sect2, r_sect3, r_sect4):
        wm['A%d' % rr].font = SECT
    hdr(wm, 4)           # header tabel masuri
    hdr(wm, r_sect2 + 1) # header reper
    hdr(wm, r_sect3 + 1) # header cifra utila

    # --- 7 Masuri_Context (context-awareness, NU clasament) ---
    wc = out.create_sheet('Masuri_Context')
    B = wc.append
    B(['MASURI_CONTEXT C10 · aceeasi masura, taieturi diferite'])
    B([])
    B(['Masura demonstrata: Valoare medie per tranzactie (definitia din Masuri!M3).'])
    B(['Scop: dovada ca ACEEASI definitie da rezultat corect pe orice taietura.'])
    B(['NU este clasament. CARE taietura e mai mare se decide la C11, nu aici.'])
    B([])
    B(['taietura', 'formula (live)', 'rezultat'])
    r_ctx = wc.max_row
    B(['tot setul (overall)',
       '=ROUND(AVERAGE(Vanzari!J2:J%d),2)' % vlast, avg_tranz])
    B(['categorie = "%s"' % cat_demo,
       '=ROUND(AVERAGEIF(Vanzari!G2:G%d,"%s",Vanzari!J2:J%d),2)'
       % (vlast, cat_demo, vlast), avg_cat])
    B(['client = "%s"' % client_demo,
       '=ROUND(AVERAGEIF(Vanzari!C2:C%d,"%s",Vanzari!J2:J%d),2)'
       % (vlast, client_demo, vlast), avg_cli])
    B(['luna = %04d-%02d' % (an_demo, mo_demo),
       '=ROUND(AVERAGEIFS(Vanzari!J2:J%d,Vanzari!B2:B%d,">="&DATE(%d,%d,1),'
       'Vanzari!B2:B%d,"<"&EDATE(DATE(%d,%d,1),1)),2)'
       % (vlast, vlast, an_demo, mo_demo, vlast, an_demo, mo_demo), avg_mo])
    B([])
    B(['Citire: valorile difera pentru ca taietura difera, nu pentru ca masura e alta.'])
    B(['Definitia este identica in toate cele patru cazuri. Aceasta e o masura robusta.'])
    wc['A1'].font = TITLE
    hdr(wc, r_ctx)

    out.save(OUT)

    visible = [w.title for w in out.worksheets if w.sheet_state == 'visible']
    hidden = [w.title for w in out.worksheets if w.sheet_state != 'visible']
    print('SCRIS:', OUT)
    print('  foi vizibile (%d):' % len(visible), visible)
    print('  foi ascunse (%d):' % len(hidden), hidden)
    print('  suma:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  M2 tranzactii=%d | clienti=%d | zile active=%d' % (n_tranz, n_clienti, n_zile))
    print('  M3 medie/tranz=%.2f | M4 medie/client=%.2f | M5 medie/zi=%.2f'
          % (avg_tranz, avg_client, avg_zi))
    print('  reper=%.2f -> semnal=%s' % (prag_mediu, semnal_txt))
    print('  context: overall=%.2f | %s=%.2f | client=%.2f | luna=%.2f'
          % (avg_tranz, cat_demo, avg_cat, avg_cli, avg_mo))
    assert len(visible) <= 7, 'PREA MULTE FOI: %d' % len(visible)
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    print('  OK: <=7 foi, suma conservata')


if __name__ == '__main__':
    main()
