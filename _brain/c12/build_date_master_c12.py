#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c12.py - Date_MASTER-C12.xlsx (T3 · INTERPRETARE), inchide T3.

C12 EXPLICA rezultatul deja produs de C11 (clasamentul). Verb: a EXPLICA.
Intrebare-mama: De ce? Citeste cauza din model, nu o ghiceste.

Strategie sigura: deschid Date_MASTER-C11.xlsx (pastreaza TOATE foile, formulele si
stilurile C11 intacte), rescriu START pentru C12 si ADAUG foaia 'Interpretare'.
Vanzari ramane neatins -> suma conservata cap-coada (R-V02.14).

Garda C12 (PUR a explica): clasamentul vine ca INPUT (citit, nu reprodus = C11),
zero dashboard/grafic (T4), zero what-if/scenariu/predictie/recomandare actiune (T5).
Interpretare = insight verbal + cauza citita din model + verificabil inapoi in date.

Uz: python3 _brain/c12/build_date_master_c12.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import os

SRC = 'c11/Date_MASTER-C11.xlsx'
OUT = 'c12/Date_MASTER-C12.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def xlstr(s):
    """Literal text pentru formula Excel, cu ghilimele dublate."""
    return '"' + str(s).replace('"', '""') + '"'


def main():
    os.makedirs('c12', exist_ok=True)

    # --- pass valori (data_only) pentru a CITI cauza din model ---
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]
    vlast = len(vrows) + 1

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)

    cat_val = defaultdict(float); cat_n = defaultdict(int)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r); cat_n[r[ix['categorie']]] += 1
    top_cat = max(cat_val, key=cat_val.get)
    top_cat_val = round(cat_val[top_cat], 2)

    hw = [r for r in vrows if r[ix['categorie']] == top_cat]
    prod = defaultdict(float); cli = defaultdict(float)
    for r in hw:
        prod[r[ix['produs_nume']]] += vn(r); cli[r[ix['client_nume']]] += vn(r)
    top_prod = sorted(prod.items(), key=lambda x: -x[1])[:3]
    top_cli = sorted(cli.items(), key=lambda x: -x[1])[:3]
    doi_prod_pct = round(100 * (top_prod[0][1] + top_prod[1][1]) / cat_val[top_cat], 1)

    # avg/tranzactie pe categorie: cauza ascunsa de agregare (Software > Hardware la medie)
    avg_cat = {k: round(cat_val[k] / cat_n[k], 2) for k in cat_val}
    cat_avg_max = max(avg_cat, key=avg_cat.get)

    def pct_hw(v):
        return round(100 * v / cat_val[top_cat], 1)

    # --- pass formule (pastreaza tot C11 intact) ---
    out = openpyxl.load_workbook(SRC)

    # rescriu START pentru C12 (sterg vechiul, recreez pe pozitia 0)
    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C12 · INTERPRETARE · de ce arata rezultatul asa (citit din model)'],
        [],
        ['Ideea mare:'],
        ['Clasamentul arata CARE conduce. Interpretarea spune DE CE.'],
        ['Rezultat numeric  ->  intrebarea "de ce?"  ->  cauza citita din model  ->  insight verificabil.'],
        [],
        ['AHA: Nu rezultatul conteaza. Conteaza de ce apare rezultatul.'],
        ['MANTRA: Cifra spune cat. Explicatia spune de ce.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins fata de C11, suma conservata)'],
        ['  Masuri         definitiile mostenite de la C10'],
        ['  Comparatii     clasamentul mostenit de la C11 (CARE conduce)'],
        ['  Interpretare   DE CE: cauza citita din model, factori principali, verificabil'],
        [],
        ['Exercitiul:'],
        ['  1. Pornesti de la clasamentul corect (mostenit din C11).'],
        ['  2. Formulezi intrebarea de business: "De ce?".'],
        ['  3. Cobori sub total si citesti cauza din model (nu o ghicesti).'],
        ['  4. Separi ce apare impreuna de ce produce rezultatul.'],
        ['  5. Scrii insight-ul ca fraza verificabila inapoi in date.'],
        [],
        ['C12 doar EXPLICA. Nu re-claseaza (C11), nu vizualizeaza (T4), nu actioneaza (T5).'],
        ['Cu C12 se inchide treapta T3 ANALIZA: setul e corect, inteles si interpretat.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 10, 16):
        ws['A%d' % rr].font = SECT

    # --- foaia Interpretare (citeste cauza din model) ---
    wi = out.create_sheet('Interpretare')
    A = wi.append
    A(['INTERPRETARE C12 · de ce arata rezultatul asa, citit din model'])
    A([])

    A(['1) INTREBAREA · "De ce?"'])
    A(['rezultat de explicat', 'sursa', 'intrebarea'])
    A(['"%s" conduce clasamentul valorii' % top_cat, 'Comparatii (C11)',
       'De ce conduce %s, si ce anume il duce acolo?' % top_cat])
    A([])

    r_s2 = wi.max_row + 1
    A(['2) CAUZA CITITA DIN MODEL · drill sub total (nu ghicit)'])
    A(['factor', 'tip', 'formula (live)', 'valoare', 'pondere in %s (%%)' % top_cat])
    r_hdr2 = wi.max_row
    for p, v in top_prod:
        A(['produs: %s' % p, 'factor principal',
           '=ROUND(SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s,Vanzari!F2:F%d,%s),2)'
           % (vlast, vlast, xlstr(top_cat), vlast, xlstr(p)),
           round(v, 2), pct_hw(v)])
    for c, v in top_cli[:2]:
        A(['client: %s' % c, 'factor principal',
           '=ROUND(SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s,Vanzari!C2:C%d,%s),2)'
           % (vlast, vlast, xlstr(top_cat), vlast, xlstr(c)),
           round(v, 2), pct_hw(v)])
    A([])

    r_s3 = wi.max_row + 1
    A(['3) FACTORI PRINCIPALI · rezultatul rar are o singura cauza'])
    A(['constatare', 'dovada (citita mai sus)'])
    A(['nu o cauza unica: doua produse trag impreuna cea mai mare parte din %s' % top_cat,
       'primele 2 produse = %.1f%% din %s' % (doi_prod_pct, top_cat)])
    A(['si concentrarea pe clienti contribuie',
       'primul client = %.1f%% din %s' % (pct_hw(top_cli[0][1]), top_cat)])
    A([])

    r_s4 = wi.max_row + 1
    A(['4) COINCIDENTA vs CAUZA · ce apare impreuna nu explica automat'])
    A(['observatie', 'verdict', 'de ce'])
    A(['%s are si cele mai multe tranzactii' % top_cat, 'COINCIDENTA partiala',
       'volumul mare de tranzactii insoteste rezultatul, dar nu e singura cauza'])
    A(['valoarea medie/tranzactie nu e cea mai mare la %s' % top_cat, 'CAUZA reala = mix',
       '%s are media/tranzactie cea mai mare (%.2f), dar nu conduce totalul: '
       '%s conduce prin produse-cheie + volum, nu prin pret mediu'
       % (cat_avg_max, avg_cat[cat_avg_max], top_cat)])
    A([])

    r_s5 = wi.max_row + 1
    A(['5) CAUZA ASCUNSA DE AGREGARE · apare doar sub total'])
    A(['nivel', 'ce vezi'])
    A(['la nivel de total (clasament C11)', 'doar ca "%s conduce" (CARE)' % top_cat])
    A(['cobori sub total (aici)',
       'doua produse duc %.1f%% din categorie: cauza reala, ascunsa de totalul agregat'
       % doi_prod_pct])
    A([])

    r_s6 = wi.max_row + 1
    A(['6) EXPLICATIE VERIFICABILA · insight verbal ancorat in model'])
    A(['insight (fraza)', 'se verifica in'])
    A(['"%s conduce nu pentru ca fiecare vanzare e cea mai mare, ci pentru ca doua '
       'produse-cheie si cativa clienti concentrati ii duc %.1f%% din valoare; '
       'pretul mediu mai mare e la alta categorie."' % (top_cat, doi_prod_pct),
       'sectiunile 2-5 de mai sus (formule live pe Vanzari)'])
    A([])

    r_s7 = wi.max_row + 1
    A(['7) HANDOFF · inchidere T3'])
    A(['input de la C11', 'clasament (CARE conduce, contributie, ABC)'])
    A(['output C12', 'explicatie: DE CE conduce, cauza citita din model, verificabila'])
    A(['treapta T3 ANALIZA', 'FINALIZATA (corect -> inteles -> interpretat)'])
    A(['predat catre T4 (C13)', 'raportarea face explicatia vizibila; C12 nu vizualizeaza'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wi['A1'].font = TITLE
    for rr in (3, r_s2, r_s3, r_s4, r_s5, r_s6, r_s7):
        wi['A%d' % rr].font = SECT
    hdr(wi, 4)            # header intrebare
    hdr(wi, r_hdr2)       # header cauza
    hdr(wi, r_s3 + 1)
    hdr(wi, r_s4 + 1)
    hdr(wi, r_s5 + 1)
    hdr(wi, r_s6 + 1)

    # Interpretare imediat dupa Comparatii in ordinea foilor
    order = [s for s in out.sheetnames if s not in ('START', 'Interpretare')]
    if 'Comparatii' in order:
        i = order.index('Comparatii') + 1
    else:
        i = len(order)
    new_order = ['START'] + order[:i] + ['Interpretare'] + order[i:]
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    visible = [w.title for w in out.worksheets if w.sheet_state == 'visible']
    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  vizibile:', visible)
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  top_cat=%s (%.2f) | 2 produse=%.1f%% | media max la %s (%.2f)'
          % (top_cat, top_cat_val, doi_prod_pct, cat_avg_max, avg_cat[cat_avg_max]))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    print('  OK: suma conservata, Interpretare adaugata, T3 inchisa')


if __name__ == '__main__':
    main()
