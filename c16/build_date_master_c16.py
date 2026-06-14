#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c16.py - Date_MASTER-C16.xlsx (T4 · LIVRARE), inchide T4.

Nume fisier = Date_MASTER-C16.xlsx (cerut de gate B2).

C16 = LIVRAREA. Verb: a LIVRA. Axa: RAPORT DECISION-READY.
C16 = foaie-raport de decizie (raportul devine obiect de decizie), NU sinteza (C15),
NU layout (C14), NU sistem recurent (C17), NU logistica.

Strategie (re-chain post-audit T4): deschid Date_MASTER-C15.xlsx (pastreaza TOATE foile,
formulele si stilurile intacte), rescriu START pentru C16 si ADAUG foaia 'Livrare'.
Vanzari ramane neatins -> suma conservata cap-coada (R-V02.14).

NOTA chain (REPARAT post-audit T4): chainez din Date_MASTER-C15.xlsx, care contine foaia
'Sinteza' (C15). Lantul T4 e complet: Vizualizare (C13) -> Compunere (C14) -> Sinteza (C15)
-> Livrare (C16). Suma e conservata identic (T4 nu schimba Vanzari).

Garda C16: workbook = SUPORT pentru construirea foii-raport de decizie (cadru de decizie
peste agregatul real + ce intra in decizie vs anexa + verificare self-standing + 6 reguli).
ZERO sinteza de mesaj (=C15). ZERO layout/dashboard (=C14). ZERO sistem recurent (=C17).
Cifrele raman in Excel prin SUMIFS live (R-V02.15).

Uz: python3 c16/build_date_master_c16.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import os

SRC = 'c15/Date_MASTER-C15.xlsx'
OUT = 'c16/Date_MASTER-C16.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def xlstr(s):
    return '"' + str(s).replace('"', '""') + '"'


def main():
    os.makedirs('c16', exist_ok=True)

    # --- pass valori (data_only) pentru a CITI rezultatul de livrat ---
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]
    vlast = len(vrows) + 1

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)

    cat_val = defaultdict(float)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r)
    cats = sorted(cat_val.items(), key=lambda x: -x[1])
    top_cat, top_val = cats[0]
    second_cat, second_val = cats[1]
    pondere_top = round(100 * top_val / suma, 1)

    # --- pass formule (pastreaza tot C14 intact) ---
    out = openpyxl.load_workbook(SRC)

    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C16 · LIVRARE · raportul devine obiect de decizie, predabil decidentului'],
        [],
        ['Ideea mare:'],
        ['Analiza a dat raspunsul, C15 a sintetizat mesajul. Dar in raport e mut: decidentul'],
        ['tot nu decide. C16 da raportului FORMA finala de decizie: il face obiect care produce'],
        ['o hotarare. Nu livrez informatie. Livrez o decizie gata de luat.'],
        ['Mesaj sintetizat (C15)  ->  cadru de decizie  ->  foaie-raport de decizie  ->  predat.'],
        [],
        ['AHA: Raportul nu se trimite, se preda ca obiect de decizie.'],
        ['MANTRA: Nu livrez informatie. Livrez o decizie gata de luat.'],
        ['MOTTO: Raportul care decide.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins, suma conservata cap-coada)'],
        ['  Comparatii     clasamentul mostenit (rezultatul de raportat)'],
        ['  Interpretare   explicatia mostenita de la C12'],
        ['  Vizualizare    forma onesta mostenita de la C13'],
        ['  Compunere      organizarea paginii mostenita de la C14'],
        ['  Sinteza        mesajul esential mostenit de la C15'],
        ['  Livrare        SUPORT: cadru de decizie + ce intra/anexa + verificare + 6 reguli'],
        [],
        ['Exercitiul:'],
        ['  1. Iei mesajul esential (sintetizat anterior) si rezultatul corect.'],
        ['  2. Urci decizia ceruta in capul foii; detaliul brut coboara in anexa.'],
        ['  3. Scrii cadrul de decizie: decizie / risc / concluzie / pas urmator.'],
        ['  4. Verifici: poate decidentul hotari doar din foaie, fara tine langa?'],
        ['  5. Aplici cele 6 reguli ale foii-raport de decizie.'],
        [],
        ['C16 doar LIVREAZA raportul ca obiect de decizie. Nu sintetizeaza mesajul (C15),'],
        ['nu aranjeaza vizual pagina (C14), nu sistematizeaza recurent (C17), nu e logistica.'],
        ['C16 preda decizia gata; C17 o sistematizeaza in T5.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 13, 22):
        ws['A%d' % rr].font = SECT

    # --- foaia Livrare (suport pentru foaia-raport de decizie) ---
    wi = out.create_sheet('Livrare')
    A = wi.append
    A(['LIVRARE C16 · foaia-raport de decizie: raportul devine obiect de decizie'])
    A([])

    A(['1) DECIZIA CERUTA · ce se hotaraste din raport (citit din rezultat, nu inventat)'])
    A(['decizie ceruta', 'sursa', 'tip'])
    A(['Pe ce categorie concentram resursele in ciclul urmator',
       'Comparatii (C11/C12) + agregat pe categorie',
       'decizie de alocare, orientata spre actiune'])
    A([])

    r2 = wi.max_row + 1
    A(['2) CADRUL DE DECIZIE · decizie / risc / concluzie / pas urmator (ancorat in agregatul live)'])
    A(['element', 'continut (ancorat in date)'])
    r_hdr2 = wi.max_row
    A(['Concluzie',
       '=CONCATENATE("Categoria ",%s," conduce valoarea neta, cu ",'
       'TEXT(ROUND(100*SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s)/SUM(Vanzari!J2:J%d),1),"0.0"),'
       '"%%%% din total")' % (xlstr(top_cat), vlast, vlast, xlstr(top_cat), vlast)])
    A(['Decizie de luat',
       '=CONCATENATE("Mentinem focusul pe ",%s," sau realocam spre ",%s," (a doua)?")'
       % (xlstr(top_cat), xlstr(second_cat))])
    A(['Risc',
       '=CONCATENATE("Dependenta de o singura categorie: ",%s," tine ",'
       'TEXT(ROUND(100*SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s)/SUM(Vanzari!J2:J%d),1),"0.0"),'
       '"%%%% din valoarea neta; prag de alerta de stabilit")'
       % (xlstr(top_cat), vlast, vlast, xlstr(top_cat), vlast)])
    A(['Pas urmator',
       'Actiunea propusa (de scris explicit de decident), ancorata in agregatul de mai jos'])
    A([])

    r3 = wi.max_row + 1
    A(['3) CE INTRA IN DECIZIE (cap de foaie) vs CE RAMANE ANEXA'])
    A(['nivel', 'continut', 'unde sta'])
    A(['in decizie', 'agregatul pe categorie (relevant pentru hotarare)', 'capul foii'])
    A(['anexa', 'detaliul brut linie cu linie (nr_factura, data_factura, client)', 'jos / foaie separata'])
    A([])
    A(['agregat pe categorie (cifra live, referinta de decizie):'])
    A(['categorie', 'valoare neta (live)', 'pondere din total (%)'])
    r_hdr_ag = wi.max_row
    for c, v in cats:
        A([c,
           '=ROUND(SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s),2)' % (vlast, vlast, xlstr(c)),
           '=ROUND(100*SUMIFS(Vanzari!J2:J%d,Vanzari!G2:G%d,%s)/SUM(Vanzari!J2:J%d),1)'
           % (vlast, vlast, xlstr(c), vlast)])
    A([])

    r4 = wi.max_row + 1
    A(['4) VERIFICARE · poate decidentul hotari doar din foaie? (self-standing)'])
    A(['proba', 'rezultat'])
    A(['decizie + risc + concluzie + pas urmator prezente?', 'da, in cadrul de decizie (sectiunea 2)'])
    A(['cifra are unitate / perioada / referinta?', 'da, agregat live + sursa declarata, fara autor langa'])
    A(['se citeste o data si se decide?', 'da -> raportul e obiect de decizie, nu doar informatie'])
    A([])

    r5 = wi.max_row + 1
    A(['5) CELE 6 REGULI ALE FOII-RAPORT DE DECIZIE'])
    A(['#', 'capcana', 'regula decision-ready'])
    for n, (d, reg) in enumerate([
        ('Concluzia ingropata', 'concluzia ceruta sta in prima linie, nu la pagina 14'),
        ('Lipsa cadrului de decizie', 'decizie / risc / concluzie / pas urmator, explicit'),
        ('Totul aglomerat in cap', 'ce nu intra in decizie coboara in anexa, nu dispare'),
        ('Risc lasat implicit', 'riscul se scrie, cu pragul lui'),
        ('Raport care se termina in constatare', 'pasul urmator se scrie, ancorat in date'),
        ('Foaie care are nevoie de autor', 'fiecare cifra poarta unitate, perioada si referinta'),
    ], 1):
        A([n, d, reg])
    A([])

    r6 = wi.max_row + 1
    A(['6) HANDOFF · C16 preda decizia gata, C17 o sistematizeaza in T5'])
    A(['input de la C15', 'mesajul esential, sintetizat'])
    A(['output C16', 'raport decision-ready, predabil ca obiect de decizie'])
    A(['predat catre T5 (C17)', 'sistematizarea recurenta a raportului (nu e treaba C16)'])
    A(['C16 NU face', 'sinteza (C15), layout vizual (C14), sistem recurent (C17), logistica'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wi['A1'].font = TITLE
    for rr in (3, r2, r3, r4, r5, r6):
        wi['A%d' % rr].font = SECT
    hdr(wi, 4)
    hdr(wi, r_hdr2)
    hdr(wi, r_hdr_ag)
    hdr(wi, r4 + 1)
    hdr(wi, r5 + 1)

    # Livrare la final (dupa ultima foaie-suport T4 disponibila)
    order = [s for s in out.sheetnames if s not in ('START', 'Livrare')]
    new_order = ['START'] + order + ['Livrare']
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  top_cat=%s (%.2f, %.1f%%) | a doua=%s (%.2f)'
          % (top_cat, top_val, pondere_top, second_cat, second_val))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    assert 'Sinteza' in out.sheetnames, 'verificare: foaia Sinteza (C15) trebuie sa existe dupa re-chain'
    print('  OK: suma conservata, Livrare adaugata, lant T4 complet (Vizualizare->Compunere->Sinteza->Livrare)')


if __name__ == '__main__':
    main()
