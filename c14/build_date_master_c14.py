#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c14.py - Date_MASTER-C14.xlsx (T4 · COMPUNERE).

Nume fisier = Date_MASTER-C14.xlsx (cerut de gate B2).

C14 ASAZA SPATIAL mai multe obiecte vizuale oneste (produse la C13) intr-o pagina
de raport cu ierarhie. Verb: a COMPUNE. Axa: ORGANIZAREA SPATIALA A ELEMENTELOR VIZUALE.
C14 = pagina cu focar/ierarhie/traseu, NU obiectul singular (=C13), NU mesajul (=C15),
NU livrarea (=C16). Dashboard = substrat tehnic, niciodata identitate.

Strategie (continuare compozitionala a Date_MASTER-C13, mandat generare C14): deschid
Date_MASTER-C13.xlsx (pastreaza TOATE foile, formulele si stilurile), rescriu START
pentru C14 si ADAUG foaia 'Compunere'. Vanzari ramane neatins -> suma conservata (R-V02.14).
Pleaca de la obiectele C13, pastreaza logica de date, NU schimba tipuri de grafice,
NU inventeaza date, NU formuleaza concluzia, NU livreaza.

Uz: python3 c14/build_date_master_c14.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
import os

SRC = 'c13/Date_MASTER-C13.xlsx'
OUT = 'c14/Date_MASTER-C14.xlsx'
SUMA_ASTEPTATA = 7986019.38

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def xlstr(s):
    return '"' + str(s).replace('"', '""') + '"'


def main():
    os.makedirs('c14', exist_ok=True)

    # --- pass valori (data_only) pentru a CITI obiectele de compus ---
    wv = openpyxl.load_workbook(SRC, data_only=True)
    V = wv['Vanzari']
    vhdr = [c.value for c in next(V.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(vhdr)}
    vrows = [r for r in V.iter_rows(min_row=2, values_only=True)]
    vlast = len(vrows) + 1

    def vn(r):
        x = r[ix['valoare_neta']]
        return float(x) if x not in (None, '') else 0.0

    suma = round(sum(vn(r) for r in vrows), 2)

    cat_val = defaultdict(float)
    for r in vrows:
        cat_val[r[ix['categorie']]] += vn(r)
    cats = sorted(cat_val.items(), key=lambda x: -x[1])
    top_cat, top_val = cats[0]
    second_cat, second_val = cats[1]
    dif_pct = round(100 * (top_val - second_val) / top_val, 1)

    # --- pass formule (pastreaza tot C13 intact) ---
    out = openpyxl.load_workbook(SRC)

    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C14 · COMPUNERE · asezarea spatiala a mai multor obiecte vizuale intr-o pagina'],
        [],
        ['Ideea mare:'],
        ['Treapta de vizualizare a dat obiecte vizuale oneste, fiecare adevarat singur.'],
        ['Dar pe o pagina fara ordine, adevarul exista peste tot si nu se vede nicaieri.'],
        ['C14 asaza obiectele: focar, ierarhie, traseu de citire, grupare, spatiu.'],
        ['Aceleasi grafice, alta ordine, alta decizie.'],
        [],
        ['AHA: Aceleasi grafice, alta ordine, alta decizie.'],
        ['MANTRA: Compun privirea, nu pagina.'],
        ['MOTTO: Ce vede ochiul intai schimba decizia.'],
        [],
        ['Ce ai in acest fisier:'],
        ['  Vanzari        tabelul fact (neatins fata de C13, suma conservata)'],
        ['  Comparatii     clasamentul mostenit (obiect vizual de asezat)'],
        ['  Vizualizare    obiectele vizuale oneste mostenite de la treapta anterioara'],
        ['  Compunere      SUPORT: focar, ierarhie, traseu, grupare, spatiu + before/after'],
        [],
        ['Exercitiul:'],
        ['  1. Pleci de la obiectele vizuale oneste (mostenite), fara sa le redesenezi.'],
        ['  2. Stabilesti un singur focar: obiectul care poarta raspunsul, atins primul.'],
        ['  3. Ordonezi traseul de citire si retrogradezi ce nu muta decizia.'],
        ['  4. Faci ierarhia cu marime, pozitie, contrast, spatiu; grupezi ce e legat.'],
        ['  5. Verifici cu al doilea ochi: in 3 secunde se prinde aceeasi prioritate?'],
        [],
        ['C14 doar ASAZA. Nu deseneaza obiectul (C13), nu formuleaza mesajul (C15),'],
        ['nu livreaza (C16), nu re-analizeaza (T3). C14 face pagina; C15 ii extrage mesajul.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 13, 19):
        ws['A%d' % rr].font = SECT

    # --- foaia Compunere (suport pentru asezarea spatiala a obiectelor) ---
    wi = out.create_sheet('Compunere')
    A = wi.append
    A(['COMPUNERE C14 · asezarea spatiala a mai multor obiecte vizuale intr-o pagina de raport'])
    A([])

    A(['1) OBIECTELE VIZUALE PRIMITE · de la treapta de vizualizare (nu redesenate, tip neschimbat)'])
    A(['obiect', 'tip (neschimbat)', 'ce arata'])
    A(['Clasament categorii', 'bare cu axa de la zero', 'comparatia valorii intre categorii'])
    A(['Pondere din total', 'structura parte-din-tot', 'cat reprezinta fiecare categorie'])
    A(['Context masura', 'indicator (numar mare)', 'o singura cifra de referinta'])
    A([])

    r2 = wi.max_row + 1
    A(['2) INTREBAREA C14 · ce vede ochiul intai? · raspunsul guverneaza focarul'])
    A(['element', 'valoare'])
    A(['raspunsul venit din analiza (ce conteaza)',
       '%s conduce, cu %.1f%% peste %s' % (top_cat, dif_pct, second_cat)])
    A(['focarul (obiectul atins primul)', 'clasamentul categoriilor, cu %s in frunte' % top_cat])
    A([])

    r3 = wi.max_row + 1
    A(['3) ASEZAREA · pozitie, ordine de citire, grup (acelasi continut, alta dispunere)'])
    A(['obiect', 'pozitie', 'ordine de citire', 'grup'])
    A(['Clasament categorii', 'sus-stanga (focar)', '1', 'rezultatul principal'])
    A(['Pondere din total', 'langa focar', '2', 'rezultatul principal'])
    A(['Context masura', 'sub focar, mai mic', '3', 'suport'])
    A(['Tabel detaliu', 'jos, retrogradat', '4', 'detaliu la cerere'])
    A([])

    r4 = wi.max_row + 1
    A(['4) BEFORE / AFTER · aceleasi obiecte, alta asezare, alta decizie'])
    A(['dimensiune', 'inainte (ordinea productiei)', 'dupa (compunere guvernata de raspuns)'])
    A(['focar', 'niciunul, toate egale', 'unul singur, atins primul'])
    A(['ordine de citire', 'cronologica, intamplatoare', 'guvernata de raspuns'])
    A(['greutate vizuala', 'egala la lucruri inegale', 'ierarhie pe trei niveluri'])
    A(['spatiu gol', 'umplut (horror vacui)', 'deliberat, izoleaza focarul'])
    A(['decizie', 'intarzie, ochiul rataceste', 'se prinde in cateva secunde'])
    A([])

    r5 = wi.max_row + 1
    A(['5) CELE 6 OPERATII DE COMPUNERE (dezordine -> compozitie pentru decizie)'])
    A(['#', 'dezordine', 'operatie de compunere'])
    for n, (d, op) in enumerate([
        ('Pagina-zid fara focar', 'stabilesti un singur focar, atins primul'),
        ('Esentialul ingropat in colt', 'esentialul urca in primul punct de contact al ochiului'),
        ('Ordinea de productie pe pagina', 'ordinea de citire guvernata de raspuns'),
        ('Greutate egala la lucruri inegale', 'ierarhie prin marime, pozitie, contrast, spatiu'),
        ('Proximitate care minte despre relatii', 'grupare spatiala a obiectelor legate'),
        ('Horror vacui, umpli orice gol', 'spatiul gol folosit ca instrument de ierarhie'),
    ], 1):
        A([n, d, op])
    A([])

    r6 = wi.max_row + 1
    A(['6) HANDOFF · C14 asaza pagina, C15 ii extrage mesajul'])
    A(['input de la treapta de vizualizare', 'obiecte vizuale oneste, fiecare adevarat singur'])
    A(['output C14', 'pagina de raport coerenta: focar, ierarhie, traseu, grupare, spatiu'])
    A(['predat catre treapta de sintetizare', 'extragerea si formularea mesajului esential'])
    A(['C14 NU face', 'obiect nou, mesaj, livrare, re-analiza; nu schimba tipuri de grafice'])
    A(['suma conservata cap-coada', SUMA_ASTEPTATA])

    wi['A1'].font = TITLE
    for rr in (3, r2, r3, r4, r5, r6):
        wi['A%d' % rr].font = SECT
    hdr(wi, 4)
    hdr(wi, r2 + 1)
    hdr(wi, r3 + 1)
    hdr(wi, r4 + 1)
    hdr(wi, r5 + 1)

    # Compunere imediat dupa Vizualizare in ordinea foilor
    order = [s for s in out.sheetnames if s not in ('START', 'Compunere')]
    if 'Vizualizare' in order:
        i = order.index('Vizualizare') + 1
    else:
        i = len(order)
    new_order = ['START'] + order[:i] + ['Compunere'] + order[i:]
    out._sheets.sort(key=lambda s: new_order.index(s.title))

    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  suma Vanzari:', suma, '| delta:', round(suma - SUMA_ASTEPTATA, 2))
    print('  top_cat=%s (%.2f) | a doua=%s (%.2f) | dif=%.1f%%'
          % (top_cat, top_val, second_cat, second_val, dif_pct))
    assert abs(suma - SUMA_ASTEPTATA) < 0.01, 'SUMA NU se conserva'
    print('  OK: suma conservata, Compunere adaugata, continuare compozitionala C13')


if __name__ == '__main__':
    main()
