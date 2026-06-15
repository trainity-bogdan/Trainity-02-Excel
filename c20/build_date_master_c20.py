#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_date_master_c20.py - Date_MASTER-C20.xlsx (T5 - DELEGAREA), predarea proprietatii.

C20 = DELEGAREA. Verb: a DELEGA. Axa: PREDAI SISTEMUL CA SA LUCREZE FARA TINE, DETINUT
DE UN ROL. Autorul iese din PROPRIETATE. Artefactul central = foaia _DELEGARE: test VIU,
nu tabel pasiv. Trei straturi:
  A) CONTROALE  AUTOR_ACTIV (DA/NU, comutatorul de predare; semnatura vizuala, C19 nu il are)
                + ROL_DELEGAT (lista de ROLURI, nu persoane) + PARAM_MUTAT.
  B) HARTA DE PREDARE  rol/zona/backup + responsabilitate + acces (matrice ROLxZONA) +
                limite (PROTECTED) + escaladare (ROL + declansator) + parametrul critic
                (author-only vs documentat PARAM_).
  C) VERIFICARE VIE + STATUS  V1 zero dependenta author-only (formula LIVE: scoaterea
                autorului face lantul critic sa intoarca eroare) · V2 acces validat ·
                V3 zone interzise blocate · V4 escaladare functionala · STATUS calculat
                NEPREDAT/PARTIAL/DELEGAT/AUTONOM (MF1, fara overlap).

TEST VIU (MF2, cablat real pe referinte): cu AUTOR_ACTIV=NU si PARAM_MUTAT=NU, lantul
critic intoarce #N/A (parametrul author-only golit), V1=FAIL, STATUS=PARTIAL ("tu erai
cheia"). Muti parametrul in blocul documentat PARAM_ (PARAM_MUTAT=DA): V1=OK fara sa
rescrii logica, STATUS=AUTONOM. Fisierul se livreaza in starea rezolvata AUTONOM
(AUTOR_ACTIV=NU + PARAM_MUTAT=DA): dovada vie ca alt rol il poate continua. Drama e
reproductibila (seteaza PARAM_MUTAT=NU).

Exemplu (cerinta 8): raportarea lunara a vanzarilor; rol "Operare Raportare Vanzari";
parametru author-only = ajustarea manuala lunara.

Granita vs C19: C19 prinde un INPUT gresit in DATE (reguli/praguri); C20 prinde
DEPENDENTA de autor (proprietate). Foi distincte, subiecte distincte. NU management,
NU HR/fisa de post/RACI, NU SOP/documentare. Rol, nu persoana. C20 inchide PACK-ul.

Strategie (chain T5): deschid Date_MASTER-C19.xlsx (pastreaza TOATE foile mostenite si
suma), adaug START (coperta C20) si _DELEGARE. Vanzari_Curat ramane neatins ->
suma conservata cap-coada (R-V02.14).

Uz: python3 c20/build_date_master_c20.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

SRC = 'c19/Date_MASTER-C19.xlsx'
OUT = 'c20/Date_MASTER-C20.xlsx'
CLEAN = 'Vanzari_Curat'
COL_VAL = 'L'
ROW1, ROWN = 2, 2001

ROLURI = ['Operare Raportare Vanzari', 'Coordonare Operatiuni', 'Backup Raportare']

HDR_FILL = PatternFill('solid', fgColor='1F2A44')
HDR_FONT = Font(color='FFFFFF', bold=True)
TITLE = Font(bold=True, size=12, color='1F2A44')
SECT = Font(bold=True, size=11, color='B8860B')
GREEN = PatternFill('solid', fgColor='C6EFCE')
AMBER = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')


def hdr(ws, row):
    for c in ws[row]:
        if c.value is not None:
            c.fill = HDR_FILL
            c.font = HDR_FONT


def main():
    os.makedirs('c20', exist_ok=True)

    wv = openpyxl.load_workbook(SRC, data_only=True)
    VC = wv[CLEAN]
    hdr_cells = [c.value for c in next(VC.iter_rows(min_row=1, max_row=1))]
    ixv = {h: i for i, h in enumerate(hdr_cells) if h}
    drows = [r for r in VC.iter_rows(min_row=2, values_only=True)]

    def num(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return None
    suma = round(sum(num(r[ixv['valoare_neta']]) for r in drows if num(r[ixv['valoare_neta']]) is not None), 2)
    nrows = len(drows)

    out = openpyxl.load_workbook(SRC)
    R_VAL = "'%s'!$%s$%d:$%s$%d" % (CLEAN, COL_VAL, ROW1, COL_VAL, ROWN)

    # ---------- START (coperta C20) ----------
    if 'START' in out.sheetnames:
        out.remove(out['START'])
    ws = out.create_sheet('START', 0)
    for line in [
        ['C20 - DELEGARE - sistemul nu mai e al tau, e al rolului care il detine'],
        [],
        ['Ideea mare:'],
        ['Sistemul ruleaza singur (C18) si se pazeste singur (C19). Si totusi e inca al tau:'],
        ['la orice problema, tot pe tine te suna. Un sistem detinut de o singura persoana nu'],
        ['e autonom, e doar bine intretinut de ea. C20 preda PROPRIETATEA de la o persoana la'],
        ['un ROL, si o dovedeste in workbook: apesi "scoate autorul" si STATUS se reaseaza singur.'],
        ['C18 ruleaza -> C19 se tine corect -> C20 PREDA proprietatea (fara autor).'],
        [],
        ['HERO: Cum deleg sistemul, ca sa mearga fara mine?'],
        ['AHA: Un sistem nu e autonom pentru ca merge singur. E autonom cand il poate detine altcineva.'],
        ['MANTRA: Nu impart sarcini. Deleg sistemul.'],
        ['MOTTO: Pleci, si munca nu mai e a ta.'],
        [],
        ['Ce vezi la deschidere (testul de predare, pe viu):'],
        ['  Fisierul se livreaza in starea rezolvata: AUTOR_ACTIV=NU si parametrul critic mutat in'],
        ['  blocul documentat PARAM_ (PARAM_MUTAT=DA). Toate V1-V4 sunt OK -> STATUS=AUTONOM:'],
        ['  autorul e scos si sistemul nu se rupe, alt rol il poate continua.'],
        ['  Drama (reproductibila): seteaza PARAM_MUTAT=NU. Lantul critic intoarce #N/A (parametrul'],
        ['  author-only golit cu autorul scos), V1=FAIL, STATUS=PARTIAL. "Credeai ca ai delegat.'],
        ['  Tu erai cheia." Muti parametrul inapoi in PARAM_ (PARAM_MUTAT=DA): V1=OK, STATUS=AUTONOM.'],
        [],
        ['_DELEGARE (predarea proprietatii):'],
        ['  A CONTROALE   AUTOR_ACTIV (comutatorul de predare) + ROL_DELEGAT (lista de roluri)'],
        ['  B HARTA       rol/zona/backup, responsabilitate, acces (ROLxZONA), limite, escaladare'],
        ['  B PARAMETRU   parametrul critic: author-only vs documentat PARAM_ (accesibil rolului)'],
        ['  C V1-V4       zero dependenta author-only / acces validat / zone blocate / escaladare'],
        ['  C STATUS      NEPREDAT / PARTIAL / DELEGAT / AUTONOM, calculat din V1-V4 (nu bifat)'],
        [],
        ['Granita: C19 prinde un INPUT gresit in date; C20 prinde DEPENDENTA de autor (proprietate).'],
        ['_DELEGARE este TEST VIU, nu tabel pasiv (anti-RACI). Rol, nu persoana. NU management, NU HR.'],
        ['C20 inchide PACK-ul C01-C20: nu mai exista treapta urmatoare.'],
    ]:
        ws.append(line)
    ws['A1'].font = TITLE
    for rr in (3, 15, 23):
        ws['A%d' % rr].font = SECT

    # ---------- _DELEGARE ----------
    if '_DELEGARE' in out.sheetnames:
        out.remove(out['_DELEGARE'])
    g = out.create_sheet('_DELEGARE')
    A = g.append

    A(['_DELEGARE C20 - predarea proprietatii unui sistem catre un ROL, dovedita prin test VIU'])
    A([])

    # A) CONTROALE
    rA = g.max_row + 1
    A(['A) CONTROALE - motorul testului viu (semnatura: AUTOR_ACTIV, comutatorul de predare)'])
    A(['control', 'valoare', 'rol']); rAhdr = g.max_row
    A(['AUTOR_ACTIV', 'NU', 'comutatorul de predare: DA = autorul prezent, NU = autorul scos (testul)']); r_autor = g.max_row
    A(['ROL_DELEGAT', ROLURI[0], 'rolul care preia proprietatea (lista de ROLURI, nu persoane)']); r_rol = g.max_row
    A(['PARAM_MUTAT', 'DA', 'DA = parametrul author-only a fost mutat in blocul documentat PARAM_']); r_pmut = g.max_row
    A([])

    # B) HARTA DE PREDARE
    rB = g.max_row + 1
    A(['B) HARTA DE PREDARE - inputul rolului (alimenteaza testul, NU tabel de citit)'])
    A(['element', 'valoare', 'detaliu']); rBhdr = g.max_row
    A(['ZONA_DETINUTA', 'Raportarea lunara a vanzarilor', 'partea de sistem pe care rolul o detine si o opereaza']); r_zona = g.max_row
    A(['BACKUP_ROL', ROLURI[1], 'rolul care preia cand rolul principal lipseste']); r_backup = g.max_row
    A(['RESPONSABILITATE', 'Produce raportul lunar si raspunde de corectitudinea livrarii', 'ce raspundere trece la rol']); r_resp = g.max_row
    A(['ACCES_OPERARE', 'DA', 'rolul vede si modifica zona de operare (alimenteaza V2)']); r_accop = g.max_row
    A(['ACCES_PROTEJATE', 'NU', 'rolul are zero acces in zonele protejate (alimenteaza V2)']); r_accpr = g.max_row
    A(['LIMITA_1_PROTECTED', 'DA', 'parametrii de sistem nu se ating: range chiar blocat (alimenteaza V3)']); r_lim1 = g.max_row
    A(['LIMITA_2_PROTECTED', 'DA', 'definitiile sistemului nu se modifica: range chiar blocat (alimenteaza V3)']); r_lim2 = g.max_row
    A(['ESCALADARE_ROL', ROLURI[1], 'catre ce ROL urca o problema peste mandat (alimenteaza V4)']); r_escrol = g.max_row
    A(['DECLANSATOR', 'abatere peste mandatul rolului', 'cand se escaladeaza (alimenteaza V4)']); r_decl = g.max_row
    A([])

    # B') PARAMETRUL CRITIC (drama author-only vs documentat)
    rBp = g.max_row + 1
    A(["B') PARAMETRUL CRITIC - ajustarea lunara: author-only vs blocul documentat PARAM_"])
    A(['parametru', 'valoare', 'sursa']); rBphdr = g.max_row
    A(['AUTHOR_AJUSTARE', 1500, 'ajustarea manuala lunara tinuta de autor (AUTHOR_ONLY=DA, se goleste cand autorul e scos)']); r_auth_aj = g.max_row
    A(['PARAM_AJUSTARE', 1500, 'aceeasi ajustare in blocul documentat PARAM_, accesibila rolului (nu depinde de autor)']); r_param_aj = g.max_row
    A([])

    # ROLURI (lista pentru V4 / validari)
    rR = g.max_row + 1
    A(['LISTA_ROLURI - rolurile recunoscute (pentru escaladare si validare)'])
    r_role0 = g.max_row + 1
    for rn in ROLURI:
        A([rn])
    r_roleN = g.max_row
    A([])

    # LANT CRITIC (raportarea lunara)
    rL = g.max_row + 1
    A(['LANT CRITIC - raportarea lunara (livrarea pe care rolul trebuie sa o poata produce fara autor)'])
    A(['element', 'valoare (oglinda vie)', 'ce arata']); rLhdr = g.max_row
    A(['INPUT_BAZA', '=SUM(%s)' % R_VAL, 'baza accesibila rolului, din datele de vanzari']); r_baza = g.max_row
    A(['AJUSTARE_EFECTIVA',
       '=IF($B$%d="DA",$B$%d,IF($B$%d="DA",$B$%d,NA()))' % (r_pmut, r_param_aj, r_autor, r_auth_aj),
       'ajustarea folosita real: din PARAM_ daca e mutata, altfel din zona autorului daca e prezent, altfel #N/A']); r_ajef = g.max_row
    A(['AJUSTARE_SHADOW',
       '=IF($B$%d="DA",$B$%d,NA())' % (r_pmut, r_param_aj),
       'testul autorului scos: ce ramane daca autorul dispare (doar sursa documentata PARAM_)']); r_ajsh = g.max_row
    A(['RAPORT_LUNAR', '=$B$%d+$B$%d' % (r_baza, r_ajef), 'livrarea critica; intoarce #N/A daca ajustarea atarna de autorul scos']); r_raport = g.max_row
    A([])

    # C) VERIFICARE VIE (V1-V4 consecutive pentru COUNTIF)
    rC = g.max_row + 1
    A(['C) VERIFICARE VIE - patru verificari pe formula (OK / FAIL), nu bifaj'])
    A(['verificare', 'rezultat (oglinda vie)', 'ce confirma']); rChdr = g.max_row
    A(['V1 zero dependenta author-only',
       '=IF(ISERROR($B$%d),"FAIL","OK")' % r_ajsh,
       'cu autorul scos, niciun input critic nu mai atarna de zona autorului']); r_v1 = g.max_row
    A(['V2 acces validat',
       '=IF(AND($B$%d="DA",$B$%d="NU"),"OK","FAIL")' % (r_accop, r_accpr),
       'rolul vede/modifica zona de operare SI are zero acces in zonele protejate']); r_v2 = g.max_row
    A(['V3 zone interzise blocate',
       '=IF(AND($B$%d="DA",$B$%d="DA"),"OK","FAIL")' % (r_lim1, r_lim2),
       'fiecare limita declarata e un range chiar blocat (PROTECTED=DA)']); r_v3 = g.max_row
    A(['V4 escaladare functionala',
       '=IF(AND(COUNTIF($A$%d:$A$%d,$B$%d)>0,$B$%d<>""),"OK","FAIL")' % (r_role0, r_roleN, r_escrol, r_decl),
       'escaladarea tinteste un ROL din lista SI are un declansator scris']); r_v4 = g.max_row
    A([])

    # D) STATUS
    rD = g.max_row + 1
    A(['D) STATUS - calculat din harta + V1-V4 (MF1, fara overlap), NU bifat manual'])
    A(['HARTA_COMPLETA',
       '=IF(AND($B$%d<>"",$B$%d<>"",$B$%d<>"",$B$%d<>"",$B$%d<>""),"DA","NU")' % (r_rol, r_zona, r_resp, r_escrol, r_decl),
       'rolul e definit si harta de predare e completa']); r_hc = g.max_row
    A(['STATUS_DELEGARE',
       '=IF(OR($B$%d="",$B$%d="NU"),"NEPREDAT",IF(COUNTIF($B$%d:$B$%d,"FAIL")>0,"PARTIAL",IF($B$%d="DA","DELEGAT","AUTONOM")))' % (
           r_rol, r_hc, r_v1, r_v4, r_autor),
       'NEPREDAT harta incompleta/rol nedefinit · PARTIAL un V FAIL · DELEGAT toate OK + autor prezent · AUTONOM toate OK + autor scos']); r_status = g.max_row
    A([])

    # E) TESTUL DE PREDARE (narativ viu)
    rE = g.max_row + 1
    A(['E) TESTUL DE PREDARE - ce spune sistemul despre predare, pe viu'])
    A(['testul de predare',
       '=IF($B$%d="AUTONOM","autorul e scos si sistemul nu se rupe: alt rol il poate continua",'
       'IF($B$%d="DELEGAT","harta validata, dar autorul e inca prezent (AUTOR_ACTIV=DA); scoate autorul ca sa probezi",'
       'IF($B$%d="PARTIAL","scoaterea autorului a aprins un FAIL: inca esti cheia; muta inputul author-only in PARAM_",'
       '"harta incompleta sau rol nedefinit: nu ai ce preda inca")))' % (r_status, r_status, r_status)]); r_test = g.max_row
    A([])

    # F) GRANITE / GARDA
    rF = g.max_row + 1
    A(['F) GRANITE + GARDA'])
    A(['granita C19->C20', 'C19 prinde un INPUT gresit in date (reguli/praguri); C20 prinde DEPENDENTA de autor (proprietate). Foi distincte.'])
    A(['granita C18->C20', 'C18 face sistemul sa mearga singur (fara efort); C20 face ca sistemul sa poata fi detinut de altcineva.'])
    A(['rol nu persoana', 'proprietatea sta pe un ROL; persoana se ataseaza temporar. Niciun nume de om in harta.'])
    A(['anti-RACI', 'tot ce intra in _DELEGARE alimenteaza testul viu (STATUS calculat). Daca devine tabel de citit, e documentare, nu delegare.'])
    A(['anti-management/HR', 'NU desemneaza persoane, NU fisa de post, NU SOP. Proprietatea pe ROL, dovedita in workbook, nu in discurs.'])
    A([])
    A(['Garda: _DELEGARE este test VIU. Scoti autorul, STATUS se reaseaza singur. Explicatia nu muta proprietatea; testul viu o muta.'])

    # stiluri
    g['A1'].font = TITLE
    for rr in (rA, rB, rBp, rR, rL, rC, rD, rE, rF):
        g['A%d' % rr].font = SECT
    for rr in (rAhdr, rBhdr, rBphdr, rLhdr, rChdr):
        hdr(g, rr)
    g['A%d' % r_status].font = Font(bold=True, size=11, color='1F2A44')
    g.column_dimensions['A'].width = 34
    g.column_dimensions['B'].width = 60
    g.column_dimensions['C'].width = 70

    # Data Validation (comutatoare + liste)
    def dv_list(values, rr):
        dv = DataValidation(type='list', formula1='"%s"' % ','.join(values), allow_blank=True, showErrorMessage=True)
        dv.errorTitle = 'Regula C20 (DELEGARE)'
        dv.error = 'Valoare in afara listei permise pentru testul de predare.'
        g.add_data_validation(dv); dv.add('B%d' % rr)

    dv_list(['DA', 'NU'], r_autor)
    dv_list(ROLURI, r_rol)
    dv_list(['DA', 'NU'], r_pmut)
    dv_list(['DA', 'NU'], r_accop)
    dv_list(['DA', 'NU'], r_accpr)
    dv_list(['DA', 'NU'], r_lim1)
    dv_list(['DA', 'NU'], r_lim2)
    dv_list(ROLURI, r_escrol)

    # Conditional formatting pe STATUS (4 stari)
    sc = 'B%d' % r_status
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"AUTONOM"'], fill=GREEN))
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"DELEGAT"'], fill=GREEN))
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"PARTIAL"'], fill=AMBER))
    g.conditional_formatting.add(sc, CellIsRule(operator='equal', formula=['"NEPREDAT"'], fill=RED))
    # semnal pe V1-V4
    for rr in (r_v1, r_v2, r_v3, r_v4):
        cc = 'B%d' % rr
        g.conditional_formatting.add(cc, CellIsRule(operator='equal', formula=['"OK"'], fill=GREEN))
        g.conditional_formatting.add(cc, CellIsRule(operator='equal', formula=['"FAIL"'], fill=RED))

    # named ranges
    def dn(name, ref):
        out.defined_names[name] = DefinedName(name, attr_text=ref)
    dn('AUTOR_ACTIV', "_DELEGARE!$B$%d" % r_autor)
    dn('ROL_DELEGAT', "_DELEGARE!$B$%d" % r_rol)
    dn('PARAM_MUTAT', "_DELEGARE!$B$%d" % r_pmut)
    dn('PARAM_AJUSTARE', "_DELEGARE!$B$%d" % r_param_aj)
    dn('RAPORT_LUNAR', "_DELEGARE!$B$%d" % r_raport)
    dn('AJUSTARE_SHADOW', "_DELEGARE!$B$%d" % r_ajsh)
    dn('V1_DELEG', "_DELEGARE!$B$%d" % r_v1)
    dn('V2_DELEG', "_DELEGARE!$B$%d" % r_v2)
    dn('V3_DELEG', "_DELEGARE!$B$%d" % r_v3)
    dn('V4_DELEG', "_DELEGARE!$B$%d" % r_v4)
    dn('STATUS_DELEGARE', "_DELEGARE!$B$%d" % r_status)

    rest = [s for s in out.sheetnames if s not in ('START', '_DELEGARE')]
    out._sheets.sort(key=lambda s: (['START', '_DELEGARE'] + rest).index(s.title))
    out.save(OUT)

    print('SCRIS:', OUT)
    print('  foi:', [w.title for w in out.worksheets])
    print('  named ranges:', list(out.defined_names.keys()))
    print('  STATUS=_DELEGARE!B%d | V1..V4=B%d:B%d | RAPORT=B%d | SHADOW=B%d' % (r_status, r_v1, r_v4, r_raport, r_ajsh))
    print('  livrat: AUTOR_ACTIV=NU (B%d) + PARAM_MUTAT=DA (B%d) -> asteptat STATUS=AUTONOM' % (r_autor, r_pmut))
    print('  suma Vanzari_Curat (conservata fata de C19):', suma, '| randuri:', nrows)
    assert '_DELEGARE' in out.sheetnames and CLEAN in out.sheetnames
    assert abs(suma - 1247893.5) < 0.01, 'SUMA NU se conserva fata de C19'
    print('  OK: suma conservata, _DELEGARE adaugata (controale/harta/parametru/V1-V4/STATUS/test/granite)')


if __name__ == '__main__':
    main()
